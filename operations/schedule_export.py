"""operations/schedule_export.py — الجدولُ ورقةَ Excel.

الورقةُ هنا صورةُ المطبوعة لا مستودعُ بيانات: الشكلُ نفسه — القسمُ ثمّ المعلّم
ثمّ خمسةٌ وثلاثون خانةً ثمّ النصاب — كي يجد من فتح الملفّ ما رآه في الورق. ومن
أراد البيانات خاماً فله واجهةُ الإسناد لا هذا الملفّ.

والأنماطُ من `reports.services.ExcelService`: هي أنماطُ المنصّة كلّها — رأسٌ من
ثلاثة سطورٍ وشعارٌ وحمايةٌ وإعدادُ طباعة — وبناءُ نسخةٍ ثانيةٍ منها هنا يفرّق
ملفّات المدرسة على شكلين. والاستيرادُ داخل الدوال لأنّ `reports` تستورد
`operations.models`، فاستيرادٌ في رأس الملفّ يعقد الحلقة.
"""

from __future__ import annotations

#: عرضُ خانة الحصّة: رمزُ الشعبة أربعةُ محارف («11/2») لا أكثر.
_CELL_WIDTH = 4.6

#: شريطُ القسم: الأقسامُ المتجاورة تتناوب على لونين فاتحين — فالحدُّ بينها
#: يُرى دون أن تُنسخ لوحةُ ألوان الورقة المطبوعة في موضعٍ ثانٍ تشيخ فيه.
_BAND_FILL = "F4F1F2"

#: خطُّ المنصّة — هو خطُّ الشاشة والورقة، فليكن خطَّ الملفّ. وأنماطُ
#: `ExcelService` مكتوبةٌ بـArial، فتُمرّ الورقةُ بعد بنائها ويُبدَّل الاسمُ
#: وحدَه: الوزنُ واللونُ والحجمُ كما ضُبطت. ومن لم يكن الخطُّ على جهازه
#: أبدله Excel بأقرب موجود — ولا تسقط الورقة.
_FONT = "Tajawal"


def schedule_workbook(ctx: dict):
    """مُصنَّفُ Excel للورقة المعروضة — الجدولُ العام مصفوفةً، وما سواه شبكة."""
    if ctx.get("view_type") == "all_teachers":
        workbook = _matrix_workbook(ctx)
    else:
        workbook = _grid_workbook(ctx)
    _apply_font(workbook.active)
    return workbook


def _apply_font(ws) -> None:
    """خطُّ المنصّة على كلّ خانةٍ مكتوبة — بقيّةُ النمط كما هي."""
    from copy import copy

    for row in ws.iter_rows():
        for cell in row:
            if cell.font is not None and cell.font.name != _FONT:
                font = copy(cell.font)
                font.name = _FONT
                cell.font = font


def _sheet(title: str, report_title: str, ctx: dict, num_cols: int):
    """ورقةٌ برأس المنصّة الثلاثيّ — وتُعاد مع أنماطها لمن يملؤها."""
    from reports.services import ExcelService

    wb, ws, styles = ExcelService._make_workbook(title)
    school = ctx.get("school")
    ExcelService._add_professional_header(
        ws,
        school.name if school else "SchoolOS",
        report_title,
        ctx.get("year") or "",
        num_cols,
    )
    return wb, ws, styles


def _matrix_workbook(ctx: dict):
    """الجدول العام: سطرٌ لكلّ معلّم، والأسبوعُ خمسةٌ وثلاثون عموداً.

    والرأسُ سطران — الأيامُ مدموجةً فوق أرقام الحصص — كما في الورقة، فلا
    يقرأ أحدٌ رقم «٣» دون أن يعرف يومه.
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from reports.services import ExcelService

    matrix = ctx.get("matrix") or []
    totals = ctx.get("matrix_totals")
    days = ctx.get("days") or []
    periods = list(ctx.get("period_numbers") or range(1, 8))

    first_period_col = 3
    num_cols = 2 + len(days) * len(periods) + 1
    wb, ws, styles = _sheet(
        "الجدول العام", ctx.get("title") or "الجدول العام للمعلمين", ctx, num_cols
    )

    # ── الرأس: سطرُ الأيام فوق سطر الحصص ──
    for column, header in ((1, "القسم"), (2, "المعلّم"), (num_cols, "النصاب")):
        letter = get_column_letter(column)
        ws.merge_cells(f"{letter}4:{letter}5")
        _head_cell(ws.cell(row=4, column=column, value=header), styles)

    column = first_period_col
    for _, day_name in days:
        span = len(periods)
        ws.merge_cells(start_row=4, start_column=column, end_row=4, end_column=column + span - 1)
        _head_cell(ws.cell(row=4, column=column, value=day_name), styles)
        for offset, number in enumerate(periods):
            _head_cell(ws.cell(row=5, column=column + offset, value=number), styles)
        column += span

    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 18

    # ── السطور: خانةُ القسم ممتدّةٌ على معلّميه، كما في الورقة ──
    band = False
    for index, row_data in enumerate(matrix):
        row_num = 6 + index
        span = row_data.get("dept_span") or 0
        if span:
            band = not band
            # قسمٌ بمعلّمٍ واحدٍ لا يُدمج: دمجُ خانةٍ بنفسها مدىً فارغُ المعنى
            # يبقى في الملفّ.
            if span > 1:
                ws.merge_cells(
                    start_row=row_num, start_column=1, end_row=row_num + span - 1, end_column=1
                )
            ws.cell(row=row_num, column=1, value=row_data["department"]["name"])
        teacher = row_data.get("teacher")
        name_cell = ws.cell(row=row_num, column=2, value=getattr(teacher, "full_name", "") or "")
        name_cell.alignment = Alignment(horizontal="right", vertical="center")

        for day_index, day in enumerate(row_data.get("days") or []):
            for period_index, cell in enumerate(day):
                ws.cell(
                    row=row_num,
                    column=first_period_col + day_index * len(periods) + period_index,
                    value=" ".join(slot.class_group.short_code for slot in cell),
                )
        ws.cell(row=row_num, column=num_cols, value=row_data.get("total") or 0)

        ExcelService._style_data_row(ws, styles, row_num, num_cols, False)
        ws.row_dimensions[row_num].height = 17
        if band:
            fill = PatternFill("solid", fgColor=_BAND_FILL)
            for column in range(1, num_cols + 1):
                ws.cell(row=row_num, column=column).fill = fill
        # المحاذاةُ تُعاد بعد الأنماط: `_style_data_row` يوسّط كلّ خانة،
        # واسمُ المعلّم يُقرأ من يمينه.
        name_cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row_num, column=1).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    # ── سطرُ المجموع ──
    last_row = 6 + len(matrix)
    if totals:
        ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=2)
        ws.cell(row=last_row, column=1, value="مجموع الحصص")
        short_columns = set()
        for day_index, day in enumerate(totals.get("days") or []):
            for period_index, column_totals in enumerate(day):
                column = first_period_col + day_index * len(periods) + period_index
                ws.cell(row=last_row, column=column, value=column_totals["count"])
                if column_totals.get("short"):
                    short_columns.add(column)
        ws.cell(row=last_row, column=num_cols, value=totals.get("total") or 0)
        ExcelService._style_data_row(ws, styles, last_row, num_cols, False)
        # سطرُ المجموع عريضٌ كلُّه، وعمودٌ فيه شعبةٌ بلا درسٍ أحمر — كالورقة
        # المطبوعة: العلامةُ هناك لونٌ وخطٌّ تحته، وهنا لونٌ وعرضٌ في خانةٍ
        # لا تُقرأ إلّا على الشاشة.
        for column in range(1, num_cols + 1):
            ws.cell(row=last_row, column=column).font = Font(
                name="Arial", bold=True, color="B00020" if column in short_columns else "000000"
            )

    # ── القياسات والطباعة ──
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 26
    for column in range(first_period_col, num_cols):
        ws.column_dimensions[get_column_letter(column)].width = _CELL_WIDTH
    ws.column_dimensions[get_column_letter(num_cols)].width = 8
    # التجميدُ عند C6: القسمُ والمعلّمُ يبقيان مع التمرير عرضاً، والرأسُ طولاً.
    ws.freeze_panes = "C6"

    ExcelService._apply_protection(ws, num_cols)
    ExcelService._setup_print(ws, num_cols, len(matrix) + 2, paper="a3", orientation="landscape")
    ws.print_title_rows = "1:5"
    ws.print_area = f"A1:{get_column_letter(num_cols)}{last_row}"
    return wb


def _grid_workbook(ctx: dict):
    """جدولُ معلّمٍ أو شعبةٍ أو المدرسة: الحصصُ سطوراً والأيامُ أعمدة."""
    from openpyxl.styles import Alignment

    from reports.services import ExcelService

    grid = ctx.get("grid") or {}
    days = ctx.get("days") or []
    periods = ctx.get("periods") or []
    view_type = ctx.get("view_type")

    num_cols = 1 + len(days)
    wb, ws, styles = _sheet("الجدول", ctx.get("title") or "الجدول الدراسي", ctx, num_cols)

    columns = [("الحصة", 12)] + [(day_name, 26) for _, day_name in days]
    ExcelService._add_header_row(ws, styles, 4, columns)

    for index, period in enumerate(periods):
        row_num = 5 + index
        # رقمُ الحصّة وحدَه في العمود — وتوقيتُها في خانتها، كالورقة.
        ws.cell(row=row_num, column=1, value=period["number"])
        for day_index, (day_num, _) in enumerate(days):
            cell_slots = (grid.get(day_num) or {}).get(period["number"]) or []
            ws.cell(
                row=row_num,
                column=2 + day_index,
                value="\n".join(_slot_text(slot, view_type) for slot in cell_slots) or "—",
            )
        ExcelService._style_data_row(ws, styles, row_num, num_cols, index % 2 == 1)
        ws.row_dimensions[row_num].height = 46

    for column in range(1, num_cols + 1):
        ws.cell(row=4, column=column).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    ws.freeze_panes = "B5"
    ExcelService._apply_protection(ws, num_cols)
    ExcelService._setup_print(ws, num_cols, len(periods), paper="a4", orientation="landscape")
    return wb


def _slot_text(slot, view_type: str | None) -> str:
    """نصُّ الخانة: المادّةُ، ثمّ من لا يُعرف من العنوان، ثمّ توقيتُ الحصّة.

    والتوقيتُ في الخانة لا في عمود الحصص: خانتان في العمود الواحد قد تختلف
    ساعتاهما (طابقان بجرسين)، فتوقيتُ العمود يكذب على إحداهما.
    """
    parts = [slot.subject.name_ar if slot.subject else "—"]
    if view_type != "teacher" and slot.teacher_id:
        parts.append(slot.teacher.full_name)
    if view_type != "class":
        parts.append(str(slot.class_group))
    if slot.start_time and slot.end_time:
        parts.append(f"{slot.start_time:%H:%M} – {slot.end_time:%H:%M}")
    return " — ".join(part for part in parts if part)


def _head_cell(cell, styles: dict):
    """خانةُ رأسٍ بأنماط المنصّة — كستنائيّةٌ بيضاءُ الخطّ موسَّطة."""
    cell.font = styles["header_font"]
    cell.fill = styles["header_fill"]
    cell.alignment = styles["header_align"]
    cell.border = styles["thin_border"]
    return cell
