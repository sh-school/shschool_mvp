"""
staging/views.py
استيراد الدرجات من Excel وتحميل قالب الإدخال
"""

import io
import logging

from django.conf import settings

logger = logging.getLogger(__name__)
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from assessments.models import Assessment
from assessments.services import GradeService
from core.models import CustomUser, StudentEnrollment
from core.permissions import role_required

from .models import ImportLog

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ── صفحة الاستيراد الرئيسية ────────────────────────────────


@login_required
@role_required(
    "principal",
    "vice_academic",
    "vice_admin",
    "coordinator",
    "teacher",
    "ese_teacher",
    "admin",
    "secretary",
)
def import_grades_select(request):
    """اختيار التقييم المراد استيراد درجاته"""
    school = request.user.get_school()
    year = request.GET.get("year", settings.CURRENT_ACADEMIC_YEAR)

    if request.user.is_admin():
        assessments = (
            Assessment.objects.filter(
                school=school,
                package__setup__academic_year=year,
                status__in=["published", "graded"],
            )
            .select_related(
                "package__setup__subject",
                "package__setup__class_group",
            )
            .order_by(
                "package__setup__class_group__grade",
                "package__setup__subject__name_ar",
            )
        )
    else:
        assessments = (
            Assessment.objects.filter(
                school=school,
                package__setup__teacher=request.user,
                package__setup__academic_year=year,
                status__in=["published", "graded"],
            )
            .select_related(
                "package__setup__subject",
                "package__setup__class_group",
            )
            .order_by(
                "package__setup__class_group__grade",
                "package__setup__subject__name_ar",
            )
        )

    logs = ImportLog.objects.filter(school=school).order_by("-started_at")[:20]

    return render(
        request,
        "staging/import_grades.html",
        {
            "assessments": assessments,
            "logs": logs,
            "year": year,
            "openpyxl_ok": OPENPYXL_OK,
        },
    )


# ── تحميل قالب Excel ───────────────────────────────────────


@login_required
@role_required(
    "principal",
    "vice_academic",
    "vice_admin",
    "coordinator",
    "teacher",
    "ese_teacher",
    "admin",
    "secretary",
)
def download_grade_template(request, assessment_id):
    """تحميل ملف Excel فارغ لإدخال الدرجات — مُعبَّأ بأسماء الطلاب"""
    if not OPENPYXL_OK:
        return HttpResponse("مكتبة openpyxl غير مثبتة.", status=500)

    school = request.user.get_school()
    assessment = get_object_or_404(Assessment, id=assessment_id, school=school)

    if not request.user.is_admin() and assessment.package.setup.teacher != request.user:
        return HttpResponse("غير مسموح", status=403)

    enrollments = (
        StudentEnrollment.objects.filter(class_group=assessment.class_group, is_active=True)
        .select_related("student")
        .order_by("student__full_name")
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الدرجات"
    ws.sheet_view.rightToLeft = True

    # ── الستايل ──
    header_fill = PatternFill("solid", fgColor="0F2347")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    info_fill = PatternFill("solid", fgColor="E8F0FE")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── معلومات التقييم (صف 1–4) ──
    info_rows = [
        ("التقييم", assessment.title),
        ("المادة", assessment.package.setup.subject.name_ar),
        ("الفصل", str(assessment.class_group)),
        ("الدرجة القصوى", str(assessment.max_grade)),
    ]
    for i, (label, val) in enumerate(info_rows, start=1):
        ws.cell(i, 1, label).font = Font(bold=True, size=10)
        ws.cell(i, 2, val).fill = info_fill
        ws.cell(i, 2).font = Font(size=10)

    # ── رأس الجدول (صف 6) ──
    headers = ["الرقم الشخصي", "اسم الطالب", "الدرجة", "غائب (1/0)", "ملاحظة"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(6, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 28

    # ── بيانات الطلاب (صف 7 فصاعداً) ──
    for row_idx, enr in enumerate(enrollments, start=7):
        st = enr.student
        ws.cell(row_idx, 1, st.national_id).alignment = center_align
        ws.cell(row_idx, 2, st.full_name).alignment = right_align
        ws.cell(row_idx, 3, "").alignment = center_align  # الدرجة يملؤها المستخدم
        ws.cell(row_idx, 4, 0).alignment = center_align  # غائب
        ws.cell(row_idx, 5, "").alignment = right_align  # ملاحظة

        for col in range(1, 6):
            ws.cell(row_idx, col).border = border

    # ── قفل العمودين A وB ──
    ws.protection.sheet = False  # يظل قابلاً للتعديل على C:E
    ws.row_dimensions[6].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename_ar = f"grades_{assessment.package.setup.subject.name_ar}_{assessment.class_group}.xlsx"
    filename_ascii = f"grades_{assessment_id}.xlsx"
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    # ── MED-003 Fix: RFC 6266 encoding لأسماء الملفات العربية ──
    from urllib.parse import quote

    resp["Content-Disposition"] = (
        f"attachment; filename=\"{filename_ascii}\"; filename*=UTF-8''{quote(filename_ar)}"
    )
    return resp


# ── رفع وتحليل الملف ───────────────────────────────────────


def _find_data_start_row(ws):
    """Find the first data row after the header containing 'الرقم الشخصي'."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and "الرقم الشخصي" in str(cell.value):
                return cell.row + 1
    return 7  # default row in template


def _parse_row(row):
    """Extract fields from an Excel row tuple."""
    national_id = str(row[0]).strip() if row[0] else ""
    grade_raw = row[2] if len(row) > 2 else None
    is_absent = bool(int(row[3])) if len(row) > 3 and row[3] is not None else False
    notes = str(row[4]).strip() if len(row) > 4 and row[4] else ""
    return national_id, grade_raw, is_absent, notes


def _validate_student(national_id, assessment, row_num):
    """Validate student exists and is enrolled. Returns (student, error)."""
    try:
        student = CustomUser.objects.get(national_id=national_id)
    except CustomUser.DoesNotExist:
        return None, f"صف {row_num}: الرقم الشخصي [{national_id}] غير موجود"

    enrolled = StudentEnrollment.objects.filter(
        student=student,
        class_group=assessment.class_group,
        is_active=True,
    ).exists()
    if not enrolled:
        return None, f"صف {row_num}: [{student.full_name}] غير مسجل في الفصل"

    return student, None


def _validate_grade(grade_raw, is_absent, max_grade, student_name, row_num):
    """Parse and validate grade value. Returns (grade, error)."""
    if is_absent or grade_raw is None:
        return None, None
    try:
        grade = float(str(grade_raw).strip())
    except (ValueError, TypeError):
        return None, (f"صف {row_num}: قيمة الدرجة غير صالحة [{grade_raw}] للطالب [{student_name}]")
    if grade < 0 or grade > float(max_grade):
        return None, (
            f"صف {row_num}: الدرجة [{grade}] خارج النطاق (0–{max_grade}) للطالب [{student_name}]"
        )
    return grade, None


def _validate_upload_request(request):
    """Validate the upload request. Returns (uploaded_file, error_redirect)."""
    if request.method != "POST":
        return None, redirect("import_grades_select")

    if not OPENPYXL_OK:
        messages.error(request, "مكتبة openpyxl غير مثبتة — شغّل: pip install openpyxl")
        return None, redirect("import_grades_select")

    uploaded = request.FILES.get("grade_file")
    if not uploaded:
        messages.error(request, "لم يتم رفع أي ملف.")
        return None, redirect("import_grades_select")

    from django.core.exceptions import ValidationError as DjangoValidationError

    from core.validators import FileTypeValidator

    try:
        FileTypeValidator(allowed_types="document")(uploaded)
    except DjangoValidationError as e:
        messages.error(request, e.message)
        return None, redirect("import_grades_select")

    return uploaded, None


@login_required
@role_required(
    "principal",
    "vice_academic",
    "vice_admin",
    "coordinator",
    "teacher",
    "ese_teacher",
    "admin",
    "secretary",
)
def upload_grade_file(request, assessment_id):
    """استيراد الدرجات من ملف Excel — يدعم وضع المعاينة (dry_run)"""
    school = request.user.get_school()
    assessment = get_object_or_404(Assessment, id=assessment_id, school=school)

    if not request.user.is_admin() and assessment.package.setup.teacher != request.user:
        return HttpResponse("غير مسموح", status=403)

    uploaded, error_redirect = _validate_upload_request(request)
    if error_redirect:
        return error_redirect

    dry_run = request.POST.get("dry_run") == "1"
    log = None
    if not dry_run:
        log = ImportLog.objects.create(
            school=school,
            uploaded_by=request.user,
            file_name=uploaded.name,
            status="validating",
        )

    errors = []
    imported = 0
    preview_rows = []

    try:
        wb = openpyxl.load_workbook(uploaded, data_only=True)
        ws = wb.active
        data_start = _find_data_start_row(ws)

        total_rows = 0
        for row in ws.iter_rows(min_row=data_start, values_only=True):
            national_id, grade_raw, is_absent, notes = _parse_row(row)
            if not national_id:
                continue

            total_rows += 1
            row_num = total_rows + data_start - 1

            student, err = _validate_student(national_id, assessment, row_num)
            if err:
                errors.append(err)
                continue

            grade, err = _validate_grade(
                grade_raw, is_absent, assessment.max_grade, student.full_name, row_num
            )
            if err:
                errors.append(err)
                continue

            if dry_run:
                preview_rows.append(
                    {
                        "name": student.full_name,
                        "national_id": national_id,
                        "grade": grade if not is_absent else "—",
                        "is_absent": is_absent,
                        "notes": notes,
                    }
                )
            else:
                GradeService.save_grade(
                    assessment=assessment,
                    student=student,
                    grade=grade,
                    is_absent=is_absent,
                    notes=notes,
                    entered_by=request.user,
                )
            imported += 1

        if not dry_run:
            if imported > 0:
                assessment.status = "graded"
                assessment.save(update_fields=["status"])
            log.status = "completed"
            log.total_rows = total_rows
            log.imported_rows = imported
            log.failed_rows = len(errors)
            log.error_log = errors
            log.completed_at = timezone.now()
            log.save()

    except Exception as exc:
        logger.exception("فشل استيراد الدرجات من Excel: %s", exc)
        if not dry_run and log:
            log.status = "failed"
            log.error_log = [str(exc)]
            log.save()
        messages.error(request, f"فشل تحليل الملف: {exc}")
        return redirect("import_grades_select")

    if dry_run:
        total_rows_count = imported + len(errors)
        log = type(
            "FakeLog",
            (),
            {
                "file_name": uploaded.name,
                "total_rows": total_rows_count,
                "imported_rows": imported,
                "failed_rows": len(errors),
            },
        )()

    return render(
        request,
        "staging/import_result.html",
        {
            "log": log,
            "assessment": assessment,
            "errors": errors,
            "imported": imported,
            "is_dry_run": dry_run,
            "preview_rows": preview_rows,
        },
    )


# ── سجل الاستيراد ──────────────────────────────────────────


@login_required
@role_required(
    "principal",
    "vice_academic",
    "vice_admin",
    "coordinator",
    "teacher",
    "ese_teacher",
    "admin",
    "secretary",
)
def import_log_list(request):
    """سجل كل عمليات الاستيراد"""
    if not request.user.is_admin():
        return HttpResponse("غير مسموح", status=403)

    school = request.user.get_school()
    logs = (
        ImportLog.objects.filter(school=school)
        .select_related("uploaded_by")
        .order_by("-started_at")[:50]
    )

    return render(request, "staging/import_log.html", {"logs": logs})
