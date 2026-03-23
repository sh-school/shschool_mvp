import logging
from decimal import Decimal

from django.conf import settings
from django.contrib import messages

logger = logging.getLogger(__name__)
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST

from core.models import ClassGroup, CustomUser, StudentEnrollment
from operations.models import Subject

from .models import (
    AnnualSubjectResult,
    Assessment,
    AssessmentPackage,
    StudentAssessmentGrade,
    StudentSubjectResult,
    SubjectClassSetup,
)
from .services import GradeService

# ── لوحة تحكم التقييمات ────────────────────────────────────


@login_required
def assessments_dashboard(request):
    school = request.user.get_school()
    semester = request.GET.get("semester", "S1")
    year = request.GET.get("year", settings.CURRENT_ACADEMIC_YEAR)

    if request.user.is_admin():
        # المدير: كل الفصول
        setups = (
            SubjectClassSetup.objects.filter(school=school, academic_year=year, is_active=True)
            .select_related("subject", "class_group", "teacher")
            .order_by("class_group__grade", "class_group__section", "subject__name_ar")
        )

        # إحصائيات عامة
        total_results = StudentSubjectResult.objects.filter(
            school=school, semester=semester
        ).count()
        passed = AnnualSubjectResult.objects.filter(
            school=school, academic_year=year, status="pass"
        ).count()
        failed = AnnualSubjectResult.objects.filter(
            school=school, academic_year=year, status="fail"
        ).count()
        failing_list = GradeService.get_failing_students(school, year)[:10]

    else:
        # المعلم: فصوله فقط
        setups = (
            SubjectClassSetup.objects.filter(
                school=school, teacher=request.user, academic_year=year, is_active=True
            )
            .select_related("subject", "class_group")
            .order_by("class_group__grade", "class_group__section")
        )
        total_results = passed = failed = 0
        failing_list = []

    return render(
        request,
        "assessments/dashboard.html",
        {
            "setups": setups,
            "semester": semester,
            "year": year,
            "total_results": total_results,
            "passed": passed,
            "failed": failed,
            "failing_list": failing_list,
            "SEMESTERS": AssessmentPackage.SEMESTER,
        },
    )


# ── إدارة الباقات والتقييمات ───────────────────────────────


@login_required
def setup_detail(request, setup_id):
    """تفاصيل إعداد مادة — الباقات الأربع"""
    school = request.user.get_school()
    setup = get_object_or_404(SubjectClassSetup, id=setup_id, school=school)

    # التحقق من الصلاحية
    if not request.user.is_admin() and setup.teacher != request.user:
        return HttpResponse("غير مسموح", status=403)

    semester = request.GET.get("semester", "S1")

    # الباقات الأربع لهذا الفصل الدراسي
    packages = (
        AssessmentPackage.objects.filter(setup=setup, semester=semester)
        .prefetch_related("assessments__grades")
        .order_by("package_type")
    )

    # إنشاء الباقات إن لم تكن موجودة
    if not packages.exists():
        semester_max = AssessmentPackage.SEMESTER_MAX.get(semester, Decimal("40"))
        weights = (
            AssessmentPackage.DEFAULT_WEIGHTS_S1
            if semester == "S1"
            else AssessmentPackage.DEFAULT_WEIGHTS_S2
        )
        for ptype, weight in weights.items():
            if weight == Decimal("0"):
                continue  # تخطي الباقات ذات الوزن صفر
            AssessmentPackage.objects.get_or_create(
                setup=setup,
                package_type=ptype,
                semester=semester,
                defaults={
                    "school": school,
                    "weight": weight,
                    "semester_max_grade": semester_max,
                    "is_active": True,
                },
            )
        packages = (
            AssessmentPackage.objects.filter(setup=setup, semester=semester)
            .prefetch_related("assessments")
            .order_by("package_type")
        )

    # نتائج الفصل لهذه المادة
    summary = GradeService.get_class_results_summary(setup)

    # عدد الطلاب
    student_count = StudentEnrollment.objects.filter(
        class_group=setup.class_group, is_active=True
    ).count()

    return render(
        request,
        "assessments/setup_detail.html",
        {
            "setup": setup,
            "packages": packages,
            "semester": semester,
            "summary": summary,
            "student_count": student_count,
            "SEMESTERS": AssessmentPackage.SEMESTER,
        },
    )


@login_required
@require_POST
def create_assessment(request, package_id):
    """إنشاء تقييم جديد في باقة"""
    school = request.user.get_school()
    package = get_object_or_404(AssessmentPackage, id=package_id, school=school)

    if not request.user.is_admin() and package.setup.teacher != request.user:
        return HttpResponse("غير مسموح", status=403)

    title = request.POST.get("title", "").strip()
    atype = request.POST.get("assessment_type", "exam")
    date = request.POST.get("date") or None
    max_g = request.POST.get("max_grade", "100")
    weight = request.POST.get("weight_in_package", "100")
    desc = request.POST.get("description", "")

    if not title:
        messages.error(request, "عنوان التقييم مطلوب")
        return redirect("setup_detail", setup_id=package.setup.id)

    try:
        assessment = Assessment.objects.create(
            package=package,
            school=school,
            title=title,
            assessment_type=atype,
            date=date,
            max_grade=Decimal(max_g),
            weight_in_package=Decimal(weight),
            status="published",
            created_by=request.user,
        )
        messages.success(request, f"تم إنشاء التقييم: {assessment.title}")
    except Exception as e:
        messages.error(request, f"خطأ: {e}")

    return redirect("setup_detail", setup_id=package.setup.id)


# ── إدخال الدرجات ──────────────────────────────────────────


@login_required
def grade_entry(request, assessment_id):
    """صفحة إدخال درجات — تعرض كل طلاب الفصل"""
    school = request.user.get_school()
    assessment = get_object_or_404(Assessment, id=assessment_id, school=school)

    if not request.user.is_admin() and assessment.package.setup.teacher != request.user:
        return HttpResponse("غير مسموح", status=403)

    # طلاب الفصل
    enrollments = (
        StudentEnrollment.objects.filter(class_group=assessment.class_group, is_active=True)
        .select_related("student")
        .order_by("student__full_name")
    )

    # درجات موجودة
    existing = {
        g.student_id: g for g in StudentAssessmentGrade.objects.filter(assessment=assessment)
    }

    students_data = [
        {
            "student": e.student,
            "grade_obj": existing.get(e.student.id),
        }
        for e in enrollments
    ]

    stats = GradeService.get_assessment_stats(assessment)

    return render(
        request,
        "assessments/grade_entry.html",
        {
            "assessment": assessment,
            "students_data": students_data,
            "stats": stats,
        },
    )


@login_required
@require_POST
def save_single_grade(request, assessment_id):
    """HTMX: حفظ درجة طالب واحد"""
    school = request.user.get_school()
    assessment = get_object_or_404(Assessment, id=assessment_id, school=school)

    if not request.user.is_admin() and assessment.package.setup.teacher != request.user:
        return HttpResponse("غير مسموح", status=403)

    student_id = request.POST.get("student_id")
    student = get_object_or_404(CustomUser, id=student_id)
    is_absent = request.POST.get("is_absent") == "1"
    is_excused = request.POST.get("is_excused") == "1"
    notes = request.POST.get("notes", "")

    grade = None
    if not is_absent and not is_excused:
        raw = request.POST.get("grade", "").strip()
        if raw:
            try:
                grade = Decimal(raw)
            except Exception:
                logger.exception("فشل تحويل الدرجة إلى Decimal: %r", raw)
                return HttpResponse("درجة غير صالحة", status=400)

    grade_obj, _ = GradeService.save_grade(
        assessment=assessment,
        student=student,
        grade=grade,
        is_absent=is_absent,
        is_excused=is_excused,
        notes=notes,
        entered_by=request.user,
    )

    stats = GradeService.get_assessment_stats(assessment)

    return render(
        request,
        "assessments/partials/grade_row.html",
        {
            "student": student,
            "grade_obj": grade_obj,
            "assessment": assessment,
            "stats": stats,
        },
    )


@login_required
@require_POST
def save_all_grades(request, assessment_id):
    """حفظ كل الدرجات دفعة واحدة من form"""
    school = request.user.get_school()
    assessment = get_object_or_404(Assessment, id=assessment_id, school=school)

    if not request.user.is_admin() and assessment.package.setup.teacher != request.user:
        return HttpResponse("غير مسموح", status=403)

    enrollments = StudentEnrollment.objects.filter(
        class_group=assessment.class_group, is_active=True
    ).select_related("student")

    saved = 0
    for enr in enrollments:
        sid = str(enr.student.id)
        is_absent = request.POST.get(f"absent_{sid}") == "1"
        is_excused = request.POST.get(f"excused_{sid}") == "1"
        notes = request.POST.get(f"notes_{sid}", "")
        grade = None

        if not is_absent and not is_excused:
            raw = request.POST.get(f"grade_{sid}", "").strip()
            if raw:
                try:
                    grade = Decimal(raw)
                except Exception:
                    logger.exception("فشل تحويل درجة الطالب %s إلى Decimal: %r", sid, raw)
                    continue

        GradeService.save_grade(
            assessment=assessment,
            student=enr.student,
            grade=grade,
            is_absent=is_absent,
            is_excused=is_excused,
            notes=notes,
            entered_by=request.user,
        )
        saved += 1

    # تحديث حالة التقييم
    assessment.status = "graded"
    assessment.save(update_fields=["status"])

    messages.success(request, f"تم حفظ {saved} درجة بنجاح")
    return redirect("grade_entry", assessment_id=assessment_id)


# ── كشوف الدرجات ───────────────────────────────────────────


@login_required
def class_gradebook(request, setup_id):
    """كشف الدرجات الكامل للفصل في مادة — يدعم عرض فصل أو السنوي"""
    school = request.user.get_school()
    setup = get_object_or_404(SubjectClassSetup, id=setup_id, school=school)

    if not request.user.is_admin() and setup.teacher != request.user:
        return HttpResponse("غير مسموح", status=403)

    semester = request.GET.get("semester", "S1")

    enrollments = (
        StudentEnrollment.objects.filter(class_group=setup.class_group, is_active=True)
        .select_related("student")
        .order_by("student__full_name")
    )

    # للعرض السنوي أو الفصلي
    show_annual = semester == "annual"
    packages = (
        AssessmentPackage.objects.filter(
            setup=setup, semester=semester if not show_annual else "S1", is_active=True
        )
        .prefetch_related("assessments")
        .order_by("package_type")
    )

    rows = []
    for enr in enrollments:
        student = enr.student
        pkg_scores = {}

        if not show_annual:
            for pkg in packages:
                score = GradeService.calc_package_score(student, pkg)
                pkg_scores[pkg.package_type] = score

        # نتيجة الفصل
        try:
            sem_result = StudentSubjectResult.objects.get(
                student=student, setup=setup, semester=semester if not show_annual else "S1"
            )
        except StudentSubjectResult.DoesNotExist:
            sem_result = None

        # النتيجة السنوية
        try:
            annual_result = AnnualSubjectResult.objects.get(
                student=student, setup=setup, academic_year=setup.academic_year
            )
        except AnnualSubjectResult.DoesNotExist:
            annual_result = None

        rows.append(
            {
                "student": student,
                "pkg_scores": pkg_scores,
                "semester_result": sem_result,
                "annual_result": annual_result,
            }
        )

    summary = GradeService.get_class_results_summary(setup)

    return render(
        request,
        "assessments/class_gradebook.html",
        {
            "setup": setup,
            "packages": packages,
            "rows": rows,
            "summary": summary,
            "semester": semester,
            "show_annual": show_annual,
            "SEMESTERS": AssessmentPackage.SEMESTER,
        },
    )


@login_required
def export_gradebook(request, setup_id):
    """تصدير كشف الدرجات إلى Excel"""
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    school = request.user.get_school()
    setup = get_object_or_404(SubjectClassSetup, id=setup_id, school=school)

    if not request.user.is_admin() and setup.teacher != request.user:
        return HttpResponse("غير مسموح", status=403)

    semester = request.GET.get("semester", "S1")

    enrollments = (
        StudentEnrollment.objects.filter(class_group=setup.class_group, is_active=True)
        .select_related("student")
        .order_by("student__full_name")
    )

    packages = (
        AssessmentPackage.objects.filter(setup=setup, semester=semester, is_active=True)
        .prefetch_related("assessments")
        .order_by("package_type")
    )

    # ── بناء الـ Workbook ──────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "كشف الدرجات"
    ws.sheet_view.rightToLeft = True

    # ألوان
    HEADER_FILL = PatternFill("solid", fgColor="8A1538")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    ALT_FILL = PatternFill("solid", fgColor="FFF0F3")
    BORDER_SIDE = Side(style="thin", color="CCCCCC")
    THIN_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")

    pkg_labels = {
        "P1": "الباقة 1",
        "P2": "الباقة 2",
        "P3": "الباقة 3",
        "P4": "الباقة 4",
    }
    sem_labels = {"S1": "الفصل الأول (40)", "S2": "الفصل الثاني (60)"}

    # ── السطر 1: عنوان ─────────────────────────────────────
    title = (
        f"كشف درجات | {setup.subject.name_ar} | "
        f"{setup.class_group} | {sem_labels.get(semester, semester)} | "
        f"{setup.academic_year}"
    )
    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13, color="8A1538")
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    # ── السطر 2: رؤوس الأعمدة ──────────────────────────────
    pkg_list = list(packages)
    headers = (
        ["م", "اسم الطالب", "الرقم الوطني"]
        + [pkg_labels.get(p.package_type, p.package_type) for p in pkg_list]
        + ["مجموع الفصل", "الحالة"]
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    ws.row_dimensions[2].height = 22

    # ── البيانات ────────────────────────────────────────────
    for row_idx, enr in enumerate(enrollments, start=1):
        student = enr.student
        excel_row = row_idx + 2
        fill = ALT_FILL if row_idx % 2 == 0 else None

        pkg_scores = {p.package_type: GradeService.calc_package_score(student, p) for p in pkg_list}

        try:
            sem_result = StudentSubjectResult.objects.get(
                student=student, setup=setup, semester=semester
            )
            total = float(sem_result.total) if sem_result.total is not None else ""
            status = (
                "ناجح ✓"
                if (sem_result.total or 0) >= (sem_result.semester_max * Decimal("0.5"))
                else "راسب ✗"
            )
        except StudentSubjectResult.DoesNotExist:
            total, status = "", "—"

        row_data = (
            [row_idx, student.full_name, student.national_id]
            + [float(pkg_scores.get(p.package_type) or 0) for p in pkg_list]
            + [total, status]
        )

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = CENTER if col_idx != 2 else RIGHT_ALIGN
            if fill:
                cell.fill = fill

    # ── عرض الأعمدة ────────────────────────────────────────
    col_widths = [5, 30, 18] + [12] * len(pkg_list) + [14, 10]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── تجميد الصفوف الأولى ────────────────────────────────
    ws.freeze_panes = "A3"

    # ── إرسال الملف ────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"gradebook_{setup.subject.name_ar}_{setup.class_group}_{semester}.xlsx".replace(
        " ", "_"
    ).replace("/", "-")
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@login_required
@require_POST
def recalculate_class(request, setup_id):
    """إعادة حساب درجات كل طلاب الفصل — Admin أو المعلم المسؤول فقط"""
    school = request.user.get_school()
    setup = get_object_or_404(SubjectClassSetup, id=setup_id, school=school)

    if not request.user.is_admin() and setup.teacher != request.user:
        return HttpResponse("غير مسموح", status=403)

    GradeService.recalculate_full_class(setup)
    messages.success(
        request, f"تم إعادة حساب درجات {setup.class_group} في {setup.subject.name_ar} بنجاح."
    )
    return redirect("class_gradebook", setup_id=setup_id)


@login_required
def student_report(request, student_id):
    """كشف درجات سنوي للطالب في كل مواده"""
    school = request.user.get_school()
    student = get_object_or_404(CustomUser, id=student_id)
    year = request.GET.get("year", settings.CURRENT_ACADEMIC_YEAR)

    if not request.user.is_admin() and not request.user.is_teacher() and request.user != student:
        return HttpResponse("غير مسموح", status=403)

    results = GradeService.get_student_annual_report(student, school, year)
    total_subjects = results.count()
    passed = results.filter(status="pass").count()
    failed = results.filter(status="fail").count()

    return render(
        request,
        "assessments/student_report.html",
        {
            "student": student,
            "results": results,
            "year": year,
            "total_subjects": total_subjects,
            "passed": passed,
            "failed": failed,
            "SEMESTERS": AssessmentPackage.SEMESTER,
        },
    )


@login_required
def failing_students(request):
    """قائمة الطلاب الراسبين — للمدير"""
    if not request.user.is_admin():
        return HttpResponse("غير مسموح", status=403)

    school = request.user.get_school()
    semester = request.GET.get("semester", "S1")
    year = request.GET.get("year", settings.CURRENT_ACADEMIC_YEAR)

    failing = GradeService.get_failing_students(school, year)

    # تجميع بالطالب
    by_student = {}
    for r in failing:
        sid = r.student.id
        if sid not in by_student:
            by_student[sid] = {"student": r.student, "subjects": []}
        by_student[sid]["subjects"].append(r)

    return render(
        request,
        "assessments/failing_students.html",
        {
            "by_student": list(by_student.values()),
            "semester": semester,
            "year": year,
            "total": len(by_student),
            "SEMESTERS": AssessmentPackage.SEMESTER,
        },
    )


# ── إعداد المواد (للمدير) ──────────────────────────────────


@login_required
def setup_subject(request):
    """ربط مادة بفصل ومعلم"""
    if not request.user.is_admin():
        return HttpResponse("غير مسموح", status=403)

    school = request.user.get_school()

    if request.method == "POST":
        subject_id = request.POST.get("subject")
        class_id = request.POST.get("class_group")
        teacher_id = request.POST.get("teacher")
        year = request.POST.get("academic_year", settings.CURRENT_ACADEMIC_YEAR)

        subject = get_object_or_404(Subject, id=subject_id, school=school)
        class_group = get_object_or_404(ClassGroup, id=class_id, school=school)
        teacher = get_object_or_404(CustomUser, id=teacher_id)

        setup, created = SubjectClassSetup.objects.get_or_create(
            school=school,
            subject=subject,
            class_group=class_group,
            academic_year=year,
            defaults={"teacher": teacher, "is_active": True},
        )
        if not created:
            setup.teacher = teacher
            setup.is_active = True
            setup.save(update_fields=["teacher", "is_active"])

        msg = "تم الإنشاء" if created else "تم التحديث"
        messages.success(request, f"{msg}: {subject.name_ar} | {class_group}")
        return redirect("assessments_dashboard")

    # GET
    subjects = Subject.objects.filter(school=school).order_by("name_ar")
    classes = ClassGroup.objects.filter(school=school, is_active=True).order_by("grade", "section")
    from core.models import Membership

    t_ids = Membership.objects.filter(
        school=school, is_active=True, role__name__in=["teacher", "coordinator"]
    ).values_list("user_id", flat=True)
    teachers = CustomUser.objects.filter(id__in=t_ids).order_by("full_name")

    return render(
        request,
        "assessments/setup_subject.html",
        {
            "subjects": subjects,
            "classes": classes,
            "teachers": teachers,
        },
    )
