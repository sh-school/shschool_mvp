"""[ENROLMENT] قيدُ العام من سجلّ القيد الوزاريّ.

كان محرّك الترفيع موقوفاً على أربعة أسئلةٍ لم تُجب: توقيتُ الدور الثاني، وهل
البقاءُ في الشعبة نفسها هو الأصل، ومن يُسنِد المسار، وهل الراسبُ يبقى. وسجلُّ
القيد يجيبها كلَّها لأنّه مخرجُ تطبيقها في الوزارة: يقول بالاسم أين يجلس كلُّ
طالب. فلا محرّكَ يجتهد، بل استيرادٌ يقرأ القرار وينفّذه.

وثلاثةُ عيوبٍ كشفها التشغيلُ الجافّ قبل أيّ كتابةٍ في الإنتاج، ولكلٍّ منها
هنا دعوى تمنع عودتَه:

  ١) السجلُّ يكتب «07/1» والقاعدةُ تُنتج «7/1» — فكانت كلُّ شُعب السابع إلى
     التاسع تبدو «مفقودة» فتُنشأ فوق القائم.
  ٢) طلابُ شُعبٍ تُنشأ في التشغيل نفسه كانوا يسقطون صامتين — أربعةُ طلابٍ في
     شُعب الدعم لم يظهروا في أيّ خانة.
  ٣) مقارنةُ «من في المنصّة وليس في السجلّ» كانت تشمل الطاقم، فيُعدّ كلُّ
     معلّمٍ غائباً عن سجلّ قيد الطلاب.
"""

import io
from pathlib import Path

import pytest
from django.core.management import CommandError, call_command

from core.management.commands.import_enrolment_register import canonical
from core.models import ClassGroup, CustomUser
from core.models.academic import StudentEnrollment

YEAR = "2026-2027"


# ── بناءُ ملفٍّ يحاكي السجلّ الرسميّ ──────────────────────────────────


def _register(tmp_path, grades, tracks=()):
    """يبني ملفّاً ببنية السجلّ: رؤوسٌ في السطر الثالث، وورقةٌ لكلّ صفّ.

    `grades`: {"07": [(رقم، اسم، صفّ، شعبة), …]}
    `tracks`: أسماءُ أوراق الشُّعب التي تحمل المسار.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for grade, rows in grades.items():
        ws = wb.create_sheet(f"الصف- {grade}")
        ws.append([f"سجل القيد (الصف: {grade})"])
        ws.append(["المدرسة: 20417 — السنة الدراسية: 2026-2027"])
        ws.append(["الرقم", "الاسم", "الصف", "الشعبة الصفية"])
        for r in rows:
            ws.append(list(r))
    for title in tracks:
        wb.create_sheet(title).append(["x"])
    path = tmp_path / "register.xlsx"
    wb.save(path)
    return str(path)


def _run(path, *args):
    out = io.StringIO()
    call_command("import_enrolment_register", path, "--year", YEAR, *args, stdout=out)
    return out.getvalue()


@pytest.fixture
def student(db, school):
    """طالبٌ قائمٌ في المنصّة برقمه الحقيقيّ في السجلّ."""
    from core.models.access import Membership, Role

    user = CustomUser.objects.create(national_id="31463401932", full_name="تميم سعد")
    role, _ = Role.objects.get_or_create(school=school, name="student")
    Membership.objects.create(user=user, school=school, role=role, is_active=True)
    return user


@pytest.fixture
def simple(tmp_path, student):
    return _register(
        tmp_path,
        {"07": [("31463401932", "تميم سعد منصور", "7", "07/1")]},
    )


# ── العيب الأوّل: الصفرُ البادئ ───────────────────────────────────────


@pytest.mark.parametrize(
    ("raw", "expected"),
    [("07/1", "7/1"), ("7/1", "7/1"), ("08/ESE", "8/ESE"), ("10/4", "10/4"), ("12/1", "12/1")],
)
def test_the_leading_zero_does_not_make_a_new_section(raw, expected):
    """«07/1» في السجلّ و«7/1» في القاعدة شعبةٌ واحدة.

    ولولا التسوية لأُنشئت شُعبُ السابع إلى التاسع كلُّها من جديد فوق القائم —
    وهو ما أظهره أوّلُ تشغيلٍ جافّ: ستَّ عشرةَ شعبةً «تُنشأ» وهي قائمة.
    """
    assert canonical(raw) == expected


def test_an_existing_section_written_with_a_zero_is_recognised(db, school, simple):
    ClassGroup.objects.create(
        school=school, grade="G7", section="1", level_type="prep", academic_year=YEAR
    )

    out = _run(simple)

    assert "تُنشأ" not in out
    assert ClassGroup.objects.filter(academic_year=YEAR).count() == 1


# ── العيب الثاني: من شعبتُه تُنشأ الآن ────────────────────────────────


def test_a_student_whose_section_is_created_in_this_run_is_still_enrolled(
    db, school, student, tmp_path
):
    """شعبةُ الدعم لم تكن في المنصّة، وطلابُها كانوا يسقطون صامتين."""
    path = _register(tmp_path, {"08": [("31463401932", "تميم سعد", "8", "08/ESE")]})

    _run(path, "--apply")

    group = ClassGroup.objects.get(academic_year=YEAR, grade="G8", section="ESE")
    assert StudentEnrollment.objects.filter(
        student=student, class_group=group, is_active=True
    ).exists()


#: الخاناتُ التي يوزَّع عليها السجلّ — وهي متباينةٌ ومستوعِبة.
BUCKETS = (
    "قيدُهم مطابقٌ أصلاً",
    "يُقيَّدون في شُعبهم",
    "يُقيَّدون في شُعبٍ تُنشأ الآن",
    "بلا حساب — لن يُقيَّدوا",
    "تُنشأ حساباتُهم",
    "يُنقلون من شعبةٍ إلى أخرى",
    "بلا شعبةٍ في السجلّ — يُتركون",
)


def _counts(report):
    """أعدادُ الخانات كما عُرضت — تُقرأ من التقرير نفسه لا من الداخل."""
    found = {}
    for line in report.splitlines():
        for bucket in BUCKETS:
            if line.strip().startswith(bucket + ":"):
                found[bucket] = int(line.rsplit(":", 1)[1].strip())
    return found


def test_every_registered_student_lands_in_exactly_one_bucket(db, school, student, tmp_path):
    """مجموعُ الخانات يساوي عددَ السجلّ — وإلّا فثمّة من سقط بلا خبر.

    وهذا هو الثابتُ الذي كشف العيب: كان المجموع ٧٠٢ من ٧٠٦، والأربعةُ
    الناقصون طلابُ شُعب الدعم التي تُنشأ في التشغيل نفسه.
    """
    path = _register(
        tmp_path,
        {
            "07": [
                ("31463401932", "تميم سعد", "7", "07/1"),
                ("31463403208", "تركي محسن", "7", "07/2"),
                ("31481800902", "اسامه اشرف", "7", "-"),
            ]
        },
    )

    counts = _counts(_run(path))

    assert sum(counts.values()) == 3, counts
    assert counts["يُقيَّدون في شُعبٍ تُنشأ الآن"] == 1, "له حسابٌ وشعبتُه تُنشأ"
    assert counts["بلا حساب — لن يُقيَّدوا"] == 1
    assert counts["بلا شعبةٍ في السجلّ — يُتركون"] == 1


# ── العيب الثالث: الطاقم ليس طالباً ──────────────────────────────────


def test_staff_are_not_counted_as_missing_from_the_student_register(
    db, school, student, teacher_user, simple
):
    """سجلُّ قيد الطلاب لا يذكر المعلّمين — ولا يُعدّون غائبين عنه."""
    out = _run(simple)

    assert "في المنصّة وليسوا في السجلّ" not in out


# ── المسارُ لا يُؤخَذ من هذا الملفّ ───────────────────────────────────


@pytest.fixture
def eleven_two(db, school):
    return ClassGroup.objects.create(
        school=school,
        grade="G11",
        section="2",
        level_type="sec",
        track="humanities",
        academic_year=YEAR,
    )


@pytest.fixture
def contradicting(tmp_path, student):
    """ورقةٌ تسمّي 11/2 «تكنولوجي» — وهو ما تكذّبه موادُّ الجدول."""
    return _register(
        tmp_path,
        {"11": [("31463401932", "تميم سعد", "11", "11/2")]},
        tracks=["الشعبة الصفية- 11-2-Technology-"],
    )


def test_the_register_sheet_names_never_overwrite_a_track(
    db, school, student, eleven_two, contradicting
):
    """أسماءُ أوراق السجلّ تكتب «11-2-Technology» و«11-4-Humanities»، وجدولُ
    المدرسة يقول عكسَه: 11/4 و12/4 وحدهما فيهما علومُ الحاسب وتكنولوجيا
    المعلومات، والباقي إدارةُ أعمالٍ وتاريخٌ وجغرافيا.

    فالحَكَمُ ما يُدرَّس فعلاً، والمنصّةُ كانت مصيبةً — ولو كتب هذا الأمرُ
    المسارَ من الملفّ لأفسد أربعَ شُعب.
    """
    _run(contradicting, "--apply")

    eleven_two.refresh_from_db()
    assert eleven_two.track == "humanities"


def test_a_contradicting_track_is_announced_not_applied(
    db, school, student, eleven_two, contradicting
):
    """يُنبّه ولا يأمر: الخلافُ إشارةٌ تستحقّ النظر، لا سبباً للكتابة."""
    out = _run(contradicting)

    assert "ولا يُغيَّر شيء" in out
    assert "المنصّة «humanities» · ورقةُ السجلّ «technology»" in out


def test_a_section_created_here_gets_no_track_from_the_file(db, school, student, tmp_path):
    """شعبةٌ جديدةٌ تُنشأ بلا مسار — وحالةُ «بلا مسار» مشروعةٌ في النموذج
    حتى يُضبط بـ`set_class_tracks` مقيساً بما يُدرَّس."""
    path = _register(
        tmp_path,
        {"11": [("31463401932", "تميم سعد", "11", "11/2")]},
        tracks=["الشعبة الصفية- 11-2-Technology-"],
    )

    _run(path, "--apply")

    assert ClassGroup.objects.get(academic_year=YEAR, grade="G11", section="2").track == ""


# ── مطابقةُ الترقيم: تُكتب ولا تُخمَّن ────────────────────────────────


def test_the_register_numbering_can_be_mapped_onto_the_schools(db, school, student, tmp_path):
    """في الثاني عشر يضع السجلُّ سبعةَ طلاب التكنولوجي في «12/2» وتسمّيهم
    المدرسة «12/4». ولولا المطابقةُ لدرس التكنولوجيّون جدولَ الإنسانيات."""
    target = ClassGroup.objects.create(
        school=school,
        grade="G12",
        section="4",
        level_type="sec",
        track="technology",
        academic_year=YEAR,
    )
    path = _register(tmp_path, {"12": [("31463401932", "تميم", "12-Technology", "12/2")]})

    _run(path, "--map", "12/2=12/4", "--apply")

    assert StudentEnrollment.objects.get(student=student, is_active=True).class_group == target


def test_a_mapping_that_sends_two_sections_to_one_is_refused(db, school, student, simple):
    with pytest.raises(CommandError, match="شعبتان"):
        _run(simple, "--map", "12/2=12/4", "12/3=12/4")


@pytest.mark.parametrize("bad", ["12/2", "12/2=12/4=12/1", "اثنعشر=12/4"])
def test_a_malformed_mapping_is_refused(db, school, student, simple, bad):
    with pytest.raises(CommandError, match="غير مفهومة"):
        _run(simple, "--map", bad)


# ── الحدود: ما لا يفعله ──────────────────────────────────────────────


def test_nothing_is_written_without_apply(db, school, student, simple):
    _run(simple)

    assert not StudentEnrollment.objects.exists()
    assert not ClassGroup.objects.exists()


def test_an_account_is_not_invented_without_the_flag(db, school, tmp_path):
    """رقمٌ شخصيٌّ حقيقيٌّ لقاصر لا يُنشأ به حسابٌ بالصدفة."""
    path = _register(tmp_path, {"07": [("31463409999", "طالبٌ جديد", "7", "07/1")]})

    assert "بلا حساب — لن يُقيَّدوا: 1" in _run(path)
    with pytest.raises(CommandError, match="--create-missing"):
        _run(path, "--apply")


def test_a_student_absent_from_the_register_is_left_alone(db, school, student, tmp_path):
    """خرّيجٌ أو منتقل — وإغلاقُ قيده قرارٌ إداريٌّ لا يُتّخذ من سطر أوامر."""
    from core.models.access import Membership, Role

    other = CustomUser.objects.create(national_id="31463400000", full_name="خرّيج")
    role = Role.objects.get(school=school, name="student")
    Membership.objects.create(user=other, school=school, role=role, is_active=True)
    path = _register(tmp_path, {"07": [("31463401932", "تميم سعد", "7", "07/1")]})

    out = _run(path, "--apply")

    assert "في المنصّة وليسوا في السجلّ — لا يُمسّون: 1" in out
    assert CustomUser.objects.filter(pk=other.pk).exists()
    assert Membership.objects.get(user=other).is_active


def test_running_twice_changes_nothing_the_second_time(db, school, student, simple):
    """المدرسةُ ستُشغّله مرّاتٍ كلّما صدر سجلٌّ محدَّث."""
    _run(simple, "--apply")
    before = StudentEnrollment.objects.count()

    out = _run(simple, "--apply")

    assert StudentEnrollment.objects.count() == before
    assert "قيدُهم مطابقٌ أصلاً: 1" in out


def test_a_student_who_changed_section_is_moved_not_duplicated(db, school, student, tmp_path):
    _run(_register(tmp_path, {"07": [("31463401932", "تميم", "7", "07/1")]}), "--apply")
    moved = tmp_path / "b"
    moved.mkdir()

    _run(_register(moved, {"07": [("31463401932", "تميم", "7", "07/2")]}), "--apply")

    active = StudentEnrollment.objects.filter(student=student, is_active=True)
    assert active.count() == 1
    assert active.first().class_group.section == "2"
    assert StudentEnrollment.objects.filter(student=student).count() == 2, "القديمُ يبقى تاريخاً"


def test_a_file_that_is_not_the_register_is_refused(db, school, tmp_path):
    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.append(["لا شيء"])
    path = tmp_path / "other.xlsx"
    wb.save(path)

    with pytest.raises(CommandError, match="سجلّ القيد"):
        _run(str(path))


def test_a_missing_file_says_so(db, school):
    with pytest.raises(CommandError, match="لا ملفّ"):
        _run(str(Path("لا-وجود-له.xlsx")))
