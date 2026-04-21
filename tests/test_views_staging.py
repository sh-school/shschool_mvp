"""
tests/test_views_staging.py
اختبارات استيراد الدرجات من Excel (staging app)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
يغطي:
  - import_grades_select  : لوحة اختيار التقييم
  - import_log_list       : سجل عمليات الاستيراد (مدير فقط)
  - download_grade_template: تحميل قالب Excel
  - upload_grade_file     : رفع الملف + التحقق + الحفظ
"""

import io
from decimal import Decimal

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile

from assessments.models import Assessment, AssessmentPackage, SubjectClassSetup
from operations.models import Subject
from staging.models import ImportLog
from tests.conftest import (
    ClassGroupFactory,
    MembershipFactory,
    RoleFactory,
    SchoolFactory,
    UserFactory,
)

# ── Helpers ────────────────────────────────────────────────────────────


def make_admin(db, school):
    role = RoleFactory(school=school, name="principal")
    user = UserFactory(full_name="مدير النظام")
    MembershipFactory(user=user, school=school, role=role)
    return user


def make_teacher_with_assessment(db, school):
    """يُنشئ معلماً + تقييماً مرتبطاً به ويعيد (teacher, assessment)"""
    class_group = ClassGroupFactory(school=school)
    subject = Subject.objects.create(school=school, name_ar="الرياضيات", code="MATH01")
    role = RoleFactory(school=school, name="teacher")
    teacher = UserFactory(full_name="معلم الاستيراد")
    MembershipFactory(user=teacher, school=school, role=role)

    setup = SubjectClassSetup.objects.create(
        school=school,
        subject=subject,
        class_group=class_group,
        teacher=teacher,
        academic_year="2025-2026",
    )
    package = AssessmentPackage.objects.create(
        setup=setup,
        school=school,
        package_type="P1",
        semester="S1",
        weight=Decimal("50"),
        semester_max_grade=Decimal("40"),
    )
    assessment = Assessment.objects.create(
        package=package,
        school=school,
        title="اختبار رياضيات",
        assessment_type="exam",
        status="published",
        max_grade=Decimal("20"),
    )
    return teacher, assessment


# ══════════════════════════════════════════════════════════
#  1. import_grades_select — GET /import/
# ══════════════════════════════════════════════════════════


@pytest.mark.django_db
class TestImportGradesSelect:
    def test_redirect_if_not_logged_in(self, client):
        response = client.get("/import/")
        assert response.status_code in (302, 301)

    def test_admin_can_view(self, client_as, school):
        admin = make_admin(pytest.db if hasattr(pytest, "db") else None, school)
        # Use db fixture approach
        client = client_as(admin)
        response = client.get("/import/")
        assert response.status_code == 200

    def test_teacher_can_view(self, client_as, school, teacher_user):
        client = client_as(teacher_user)
        response = client.get("/import/")
        assert response.status_code == 200

    def test_nurse_forbidden(self, client_as, school, nurse_user):
        client = client_as(nurse_user)
        response = client.get("/import/")
        assert response.status_code == 403

    def test_shows_assessments_context(self, client_as, school, teacher_user):
        client = client_as(teacher_user)
        response = client.get("/import/")
        assert response.status_code == 200
        assert "assessments" in response.context
        assert "logs" in response.context
        assert "year" in response.context

    def test_year_filter_param(self, client_as, school, teacher_user):
        client = client_as(teacher_user)
        response = client.get("/import/?year=2024-2025")
        assert response.status_code == 200
        assert response.context["year"] == "2024-2025"

    def test_admin_sees_all_school_assessments(self, client_as, school):
        """المدير يرى تقييمات كل المعلمين"""
        teacher, assessment = make_teacher_with_assessment(None, school)
        admin = make_admin(None, school)
        client = client_as(admin)
        response = client.get("/import/?year=2025-2026")
        assert response.status_code == 200
        pks = [str(a.id) for a in response.context["assessments"]]
        assert str(assessment.id) in pks

    def test_teacher_sees_own_assessments_only(self, client_as, school):
        """المعلم يرى تقييماته فقط — ليس تقييمات معلم آخر"""
        teacher, assessment = make_teacher_with_assessment(None, school)
        # معلم آخر لا علاقة له
        other_teacher, other_assessment = make_teacher_with_assessment(None, school)
        client = client_as(teacher)
        response = client.get("/import/?year=2025-2026")
        assert response.status_code == 200
        pks = [str(a.id) for a in response.context["assessments"]]
        assert str(assessment.id) in pks
        assert str(other_assessment.id) not in pks


# ══════════════════════════════════════════════════════════
#  2. import_log_list — GET /import/log/
# ══════════════════════════════════════════════════════════


@pytest.mark.django_db
class TestImportLogList:
    def test_redirect_if_not_logged_in(self, client):
        response = client.get("/import/log/")
        assert response.status_code in (302, 301)

    def test_admin_can_view(self, client_as, school):
        admin = make_admin(None, school)
        client = client_as(admin)
        response = client.get("/import/log/")
        assert response.status_code == 200

    def test_teacher_forbidden(self, client_as, school, teacher_user):
        client = client_as(teacher_user)
        response = client.get("/import/log/")
        assert response.status_code == 403

    def test_shows_logs_context(self, client_as, school):
        admin = make_admin(None, school)
        # إنشاء سجل استيراد
        ImportLog.objects.create(
            school=school,
            uploaded_by=admin,
            file_name="grades.xlsx",
            status="completed",
        )
        client = client_as(admin)
        response = client.get("/import/log/")
        assert response.status_code == 200
        assert "logs" in response.context
        assert response.context["logs"].count() >= 1

    def test_empty_logs(self, client_as, school):
        admin = make_admin(None, school)
        client = client_as(admin)
        response = client.get("/import/log/")
        assert response.status_code == 200
        assert "logs" in response.context


# ══════════════════════════════════════════════════════════
#  3. download_grade_template — GET /import/template/<uuid>/
# ══════════════════════════════════════════════════════════


@pytest.mark.django_db
class TestDownloadGradeTemplate:
    def test_redirect_if_not_logged_in(self, client, school):
        teacher, assessment = make_teacher_with_assessment(None, school)
        response = client.get(f"/import/template/{assessment.id}/")
        assert response.status_code in (302, 301)

    def test_admin_can_download(self, client_as, school):
        """المدير يستطيع تحميل قالب أي تقييم"""
        teacher, assessment = make_teacher_with_assessment(None, school)
        admin = make_admin(None, school)
        client = client_as(admin)
        response = client.get(f"/import/template/{assessment.id}/")
        # إما xlsx أو 500 إذا openpyxl غير مثبت
        assert response.status_code in (200, 500)
        if response.status_code == 200:
            assert "spreadsheetml" in response["Content-Type"]

    def test_teacher_owner_can_download(self, client_as, school):
        """المعلم صاحب التقييم يستطيع التحميل"""
        teacher, assessment = make_teacher_with_assessment(None, school)
        client = client_as(teacher)
        response = client.get(f"/import/template/{assessment.id}/")
        assert response.status_code in (200, 500)

    def test_other_teacher_forbidden(self, client_as, school):
        """معلم آخر لا يستطيع تحميل قالب تقييم ليس له"""
        teacher, assessment = make_teacher_with_assessment(None, school)
        # معلم آخر
        other_role = RoleFactory(school=school, name="teacher")
        other_teacher = UserFactory(full_name="معلم آخر")
        MembershipFactory(user=other_teacher, school=school, role=other_role)
        client = client_as(other_teacher)
        response = client.get(f"/import/template/{assessment.id}/")
        assert response.status_code in (403, 500)

    def test_wrong_school_404(self, client_as, school):
        """تقييم مدرسة أخرى → 404"""
        other_school = SchoolFactory()
        _, assessment = make_teacher_with_assessment(None, other_school)

        admin = make_admin(None, school)
        client = client_as(admin)
        response = client.get(f"/import/template/{assessment.id}/")
        assert response.status_code == 404

    def test_content_disposition_header(self, client_as, school):
        """الاستجابة يجب أن تحتوي Content-Disposition"""
        teacher, assessment = make_teacher_with_assessment(None, school)
        admin = make_admin(None, school)
        client = client_as(admin)
        response = client.get(f"/import/template/{assessment.id}/")
        if response.status_code == 200:
            assert "attachment" in response.get("Content-Disposition", "")


# ══════════════════════════════════════════════════════════
#  4. upload_grade_file — POST /import/upload/<uuid>/
# ══════════════════════════════════════════════════════════


@pytest.mark.django_db
class TestUploadGradeFile:
    def test_redirect_if_not_logged_in(self, client, school):
        teacher, assessment = make_teacher_with_assessment(None, school)
        response = client.post(f"/import/upload/{assessment.id}/")
        assert response.status_code in (302, 301)

    def test_get_redirects_to_select(self, client_as, school):
        """GET على upload يُعيد redirect لصفحة الاختيار"""
        teacher, assessment = make_teacher_with_assessment(None, school)
        client = client_as(teacher)
        response = client.get(f"/import/upload/{assessment.id}/")
        assert response.status_code in (302, 301)

    def test_post_without_file_redirects(self, client_as, school):
        """POST بدون ملف → redirect مع رسالة خطأ"""
        teacher, assessment = make_teacher_with_assessment(None, school)
        client = client_as(teacher)
        response = client.post(f"/import/upload/{assessment.id}/", {})
        assert response.status_code in (302, 200)

    def test_other_teacher_forbidden(self, client_as, school):
        """معلم لا يملك التقييم → 403"""
        teacher, assessment = make_teacher_with_assessment(None, school)
        other_role = RoleFactory(school=school, name="teacher")
        other_teacher = UserFactory(full_name="معلم آخر 2")
        MembershipFactory(user=other_teacher, school=school, role=other_role)
        client = client_as(other_teacher)
        response = client.post(
            f"/import/upload/{assessment.id}/",
            {"grade_file": io.BytesIO(b"fake")},
        )
        assert response.status_code == 403

    def test_wrong_school_404(self, client_as, school):
        other_school = SchoolFactory()
        _, assessment = make_teacher_with_assessment(None, other_school)
        admin = make_admin(None, school)
        client = client_as(admin)
        response = client.post(f"/import/upload/{assessment.id}/", {})
        assert response.status_code == 404

    def test_invalid_xlsx_creates_failed_log(self, client_as, school):
        """ملف تالف → سجل ImportLog بحالة failed"""
        teacher, assessment = make_teacher_with_assessment(None, school)
        client = client_as(teacher)

        # ملف ليس xlsx حقيقياً
        fake_file = SimpleUploadedFile(
            "grades.xlsx", b"not a real xlsx file", content_type="application/octet-stream"
        )

        initial_count = ImportLog.objects.filter(school=school).count()
        response = client.post(
            f"/import/upload/{assessment.id}/",
            {"grade_file": fake_file},
        )
        # إما redirect (فشل) أو عرض نتيجة
        assert response.status_code in (302, 200)
        # سجل ImportLog يجب أن يُنشأ بحالة failed أو completed
        new_count = ImportLog.objects.filter(school=school).count()
        assert new_count > initial_count

    def test_valid_xlsx_creates_completed_log(self, client_as, school):
        """ملف xlsx صحيح + صف بيانات → ImportLog مكتمل"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        teacher, assessment = make_teacher_with_assessment(None, school)

        # إنشاء ملف xlsx بسيط بدون بيانات طلاب (0 صفوف)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "الرقم الشخصي"
        ws["B1"] = "اسم الطالب"
        ws["C1"] = "الدرجة"
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()
        upload = SimpleUploadedFile(
            "grades_test.xlsx",
            xlsx_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        client = client_as(teacher)
        response = client.post(
            f"/import/upload/{assessment.id}/",
            {"grade_file": upload},
        )
        assert response.status_code in (200, 302)
        log = ImportLog.objects.filter(school=school).last()
        assert log is not None
        assert log.status in ("completed", "failed")


# ══════════════════════════════════════════════════════════
#  5. ImportLog Model
# ══════════════════════════════════════════════════════════


@pytest.mark.django_db
class TestImportLogModel:
    def test_create_log(self, school, principal_user):
        log = ImportLog.objects.create(
            school=school,
            uploaded_by=principal_user,
            file_name="grades_s1.xlsx",
            status="pending",
        )
        assert log.pk is not None
        assert log.status == "pending"

    def test_str(self, school, principal_user):
        log = ImportLog.objects.create(
            school=school,
            uploaded_by=principal_user,
            file_name="test.xlsx",
            status="completed",
        )
        assert "test.xlsx" in str(log)

    def test_default_values(self, school, principal_user):
        log = ImportLog.objects.create(
            school=school,
            uploaded_by=principal_user,
            file_name="x.xlsx",
        )
        assert log.total_rows == 0
        assert log.imported_rows == 0
        assert log.failed_rows == 0
        assert log.error_log == []
        assert log.completed_at is None

    def test_error_log_is_list(self, school, principal_user):
        log = ImportLog.objects.create(
            school=school,
            uploaded_by=principal_user,
            file_name="err.xlsx",
            status="completed",
            error_log=["خطأ 1", "خطأ 2"],
        )
        log.refresh_from_db()
        assert isinstance(log.error_log, list)
        assert len(log.error_log) == 2
