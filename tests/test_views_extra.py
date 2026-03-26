"""
tests/test_views_extra.py
Combined tests for clinic/views.py, library/views.py, staging/views.py
"""

import io
from datetime import date, timedelta
from decimal import Decimal

import pytest
from django.urls import reverse
from django.utils import timezone

from assessments.models import Assessment, AssessmentPackage, SubjectClassSetup
from clinic.models import ClinicVisit, HealthRecord
from library.models import BookBorrowing
from operations.models import Subject
from staging.models import ImportLog
from tests.conftest import (
    LibraryBookFactory,
    MembershipFactory,
    RoleFactory,
    UserFactory,
)

# ══════════════════════════════════════════════════════════════
#  CLINIC VIEWS
# ══════════════════════════════════════════════════════════════


class TestClinicDashboard:
    @pytest.mark.django_db
    def test_nurse_can_access_dashboard(self, client_as, nurse_user):
        c = client_as(nurse_user)
        resp = c.get(reverse("clinic:dashboard"))
        assert resp.status_code == 200

    @pytest.mark.django_db
    def test_student_cannot_access_dashboard(self, client_as, student_user):
        c = client_as(student_user)
        resp = c.get(reverse("clinic:dashboard"))
        assert resp.status_code == 403

    @pytest.mark.django_db
    def test_teacher_cannot_access_dashboard(self, client_as, teacher_user):
        c = client_as(teacher_user)
        resp = c.get(reverse("clinic:dashboard"))
        assert resp.status_code == 403

    @pytest.mark.django_db
    def test_anonymous_redirected(self, client):
        resp = client.get(reverse("clinic:dashboard"))
        assert resp.status_code == 302

    @pytest.mark.django_db
    def test_dashboard_shows_visits(self, client_as, nurse_user, clinic_visit):
        c = client_as(nurse_user)
        resp = c.get(reverse("clinic:dashboard"))
        assert resp.status_code == 200


class TestStudentHealthRecord:
    @pytest.mark.django_db
    def test_get_health_record(self, client_as, nurse_user, student_user, health_record):
        c = client_as(nurse_user)
        url = reverse("clinic:health_record", args=[student_user.pk])
        resp = c.get(url)
        assert resp.status_code == 200

    @pytest.mark.django_db
    def test_get_health_record_creates_if_missing(self, client_as, nurse_user, student_user):
        """If no HealthRecord exists, view creates one."""
        c = client_as(nurse_user)
        assert not HealthRecord.objects.filter(student=student_user).exists()
        url = reverse("clinic:health_record", args=[student_user.pk])
        resp = c.get(url)
        assert resp.status_code == 200
        assert HealthRecord.objects.filter(student=student_user).exists()

    @pytest.mark.django_db
    def test_post_health_record(self, client_as, nurse_user, student_user, health_record):
        c = client_as(nurse_user)
        url = reverse("clinic:health_record", args=[student_user.pk])
        data = {
            "blood_type": "A+",
            "emergency_contact_name": "Parent Name",
            "emergency_contact_phone": "12345",
            "allergies": "nuts",
            "chronic_diseases": "asthma",
            "medications": "inhaler",
        }
        resp = c.post(url, data)
        assert resp.status_code == 302  # redirect after save

    @pytest.mark.django_db
    def test_student_forbidden(self, client_as, student_user):
        c = client_as(student_user)
        url = reverse("clinic:health_record", args=[student_user.pk])
        resp = c.get(url)
        assert resp.status_code == 403

    @pytest.mark.django_db
    def test_nonexistent_student_404(self, client_as, nurse_user):
        import uuid

        c = client_as(nurse_user)
        url = reverse("clinic:health_record", args=[uuid.uuid4()])
        resp = c.get(url)
        assert resp.status_code == 404


class TestRecordVisit:
    @pytest.mark.django_db
    def test_get_record_visit_form(self, client_as, nurse_user):
        c = client_as(nurse_user)
        resp = c.get(reverse("clinic:record_visit"))
        assert resp.status_code == 200

    @pytest.mark.django_db
    def test_get_record_visit_with_student_id(self, client_as, nurse_user, student_user):
        c = client_as(nurse_user)
        url = reverse("clinic:record_visit_student", args=[student_user.pk])
        resp = c.get(url)
        assert resp.status_code == 200

    @pytest.mark.django_db
    def test_post_record_visit(self, client_as, nurse_user, student_user, school):
        c = client_as(nurse_user)
        data = {
            "student_id": str(student_user.pk),
            "reason": "headache",
            "symptoms": "pain",
            "temperature": "37.5",
            "treatment": "rest",
        }
        resp = c.post(reverse("clinic:record_visit"), data)
        assert resp.status_code == 302
        assert ClinicVisit.objects.filter(student=student_user).exists()

    @pytest.mark.django_db
    def test_post_record_visit_sent_home(self, client_as, nurse_user, student_user, school):
        c = client_as(nurse_user)
        data = {
            "student_id": str(student_user.pk),
            "reason": "fever",
            "symptoms": "high temp",
            "treatment": "sent home",
            "is_sent_home": "on",
        }
        # Just test the view works even if hub fails
        resp = c.post(reverse("clinic:record_visit"), data)
        assert resp.status_code == 302
        visit = ClinicVisit.objects.get(student=student_user)
        assert visit.is_sent_home is True

    @pytest.mark.django_db
    def test_post_record_visit_htmx(self, client_as, nurse_user, student_user, school):
        c = client_as(nurse_user)
        data = {
            "student_id": str(student_user.pk),
            "reason": "stomachache",
            "symptoms": "",
            "treatment": "",
        }
        resp = c.post(
            reverse("clinic:record_visit"),
            data,
            HTTP_HX_REQUEST="true",
        )
        assert resp.status_code == 200  # renders partial template

    @pytest.mark.django_db
    def test_teacher_forbidden(self, client_as, teacher_user):
        c = client_as(teacher_user)
        resp = c.get(reverse("clinic:record_visit"))
        assert resp.status_code == 403


class TestVisitsList:
    @pytest.mark.django_db
    def test_get_visits_list(self, client_as, nurse_user, clinic_visit):
        c = client_as(nurse_user)
        resp = c.get(reverse("clinic:visits_list"))
        assert resp.status_code == 200

    @pytest.mark.django_db
    def test_filter_by_date(self, client_as, nurse_user, clinic_visit):
        c = client_as(nurse_user)
        today = timezone.now().date().strftime("%Y-%m-%d")
        resp = c.get(reverse("clinic:visits_list"), {"date": today})
        assert resp.status_code == 200

    @pytest.mark.django_db
    def test_filter_by_invalid_date(self, client_as, nurse_user):
        c = client_as(nurse_user)
        resp = c.get(reverse("clinic:visits_list"), {"date": "not-a-date"})
        assert resp.status_code == 200  # gracefully ignores bad date

    @pytest.mark.django_db
    def test_filter_by_student_name(self, client_as, nurse_user, clinic_visit):
        c = client_as(nurse_user)
        resp = c.get(reverse("clinic:visits_list"), {"student": "طالب"})
        assert resp.status_code == 200

    @pytest.mark.django_db
    def test_forbidden_for_student(self, client_as, student_user):
        c = client_as(student_user)
        resp = c.get(reverse("clinic:visits_list"))
        assert resp.status_code == 403


class TestHealthStatistics:
    @pytest.mark.django_db
    def test_nurse_can_view_statistics(self, client_as, nurse_user):
        c = client_as(nurse_user)
        resp = c.get(reverse("clinic:statistics"))
        assert resp.status_code == 200

    @pytest.mark.django_db
    def test_student_forbidden(self, client_as, student_user):
        c = client_as(student_user)
        resp = c.get(reverse("clinic:statistics"))
        assert resp.status_code == 403

    @pytest.mark.django_db
    def test_statistics_with_data(self, client_as, nurse_user, health_record, clinic_visit):
        c = client_as(nurse_user)
        resp = c.get(reverse("clinic:statistics"))
        assert resp.status_code == 200


# ══════════════════════════════════════════════════════════════
#  LIBRARY VIEWS
# ══════════════════════════════════════════════════════════════


class TestLibraryDashboard:
    @pytest.mark.django_db
    def test_staff_can_access(self, client_as, teacher_user):
        c = client_as(teacher_user)
        resp = c.get(reverse("library:dashboard"))
        assert resp.status_code == 200

    @pytest.mark.django_db
    def test_librarian_can_access(self, client_as, librarian_user):
        c = client_as(librarian_user)
        resp = c.get(reverse("library:dashboard"))
        assert resp.status_code == 200

    @pytest.mark.django_db
    def test_student_can_access(self, client_as, student_user):
        """الطالب ضمن LIBRARY_VIEW — يمكنه تصفح المكتبة"""
        c = client_as(student_user)
        resp = c.get(reverse("library:dashboard"))
        assert resp.status_code == 200

    @pytest.mark.django_db
    def test_anonymous_redirected(self, client):
        resp = client.get(reverse("library:dashboard"))
        assert resp.status_code == 302

    @pytest.mark.django_db
    def test_dashboard_with_data(self, client_as, librarian_user, library_book, book_borrowing):
        c = client_as(librarian_user)
        resp = c.get(reverse("library:dashboard"))
        assert resp.status_code == 200


class TestBookList:
    @pytest.mark.django_db
    def test_book_list_view(self, client_as, teacher_user, library_book):
        c = client_as(teacher_user)
        resp = c.get(reverse("library:book_list"))
        assert resp.status_code == 200

    @pytest.mark.django_db
    def test_book_list_search(self, client_as, teacher_user, library_book):
        c = client_as(teacher_user)
        resp = c.get(reverse("library:book_list"), {"q": "كتاب"})
        assert resp.status_code == 200

    @pytest.mark.django_db
    def test_book_list_pagination(self, client_as, teacher_user, school):
        # Create 30 books to trigger pagination
        for i in range(30):
            LibraryBookFactory(school=school, title=f"Book {i}")
        c = client_as(teacher_user)
        resp = c.get(reverse("library:book_list"), {"page": "2"})
        assert resp.status_code == 200

    @pytest.mark.django_db
    def test_student_can_browse_books(self, client_as, student_user):
        """الطالب ضمن LIBRARY_VIEW — يمكنه تصفح الكتب"""
        c = client_as(student_user)
        resp = c.get(reverse("library:book_list"))
        assert resp.status_code == 200


class TestBorrowBook:
    @pytest.mark.django_db
    def test_get_borrow_form(self, client_as, librarian_user):
        c = client_as(librarian_user)
        resp = c.get(reverse("library:borrow_book"))
        assert resp.status_code == 200

    @pytest.mark.django_db
    def test_post_borrow_book_success(self, client_as, librarian_user, library_book, student_user):
        c = client_as(librarian_user)
        data = {
            "book_id": str(library_book.pk),
            "user_id": str(student_user.pk),
            "due_date": (date.today() + timedelta(days=14)).isoformat(),
        }
        resp = c.post(reverse("library:borrow_book"), data)
        assert resp.status_code == 302
        assert BookBorrowing.objects.filter(book=library_book, user=student_user).exists()
        library_book.refresh_from_db()
        assert library_book.available_qty == 2  # was 3, now 2

    @pytest.mark.django_db
    def test_post_borrow_book_unavailable(
        self, client_as, librarian_user, library_book, student_user
    ):
        library_book.available_qty = 0
        library_book.save()
        c = client_as(librarian_user)
        data = {
            "book_id": str(library_book.pk),
            "user_id": str(student_user.pk),
            "due_date": (date.today() + timedelta(days=14)).isoformat(),
        }
        resp = c.post(reverse("library:borrow_book"), data)
        assert resp.status_code == 302
        assert not BookBorrowing.objects.filter(book=library_book, user=student_user).exists()

    @pytest.mark.django_db
    def test_teacher_forbidden(self, client_as, teacher_user):
        """teacher has staff_required but not librarian_required"""
        c = client_as(teacher_user)
        resp = c.get(reverse("library:borrow_book"))
        assert resp.status_code == 403


class TestReturnBook:
    @pytest.mark.django_db
    def test_return_book_success(self, client_as, librarian_user, book_borrowing):
        c = client_as(librarian_user)
        old_qty = book_borrowing.book.available_qty
        url = reverse("library:return_book", args=[book_borrowing.pk])
        resp = c.get(url)
        assert resp.status_code == 302
        book_borrowing.refresh_from_db()
        assert book_borrowing.status == "RETURNED"
        book_borrowing.book.refresh_from_db()
        assert book_borrowing.book.available_qty == old_qty + 1

    @pytest.mark.django_db
    def test_return_already_returned(self, client_as, librarian_user, book_borrowing):
        book_borrowing.status = "RETURNED"
        book_borrowing.save()
        c = client_as(librarian_user)
        old_qty = book_borrowing.book.available_qty
        url = reverse("library:return_book", args=[book_borrowing.pk])
        resp = c.get(url)
        assert resp.status_code == 302
        book_borrowing.book.refresh_from_db()
        assert book_borrowing.book.available_qty == old_qty  # no change

    @pytest.mark.django_db
    def test_teacher_forbidden(self, client_as, teacher_user, book_borrowing):
        c = client_as(teacher_user)
        url = reverse("library:return_book", args=[book_borrowing.pk])
        resp = c.get(url)
        assert resp.status_code == 403


# ══════════════════════════════════════════════════════════════
#  STAGING VIEWS
# ══════════════════════════════════════════════════════════════


@pytest.fixture
def subject(db, school):
    return Subject.objects.create(school=school, name_ar="رياضيات", code="MATH")


@pytest.fixture
def subject_setup(db, school, teacher_user, class_group, subject):
    return SubjectClassSetup.objects.create(
        school=school,
        subject=subject,
        class_group=class_group,
        teacher=teacher_user,
        academic_year="2025-2026",
    )


@pytest.fixture
def assessment_package(db, school, subject_setup):
    return AssessmentPackage.objects.create(
        setup=subject_setup,
        school=school,
        package_type="P1",
        semester="S1",
        weight=Decimal("50"),
    )


@pytest.fixture
def assessment(db, school, assessment_package):
    return Assessment.objects.create(
        package=assessment_package,
        school=school,
        title="اختبار نصفي",
        assessment_type="exam",
        max_grade=Decimal("10"),
        status="published",
    )


class TestImportGradesSelect:
    @pytest.mark.django_db
    def test_admin_can_access(self, client_as, principal_user):
        c = client_as(principal_user)
        resp = c.get(reverse("import_grades_select"))
        assert resp.status_code == 200

    @pytest.mark.django_db
    def test_teacher_can_access(self, client_as, teacher_user, assessment):
        c = client_as(teacher_user)
        resp = c.get(reverse("import_grades_select"))
        assert resp.status_code == 200

    @pytest.mark.django_db
    def test_student_forbidden(self, client_as, student_user):
        c = client_as(student_user)
        resp = c.get(reverse("import_grades_select"))
        assert resp.status_code == 403

    @pytest.mark.django_db
    def test_with_year_param(self, client_as, principal_user):
        c = client_as(principal_user)
        resp = c.get(reverse("import_grades_select"), {"year": "2025-2026"})
        assert resp.status_code == 200

    @pytest.mark.django_db
    def test_teacher_sees_own_assessments(self, client_as, teacher_user, assessment):
        """Teacher should see assessments where they are the setup teacher."""
        c = client_as(teacher_user)
        resp = c.get(reverse("import_grades_select"))
        assert resp.status_code == 200


class TestDownloadGradeTemplate:
    @pytest.mark.django_db
    def test_admin_downloads_template(
        self, client_as, principal_user, assessment, enrolled_student
    ):
        c = client_as(principal_user)
        url = reverse("download_grade_template", args=[assessment.pk])
        resp = c.get(url)
        assert resp.status_code == 200
        assert "spreadsheetml" in resp["Content-Type"]
        assert resp.has_header("Content-Disposition")

    @pytest.mark.django_db
    def test_teacher_downloads_own(self, client_as, teacher_user, assessment, enrolled_student):
        c = client_as(teacher_user)
        url = reverse("download_grade_template", args=[assessment.pk])
        resp = c.get(url)
        assert resp.status_code == 200

    @pytest.mark.django_db
    def test_other_teacher_forbidden(self, client_as, school, assessment):
        """A teacher not assigned to the assessment should get 403."""
        role = RoleFactory(school=school, name="teacher")
        other = UserFactory(full_name="Other Teacher")
        MembershipFactory(user=other, school=school, role=role)
        c = client_as(other)
        url = reverse("download_grade_template", args=[assessment.pk])
        resp = c.get(url)
        assert resp.status_code == 403

    @pytest.mark.django_db
    def test_student_forbidden(self, client_as, student_user, assessment):
        c = client_as(student_user)
        url = reverse("download_grade_template", args=[assessment.pk])
        resp = c.get(url)
        assert resp.status_code == 403


class TestUploadGradeFile:
    @pytest.mark.django_db
    def test_get_redirects(self, client_as, principal_user, assessment):
        c = client_as(principal_user)
        url = reverse("upload_grade_file", args=[assessment.pk])
        resp = c.get(url)
        assert resp.status_code == 302

    @pytest.mark.django_db
    def test_no_file_uploaded(self, client_as, principal_user, assessment):
        c = client_as(principal_user)
        url = reverse("upload_grade_file", args=[assessment.pk])
        resp = c.post(url)
        assert resp.status_code == 302  # redirect with error message

    @pytest.mark.django_db
    def test_upload_valid_excel(
        self, client_as, principal_user, assessment, enrolled_student, student_user
    ):
        """Upload a valid Excel file with grade data."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        # Header row at row 6
        ws.cell(6, 1, "الرقم الوطني")
        ws.cell(6, 2, "اسم الطالب")
        ws.cell(6, 3, "الدرجة")
        ws.cell(6, 4, "غائب (1/0)")
        ws.cell(6, 5, "ملاحظة")
        # Student data row at row 7
        ws.cell(7, 1, student_user.national_id)
        ws.cell(7, 2, student_user.full_name)
        ws.cell(7, 3, 8)
        ws.cell(7, 4, 0)
        ws.cell(7, 5, "")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "grades.xlsx"

        c = client_as(principal_user)
        url = reverse("upload_grade_file", args=[assessment.pk])
        resp = c.post(url, {"grade_file": buf})
        assert resp.status_code == 200  # renders result template
        log = ImportLog.objects.filter(school=assessment.school).first()
        assert log is not None
        assert log.imported_rows == 1

    @pytest.mark.django_db
    def test_upload_invalid_national_id(
        self, client_as, principal_user, assessment, enrolled_student
    ):
        """Upload with a national ID that doesn't exist."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(6, 1, "الرقم الوطني")
        ws.cell(7, 1, "INVALID_ID_999")
        ws.cell(7, 2, "Ghost Student")
        ws.cell(7, 3, 5)
        ws.cell(7, 4, 0)
        ws.cell(7, 5, "")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "grades.xlsx"

        c = client_as(principal_user)
        url = reverse("upload_grade_file", args=[assessment.pk])
        resp = c.post(url, {"grade_file": buf})
        assert resp.status_code == 200
        log = ImportLog.objects.filter(school=assessment.school).first()
        assert log.failed_rows >= 1

    @pytest.mark.django_db
    def test_upload_grade_out_of_range(
        self, client_as, principal_user, assessment, enrolled_student, student_user
    ):
        """Grade exceeding max_grade should be flagged as error."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(6, 1, "الرقم الوطني")
        ws.cell(7, 1, student_user.national_id)
        ws.cell(7, 2, student_user.full_name)
        ws.cell(7, 3, 999)  # way above max_grade=10
        ws.cell(7, 4, 0)
        ws.cell(7, 5, "")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "grades.xlsx"

        c = client_as(principal_user)
        url = reverse("upload_grade_file", args=[assessment.pk])
        resp = c.post(url, {"grade_file": buf})
        assert resp.status_code == 200
        log = ImportLog.objects.filter(school=assessment.school).first()
        assert log.failed_rows >= 1

    @pytest.mark.django_db
    def test_upload_corrupt_file(self, client_as, principal_user, assessment):
        """Uploading a non-Excel file should fail gracefully."""
        buf = io.BytesIO(b"this is not an excel file")
        buf.name = "bad.xlsx"

        c = client_as(principal_user)
        url = reverse("upload_grade_file", args=[assessment.pk])
        resp = c.post(url, {"grade_file": buf})
        assert resp.status_code == 302  # redirect with error msg
        log = ImportLog.objects.filter(school=assessment.school).first()
        assert log.status == "failed"

    @pytest.mark.django_db
    def test_other_teacher_forbidden(self, client_as, school, assessment):
        role = RoleFactory(school=school, name="teacher")
        other = UserFactory(full_name="Other Teacher")
        MembershipFactory(user=other, school=school, role=role)

        buf = io.BytesIO(b"dummy")
        buf.name = "grades.xlsx"

        c = client_as(other)
        url = reverse("upload_grade_file", args=[assessment.pk])
        resp = c.post(url, {"grade_file": buf})
        assert resp.status_code == 403

    @pytest.mark.django_db
    def test_upload_absent_student(
        self, client_as, principal_user, assessment, enrolled_student, student_user
    ):
        """Student marked as absent (1) should be imported successfully."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(6, 1, "الرقم الوطني")
        ws.cell(7, 1, student_user.national_id)
        ws.cell(7, 2, student_user.full_name)
        ws.cell(7, 3, "")
        ws.cell(7, 4, 1)  # absent
        ws.cell(7, 5, "غائب")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "grades.xlsx"

        c = client_as(principal_user)
        url = reverse("upload_grade_file", args=[assessment.pk])
        resp = c.post(url, {"grade_file": buf})
        assert resp.status_code == 200


class TestImportLogList:
    @pytest.mark.django_db
    def test_admin_can_view_logs(self, client_as, principal_user, school):
        ImportLog.objects.create(school=school, file_name="test.xlsx", status="completed")
        c = client_as(principal_user)
        resp = c.get(reverse("import_log_list"))
        assert resp.status_code == 200

    @pytest.mark.django_db
    def test_teacher_forbidden(self, client_as, teacher_user):
        c = client_as(teacher_user)
        resp = c.get(reverse("import_log_list"))
        assert resp.status_code == 403

    @pytest.mark.django_db
    def test_student_forbidden(self, client_as, student_user):
        c = client_as(student_user)
        resp = c.get(reverse("import_log_list"))
        assert resp.status_code == 403
