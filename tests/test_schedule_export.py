"""[SCHEDULE] الجدولُ يخرج من الشاشة: ورقةً وملفَّ PDF ومصنَّفَ Excel.

الورقةُ المعلَّقة في المدرسة تُنسخ وتُرسَل وتُؤرشَف، وكان مخرجُها الوحيدُ
زرَّ طباعةٍ يفتح حوار المتصفّح. فأُضيف مخرجان يقرآن السياق نفسه — فما يخرج
ملفّاً هو ما يُرى على الشاشة، لا جدولٌ ثانٍ يُبنى على حدة.

وعمودُ القسم أُضيف إلى يمين أسماء المعلّمين: خانةٌ واحدةٌ ممتدّةٌ على معلّمي
القسم، لا اسمٌ يتكرّر في ثلاثةٍ وسبعين سطراً.
"""

from datetime import time

import pytest
from django.test import Client
from django.urls import reverse

from core.models import CustomUser
from core.models.access import Membership, Role
from operations.models import ScheduleSlot, Subject


@pytest.fixture
def teaching_school(db, school, class_group):
    """معلّمان في قسمين: الشرعيةُ بواحد، والعربيةُ بواحد."""
    made = []
    for index, (name, subject_name) in enumerate(
        (
            ("معلّم الشرعية", "التربية الإسلامية"),
            ("معلّم العربية", "اللغة العربية"),
        )
    ):
        user = CustomUser.objects.create(national_id=f"2864400{index:04d}", full_name=name)
        role, _ = Role.objects.get_or_create(school=school, name="teacher")
        Membership.objects.create(user=user, school=school, role=role)
        ScheduleSlot.objects.create(
            school=school,
            class_group=class_group,
            teacher=user,
            subject=Subject.objects.create(school=school, name_ar=subject_name),
            day_of_week=index,
            period_number=1,
            start_time=time(7, 30),
            end_time=time(8, 15),
            academic_year=class_group.academic_year,
        )
        made.append(user)
    return made


@pytest.fixture
def principal(db, school):
    user = CustomUser.objects.create(national_id="28644099999", full_name="مدير")
    role, _ = Role.objects.get_or_create(school=school, name="principal")
    Membership.objects.create(user=user, school=school, role=role)
    return user


def _get(user, name, query=""):
    client = Client()
    client.force_login(user)
    return client.get(reverse(name) + query, HTTP_HOST="localhost")


# ── عمودُ القسم في الورقة ────────────────────────────────────────────


def test_the_sheet_names_each_department_beside_its_teachers(db, principal, teaching_school):
    """كان القسمُ لوناً في الخلفيّة وعنواناً في `title` — لا يُقرأ على ورقٍ مطبوع."""
    body = _get(principal, "schedule_print", "?view=all_teachers").content.decode()

    assert "الشرعية" in body
    assert "اللغة العربية" in body
    assert 'class="m-dept"' in body


def test_the_department_cell_carries_a_rowspan(db, principal, teaching_school, school, class_group):
    """معلّمان في قسمٍ واحد ⇐ خانةٌ واحدةٌ ممتدّةٌ على سطرين."""
    second = CustomUser.objects.create(national_id="28644012345", full_name="معلّم عربية ثانٍ")
    role, _ = Role.objects.get_or_create(school=school, name="teacher")
    Membership.objects.create(user=second, school=school, role=role)
    ScheduleSlot.objects.create(
        school=school,
        class_group=class_group,
        teacher=second,
        subject=Subject.objects.create(school=school, name_ar="اللغة العربية"),
        day_of_week=3,
        period_number=2,
        start_time=time(8, 15),
        end_time=time(9, 0),
        academic_year=class_group.academic_year,
    )

    body = _get(principal, "schedule_print", "?view=all_teachers").content.decode()

    assert 'rowspan="2"' in body


# ── مخرجا التصدير ────────────────────────────────────────────────────


def test_excel_export_returns_a_workbook(db, principal, teaching_school):
    resp = _get(principal, "schedule_export_excel", "?view=all_teachers&paper=a3")

    assert resp.status_code == 200
    assert "spreadsheetml.sheet" in resp["Content-Type"]
    assert resp["Content-Disposition"].startswith("attachment;")
    # الاسمُ العربيّ يصل بترميز RFC 5987 — والترويسةُ كلُّها ASCII، وإلّا
    # رمّزها Django بـRFC 2047 فضاعت كلمةُ `attachment` على المتصفّح.
    assert resp["Content-Disposition"].isascii()
    assert "filename*=UTF-8''" in resp["Content-Disposition"]


def test_the_workbook_holds_the_department_column(db, principal, teaching_school):
    from io import BytesIO

    import openpyxl

    resp = _get(principal, "schedule_export_excel", "?view=all_teachers&paper=a3")
    sheet = openpyxl.load_workbook(BytesIO(resp.content)).active

    assert sheet.cell(row=4, column=1).value == "القسم"
    assert sheet.cell(row=4, column=2).value == "المعلّم"
    assert sheet.cell(row=6, column=1).value == "الشرعية"
    # الورقةُ عربيّةٌ: العمودُ الأوّل أقصى اليمين.
    assert sheet.sheet_view.rightToLeft


def test_pdf_export_returns_a_document(db, principal, teaching_school):
    """مولّدُ PDF يتدهور إلى 503 حين تغيب مكتبته — والمسارُ لا ينهار."""
    resp = _get(principal, "schedule_export_pdf", "?view=all_teachers&paper=a3")

    assert resp.status_code in (200, 503)
    if resp.status_code == 200:
        assert resp["Content-Type"] == "application/pdf"
        assert resp.content.startswith(b"%PDF")


def test_a_teacher_exports_only_their_own_schedule(db, teaching_school):
    """قيدُ التصفّح نفسُه على المخرجين — وإلّا فُتح بابٌ من خلف الورقة."""
    from io import BytesIO

    import openpyxl

    teacher, colleague = teaching_school
    resp = Client()
    resp.force_login(teacher)
    response = resp.get(
        reverse("schedule_export_excel") + f"?view=teacher&teacher={colleague.id}",
        HTTP_HOST="localhost",
    )

    assert response.status_code == 200
    sheet = openpyxl.load_workbook(BytesIO(response.content)).active
    printed = "\n".join(str(cell.value or "") for row in sheet.iter_rows() for cell in row)
    assert colleague.full_name not in printed


# ══════════════════════════════════════════════════════════════════════
#  جدولُ المعلّم هو ورقةُ المدرسة نفسُها
# ══════════════════════════════════════════════════════════════════════


def test_the_teacher_schedule_opens_the_school_schedule_page(client_as, school, principal_user):
    """شكلان لشيءٍ واحدٍ يُشتّتان (قرارُ المستخدم 2026-09-06) — فصفحةٌ واحدة.

    وهي عينُها التي تفتحها القائمةُ المنسدلة حين يُختار معلّمٌ بعينه: قالبٌ
    واحدٌ وتنسيقٌ واحدٌ وبياناتٌ واحدة.
    """
    from django.urls import reverse

    response = client_as(principal_user).get(
        reverse("teacher_weekly_view", args=[principal_user.id])
    )

    assert response.status_code == 302
    assert reverse("weekly_schedule") in response.url
    assert f"teacher={principal_user.id}" in response.url
    assert "view=teacher" in response.url


def test_a_teacher_reaching_the_page_sees_their_own_schedule(client_as, school, teacher_user):
    """ومن لا يتصفّح جداولَ غيره يرى جدولَه هو، مهما كُتب في الرابط."""
    from django.urls import reverse

    response = client_as(teacher_user).get(
        reverse("weekly_schedule"),
        {"view": "teacher", "teacher": "00000000-0000-0000-0000-000000000001"},
    )

    assert response.status_code == 200
    assert teacher_user.full_name in response.content.decode()
