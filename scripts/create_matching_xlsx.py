import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

wb = Workbook()

BLUE_HEADER = PatternFill('solid', fgColor='1F4E79')
RED_HEADER = PatternFill('solid', fgColor='C0392B')
BLUE2_HEADER = PatternFill('solid', fgColor='2E86C1')
WHITE_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
GREEN_FILL = PatternFill('solid', fgColor='C6EFCE')
YELLOW_FILL = PatternFill('solid', fgColor='FFEB9C')
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
LIGHT_GRAY = PatternFill('solid', fgColor='F2F2F2')
BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
AR_ALIGN = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)
AR_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True, readingOrder=2)
CELL_FONT = Font(name='Arial', size=10)

db_names = [
    'ابراهيم سليمان طه حمد',
    'ابراهيم محمد محمود الزريقات',
    'احمد المنصف الحاج قاسم',
    'احمد بكر محمد الزبط',
    'احمد جبر جبر خلف',
    'احمد جعفر عبد الفتاح شاهين',
    'احمد رمضان خطاب ابراهيم حامد',
    'احمد عبدالعزيز جامع مرسى',
    'احمد محمد أوغلو',
    'أكرم رابح قمودي',
    'البشير بوحلاب سودان',
    'السيد محمدي رفاعي عبدالوهاب الرفاعي',
    'حسام محمود مبروك غانم',
    'خليفه صالح ظاهر عودات',
    'سامر غازي مصطفى محمد',
    'سامر نصر سليمان جديع',
    'سفيان احمد محمد مسيف',
    'طارق باسم مصطفى شملاوي',
    'عامر محمد نجار',
    'عبدالباسط عبدالسلام الجاسم',
    'عبدالرحمن فيصل اسماعيل راجه',
    'عبدالرحمن محمد عبدالله لطف الله الاحزم',
    'عبدالله الرمضان',
    'عبدالله حسين مفلح نوفل',
    'عبدالله خالد كامل محمود عبدربه',
    'عثمان عبدالرحمن فاروسي',
    'عدنان بركات عدنان المصطفى',
    'عربى السيد يوسف السيد رجب',
    'عزام احمد يوسف الزعبى',
    'علاء محمد عبد الهادي القضاه',
    'على ضيف الله حمد على',
    'علي ضيف الله خليل خريسات',
    'علي محمد محمد دار ناصر',
    'علي مصطفى الدروبى',
    'عماد يحيى محمد العبسي نوح',
    'عمادالدين محمد الحبشى قاسم',
    'عمر حسن عقله بني عطا',
    'عمر مصطفى محمد العباس',
    'عمرو محمد حمدان حمايده',
    'عمرو محمد خليل الشحروري',
    'ليث حامد محمد السعودي',
    'مجدى محمد على احمد قنديل',
    'محمد احمد حسن عنانبه',
    'محمد سلام سليمان حسين',
    'محمد صبرى محمود درويش',
    'محمد عبدالعزيز يونس عدوان',
    'محمد عبدالله عارف العجلوني',
    'محمد عبدالوهاب عبدالبديع عويس',
    'محمد فرحان سالم النوايسه',
    'محمد فلاح صالح درادكه',
    'محمود عبد المهدى عبد القادر الاسطه',
    'محمود ماجد يوسف الجرادات',
    'مرتضي امين ابوالبشر عبدالله',
    'مشل مرفى سعدى الرمالى الشمرى',
    'مصطفى عمر حسين النزهاوى',
    'معز بن احمد السعداوي',
    'منير رافع شتيوي شتيات',
    'مؤيد احمد محمد المومني',
    'نادر جمعه عثمان حنفيه',
    'نادر على لطفى محمد لطفى',
    'ناصر فايز مناحى سعد الهاجرى',
    'هاني محمد زغلول طه حسن',
    'وجدي بن محمد بن عمارة يوسفي',
    'وليد عبد اللطيف',
    'ياسر حجى شلبى احمد',
    'يوسف جميل سليمان العبدالله',
]

rows = [
    ('فيصل جليل الرويلي', '', '0%', 'غير موجود في المنصة'),
    ('محمد فرحان النوايسة', 'محمد فرحان سالم النوايسه', '75%', ''),
    ('ابراهيم محمد زريقات', 'ابراهيم محمد محمود الزريقات', '50%', ''),
    ('أحمد جبر خلف', 'احمد جبر جبر خلف', '100%', ''),
    ('ابراهيم سليمان حمد', 'ابراهيم سليمان طه حمد', '75%', ''),
    ('عماد محمد قاسم', 'عمادالدين محمد الحبشى قاسم', '50%', 'تحقق من التطابق'),
    ('عماد العبسي', 'عماد يحيى محمد العبسي نوح', '40%', ''),
    ('محمود الأسطة', 'محمود عبد المهدى عبد القادر الاسطه', '40%', 'اسم مختصر جداً'),
    ('محمود سعد', '', '0%', 'غير موجود في المنصة'),
    ('احمد عبدالعزيز جامع', 'احمد عبدالعزيز جامع مرسى', '75%', ''),
    ('خليفة صالح عودات', 'خليفه صالح ظاهر عودات', '75%', ''),
    ('عربي السيد رجب', 'عربى السيد يوسف السيد رجب', '75%', ''),
    ('أحمد أغلو', 'احمد محمد أوغلو', '50%', 'مدرس لغة عربية'),
    ('ياسر حجي شلبي', 'ياسر حجى شلبى احمد', '50%', ''),
    ('عزام احمد الزعبي', 'عزام احمد يوسف الزعبى', '75%', ''),
    ('نادر جمعة حنفية', 'نادر جمعه عثمان حنفيه', '75%', ''),
    ('عامر النجار', 'عامر محمد نجار', '50%', ''),
    ('أحمد بكر', 'احمد بكر محمد الزبط', '50%', ''),
    ('محمد العرامين', '', '0%', 'غير موجود في المنصة'),
    ('نادر على لطفى', 'نادر على لطفى محمد لطفى', '75%', ''),
    ('عمرو حمايدة', 'عمرو محمد حمدان حمايده', '50%', ''),
    ('محمد صبرى درويش', 'محمد صبرى محمود درويش', '75%', ''),
    ('مصطفى عمر النزهاوى', 'مصطفى عمر حسين النزهاوى', '75%', ''),
    ('مجدى محمد قنديل', 'مجدى محمد على احمد قنديل', '60%', ''),
    ('أكرم القمودي', 'أكرم رابح قمودي', '50%', ''),
    ('سفيان أحمد مسيف', 'سفيان احمد محمد مسيف', '75%', ''),
    ('مرتضى أمين', 'مرتضي امين ابوالبشر عبدالله', '50%', ''),
    ('عبدالرحمن رجب', '', '0%', 'يحتاج تحديد يدوي - هل هو راجه أو الاحزم؟'),
    ('احمد محمد إبراهيم', '', '0%', 'منسق الأحياء - غير موجود كمعلم بالمنصة'),
    ('عدنان المصطفى', 'عدنان بركات عدنان المصطفى', '67%', ''),
    ('محمد احمد عنانبه', 'محمد احمد حسن عنانبه', '75%', ''),
    ('علي خريسات', 'علي ضيف الله خليل خريسات', '40%', ''),
    ('حسن الصافي', '', '0%', 'غير موجود في المنصة'),
    ('وليد جمعه عبد اللطيف', 'وليد عبد اللطيف', '75%', ''),
    ('عطية محمود', '', '0%', 'غير موجود في المنصة'),
    ('أحمد شاهين', 'احمد جعفر عبد الفتاح شاهين', '50%', ''),
    ('عادل محمد نصر', '', '0%', 'غير موجود في المنصة'),
    ('حسام محمود غانم', 'حسام محمود مبروك غانم', '75%', ''),
    ('عبدالباسط عبدالسلام الجاسم', 'عبدالباسط عبدالسلام الجاسم', '100%', ''),
    ('محمود ماجد الجرادات', 'محمود ماجد يوسف الجرادات', '75%', ''),
    ('عبدالرحمن الأحزم', 'عبدالرحمن محمد عبدالله لطف الله الاحزم', '40%', ''),
    ('سلطان عواد', '', '0%', 'غير موجود في المنصة'),
    ('سامر غازي', 'سامر غازي مصطفى محمد', '50%', ''),
    ('طارق باسم شملاوي', 'طارق باسم مصطفى شملاوي', '75%', ''),
    ('محمد سلام سليمان', 'محمد سلام سليمان حسين', '75%', ''),
    ('محمد عبدالوهاب عويس', 'محمد عبدالوهاب عبدالبديع عويس', '75%', ''),
    ('علي محمد دار ناصر', 'علي محمد محمد دار ناصر', '100%', ''),
    ('محمد درادكة', 'محمد فلاح صالح درادكه', '50%', ''),
    ('عمر بني عطا', 'عمر حسن عقله بني عطا', '60%', ''),
    ('ليث السعودي', 'ليث حامد محمد السعودي', '50%', ''),
    ('البشير بو حلاب', 'البشير بوحلاب سودان', '50%', ''),
    ('إمام رشدي', '', '0%', 'غير موجود في المنصة'),
    ('علي مصطفى الدروبى', 'علي مصطفى الدروبى', '100%', ''),
    ('السيد محمدي رفاعي', 'السيد محمدي رفاعي عبدالوهاب الرفاعي', '60%', ''),
    ('محمد عبدالعزيز عدوان', 'محمد عبدالعزيز يونس عدوان', '75%', ''),
    ('مؤيد احمد المومني', 'مؤيد احمد محمد المومني', '75%', ''),
    ('محمد عبدالله العجلوني', 'محمد عبدالله عارف العجلوني', '75%', ''),
    ('علي ضيف', 'على ضيف الله حمد على', '50%', ''),
    ('عمر العباس', 'عمر مصطفى محمد العباس', '50%', ''),
    ('علاء القضاه', 'علاء محمد عبد الهادي القضاه', '50%', ''),
    ('احمد رمضان حامد', 'احمد رمضان خطاب ابراهيم حامد', '60%', ''),
    ('عبد الله نوفل', 'عبدالله حسين مفلح نوفل', '50%', ''),
    ('عثمان الفاروسي', 'عثمان عبدالرحمن فاروسي', '50%', ''),
    ('محمد اسماعيل السيد', '', '0%', 'غير موجود في المنصة'),
    ('سامر نصر جديع', 'سامر نصر سليمان جديع', '75%', ''),
    ('منير شتيات', 'منير رافع شتيوي شتيات', '50%', ''),
    ('أحمد الحاج', 'احمد المنصف الحاج قاسم', '50%', 'تحقق - قد يكون شخص مختلف'),
    ('يوسف يعقوب عوض', '', '0%', 'غير موجود في المنصة'),
    ('يوسف عثامنه', '', '0%', 'غير موجود في المنصة'),
    ('عبد الله الرمضان', 'عبدالله الرمضان', '100%', ''),
    ('بنجر الدوسري', '', '0%', 'غير موجود في المنصة'),
    ('وجدي يوسفي', 'وجدي بن محمد بن عمارة يوسفي', '50%', ''),
    ('معز السعداوي', 'معز بن احمد السعداوي', '50%', ''),
    ('عبد الله خالد', 'عبدالله خالد كامل محمود عبدربه', '50%', ''),
    ('ناصر الهاجري', 'ناصر فايز مناحى سعد الهاجرى', '40%', ''),
]

# ============ SHEET 1: المطابقة ============
ws1 = wb.active
ws1.title = 'المطابقة'
ws1.sheet_view.rightToLeft = True

headers = ['رقم', 'اسم المعلم في الجدول (PDF)', 'الاسم المقترح من المنصة', 'نسبة التطابق', 'التأكيد', 'الاسم الصحيح (اختر من القائمة)', 'ملاحظات']
col_widths = [6, 35, 42, 14, 18, 42, 35]

for c, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws1.cell(row=1, column=c, value=h)
    cell.font = WHITE_FONT
    cell.fill = BLUE_HEADER
    cell.alignment = AR_CENTER
    cell.border = BORDER
    ws1.column_dimensions[get_column_letter(c)].width = w

ws1.freeze_panes = 'A2'

for i, (pdf, db, pct, note) in enumerate(rows, 1):
    r = i + 1
    ws1.cell(row=r, column=1, value=i).alignment = AR_CENTER
    ws1.cell(row=r, column=1).font = CELL_FONT
    ws1.cell(row=r, column=1).border = BORDER

    ws1.cell(row=r, column=2, value=pdf).alignment = AR_ALIGN
    ws1.cell(row=r, column=2).font = CELL_FONT
    ws1.cell(row=r, column=2).border = BORDER

    ws1.cell(row=r, column=3, value=db).alignment = AR_ALIGN
    ws1.cell(row=r, column=3).font = CELL_FONT
    ws1.cell(row=r, column=3).border = BORDER

    pct_cell = ws1.cell(row=r, column=4, value=pct)
    pct_cell.alignment = AR_CENTER
    pct_cell.font = Font(name='Arial', size=10, bold=True)
    pct_cell.border = BORDER
    pct_val = int(pct.replace('%', '')) if '%' in pct else 0
    if pct_val >= 75:
        pct_cell.fill = GREEN_FILL
    elif pct_val >= 40:
        pct_cell.fill = YELLOW_FILL
    else:
        pct_cell.fill = RED_FILL

    for c in [5, 6]:
        ws1.cell(row=r, column=c).alignment = AR_ALIGN
        ws1.cell(row=r, column=c).font = CELL_FONT
        ws1.cell(row=r, column=c).border = BORDER

    ws1.cell(row=r, column=7, value=note).alignment = AR_ALIGN
    ws1.cell(row=r, column=7).font = CELL_FONT
    ws1.cell(row=r, column=7).border = BORDER

    if i % 2 == 0:
        for c in [1, 2, 3, 5, 6, 7]:
            ws1.cell(row=r, column=c).fill = LIGHT_GRAY

dv_confirm = DataValidation(type='list', formula1='"صحيح,خاطئ,يحتاج تعديل"', allow_blank=True)
dv_confirm.error = 'اختر من القائمة'
ws1.add_data_validation(dv_confirm)
dv_confirm.add(f'E2:E{len(rows)+1}')

dv_db = DataValidation(type='list', formula1=f"'كل_المعلمين'!$A$2:$A${len(db_names)+1}", allow_blank=True)
ws1.add_data_validation(dv_db)
dv_db.add(f'F2:F{len(rows)+1}')

# ============ SHEET 2: غير موجودين بالمنصة ============
ws2 = wb.create_sheet('غير_موجودين_بالمنصة')
ws2.sheet_view.rightToLeft = True

unmatched_pdf = [(i+1, r[0], r[3]) for i, r in enumerate(rows) if r[1] == '']
headers2 = ['رقم في الجدول', 'اسم المعلم في الجدول', 'ملاحظات', 'الإجراء المطلوب']
col_widths2 = [14, 35, 35, 25]

for c, (h, w) in enumerate(zip(headers2, col_widths2), 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = WHITE_FONT
    cell.fill = RED_HEADER
    cell.alignment = AR_CENTER
    cell.border = BORDER
    ws2.column_dimensions[get_column_letter(c)].width = w

ws2.freeze_panes = 'A2'
for i, (num, name, note) in enumerate(unmatched_pdf, 1):
    ws2.cell(row=i+1, column=1, value=num).alignment = AR_CENTER
    ws2.cell(row=i+1, column=1).font = CELL_FONT
    ws2.cell(row=i+1, column=1).border = BORDER
    ws2.cell(row=i+1, column=2, value=name).alignment = AR_ALIGN
    ws2.cell(row=i+1, column=2).font = CELL_FONT
    ws2.cell(row=i+1, column=2).border = BORDER
    ws2.cell(row=i+1, column=3, value=note).alignment = AR_ALIGN
    ws2.cell(row=i+1, column=3).font = CELL_FONT
    ws2.cell(row=i+1, column=3).border = BORDER
    ws2.cell(row=i+1, column=4).alignment = AR_ALIGN
    ws2.cell(row=i+1, column=4).font = CELL_FONT
    ws2.cell(row=i+1, column=4).border = BORDER

dv_action = DataValidation(type='list', formula1='"إضافة للمنصة,ربط باسم موجود,تجاهل"', allow_blank=True)
ws2.add_data_validation(dv_action)
dv_action.add(f'D2:D{len(unmatched_pdf)+1}')

# ============ SHEET 3: كل المعلمين (reference for dropdown) ============
ws3 = wb.create_sheet('كل_المعلمين')
ws3.sheet_view.rightToLeft = True

ws3.cell(row=1, column=1, value='اسم المعلم في المنصة').font = WHITE_FONT
ws3.cell(row=1, column=1).fill = BLUE2_HEADER
ws3.cell(row=1, column=1).alignment = AR_CENTER
ws3.cell(row=1, column=1).border = BORDER
ws3.column_dimensions['A'].width = 45

ws3.cell(row=1, column=2, value='حالة الربط بالجدول').font = WHITE_FONT
ws3.cell(row=1, column=2).fill = BLUE2_HEADER
ws3.cell(row=1, column=2).alignment = AR_CENTER
ws3.cell(row=1, column=2).border = BORDER
ws3.column_dimensions['B'].width = 25

ws3.freeze_panes = 'A2'

matched_db = set(r[1] for r in rows if r[1])
for i, name in enumerate(db_names, 1):
    ws3.cell(row=i+1, column=1, value=name).alignment = AR_ALIGN
    ws3.cell(row=i+1, column=1).font = CELL_FONT
    ws3.cell(row=i+1, column=1).border = BORDER

    status = 'مرتبط' if name in matched_db else 'بدون جدول'
    ws3.cell(row=i+1, column=2, value=status).alignment = AR_CENTER
    ws3.cell(row=i+1, column=2).font = CELL_FONT
    ws3.cell(row=i+1, column=2).border = BORDER
    if status == 'بدون جدول':
        ws3.cell(row=i+1, column=1).fill = YELLOW_FILL
        ws3.cell(row=i+1, column=2).fill = YELLOW_FILL

output = 'D:/shschool_mvp/teacher_matching.xlsx'
wb.save(output)
print(f'Done: {output}')
