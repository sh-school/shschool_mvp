"""
core/export_utils.py — أدوات تصدير موحّدة لكل المنصة
أسماء ملفات + هيدر + فوتر + توقيع المُصدِّر
"""

import uuid
from pathlib import Path

from django.conf import settings
from django.utils import timezone


def generate_export_filename(module: str, report_type: str, ext: str) -> str:
    """
    اسم ملف موحّد: SchoolOS_{module}_{type}_{YYYYMMDD}_{HHMMSS}_{6hex}.{ext}
    مثال: SchoolOS_students_list_20260403_143022_a7f3b2.xlsx
    """
    now = timezone.localtime()
    short_id = uuid.uuid4().hex[:6]
    date_str = now.strftime("%Y%m%d")
    time_str = now.strftime("%H%M%S")
    return f"SchoolOS_{module}_{report_type}_{date_str}_{time_str}_{short_id}.{ext}"


def get_export_context(request, title: str) -> dict:
    """
    context موحّد لجميع التصديرات (Excel + PDF):
    - school_name, school_logo_path
    - exported_by (الاسم), exporter_role (الدور)
    - export_date, export_time, export_datetime
    - title, academic_year
    """
    user = request.user
    school = user.get_school()
    now = timezone.localtime()

    ROLE_AR = {
        "principal": "مدير المدرسة",
        "vice_admin": "نائب المدير الإداري",
        "vice_academic": "نائب المدير الأكاديمي",
        "coordinator": "المنسق",
        "teacher": "المعلم",
        "social_worker": "الأخصائي الاجتماعي",
        "psychologist": "الأخصائي النفسي",
        "admin": "الإداري",
        "platform_developer": "مطور المنصة",
        "nurse": "الممرض",
        "librarian": "أمين المكتبة",
    }

    role_en = user.get_role() or "—"
    logo_path = str(Path(settings.BASE_DIR) / "static" / "brand" / "logoMaroon.png")

    return {
        "school_name": school.name if school else "مدرسة الشحانية الإعدادية الثانوية للبنين",
        "school_logo_path": str(Path(settings.BASE_DIR) / "static" / "brand" / "logowhite.png"),
        "logo_path": logo_path,
        "exported_by": user.full_name,
        "exporter_role": role_en,
        "exporter_role_ar": ROLE_AR.get(role_en, role_en),
        "export_date": now.strftime("%d/%m/%Y"),
        "export_time": now.strftime("%H:%M"),
        "export_datetime": now.strftime("%d/%m/%Y %H:%M"),
        "title": title,
        "academic_year": getattr(settings, "CURRENT_ACADEMIC_YEAR", "2025-2026"),
        "ministry": "وزارة التربية والتعليم والتعليم العالي — دولة قطر",
    }


def add_excel_header(ws, context: dict, num_cols: int):
    """
    هيدر Excel احترافي موحّد:
    صف 1: [شعار] اسم المدرسة — الوزارة (merged)
    صف 2: عنوان التقرير — العام الدراسي (merged)
    صف 3: صدر بواسطة: الاسم — الدور — التاريخ (merged)
    صف 4: (فارغ)
    صف 5: بداية headers البيانات
    يُعيد رقم أول صف للبيانات (5)
    """
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    MAROON = "8A1538"
    header_fill = PatternFill(start_color=MAROON, end_color=MAROON, fill_type="solid")
    white_font_lg = Font(name="Tajawal", bold=True, color="FFFFFF", size=14)
    white_font_md = Font(name="Tajawal", bold=True, color="FFFFFF", size=11)
    gray_font_sm = Font(name="Tajawal", color="666666", size=9)
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    last_col_letter = chr(64 + min(num_cols, 26))  # A=65

    # صف 1: اسم المدرسة
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell1 = ws.cell(row=1, column=1, value=f"{context['school_name']}  —  {context['ministry']}")
    cell1.fill = header_fill
    cell1.font = white_font_lg
    cell1.alignment = center
    ws.row_dimensions[1].height = 40

    # شعار المدرسة
    logo_path = context.get("school_logo_path", "")
    if logo_path and Path(logo_path).exists():
        try:
            img = XLImage(logo_path)
            img.width = 36
            img.height = 36
            ws.add_image(img, "A1")
        except Exception:
            pass

    # صف 2: عنوان التقرير
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    cell2 = ws.cell(
        row=2, column=1, value=f"{context['title']}  —  العام الدراسي {context['academic_year']}"
    )
    cell2.fill = PatternFill(start_color="B91C38", end_color="B91C38", fill_type="solid")
    cell2.font = white_font_md
    cell2.alignment = center
    ws.row_dimensions[2].height = 28

    # صف 3: معلومات التصدير
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=num_cols)
    export_info = f"صدر بواسطة: {context['exported_by']} — {context['exporter_role']}  |  {context['export_datetime']}"
    cell3 = ws.cell(row=3, column=1, value=export_info)
    cell3.font = gray_font_sm
    cell3.alignment = Alignment(horizontal="center", vertical="center")
    cell3.fill = PatternFill(start_color="FDF2F5", end_color="FDF2F5", fill_type="solid")
    ws.row_dimensions[3].height = 22

    # صف 4: فارغ
    ws.row_dimensions[4].height = 8

    return 5  # أول صف للبيانات


def add_excel_footer(ws, context: dict, row: int, num_cols: int):
    """
    فوتر Excel: توقيع المُصدِّر + معلومات المدرسة
    """
    from openpyxl.styles import Alignment, Border, Font, Side

    thin_top = Border(top=Side(style="medium", color="8A1538"))
    footer_font = Font(name="Tajawal", size=9, color="666666")
    sig_font = Font(name="Tajawal", size=9, color="333333", bold=True)

    # صف فارغ
    row += 1

    # توقيع المُصدِّر
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(
        row=row,
        column=1,
        value=f"صدر بواسطة: {context['exported_by']}  —  {context['exporter_role']}",
    )
    cell.font = sig_font
    cell.alignment = Alignment(horizontal="right")
    cell.border = thin_top

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=f"التاريخ: {context['export_datetime']}")
    cell.font = footer_font
    cell.alignment = Alignment(horizontal="right")

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(
        row=row, column=1, value=f"{context['school_name']}  —  وثيقة رسمية  —  SchoolOS"
    )
    cell.font = footer_font
    cell.alignment = Alignment(horizontal="center")

    return row


def get_pdf_header_html(context: dict) -> str:
    """
    HTML هيدر PDF موحّد — يُدرج في أعلى كل template PDF
    """
    return f"""
    <div style="text-align:center;border-bottom:3px solid #8A1538;padding-bottom:12px;margin-bottom:20px">
      <h1 style="color:#8A1538;font-size:16pt;margin:0">{context["school_name"]}</h1>
      <p style="font-size:9pt;color:#666;margin:2px 0 0">{context["ministry"]}</p>
      <p style="font-size:11pt;font-weight:700;color:#333;margin:8px 0 0">{context["title"]}</p>
      <p style="font-size:8pt;color:#999;margin:4px 0 0">العام الدراسي {context["academic_year"]}</p>
    </div>
    """


def get_pdf_footer_html(context: dict) -> str:
    """
    HTML فوتر PDF — توقيع المُصدِّر في أسفل كل صفحة
    """
    return f"""
    <div style="text-align:center;font-size:8pt;color:#999;border-top:1px solid #ddd;padding-top:8px;margin-top:30px">
      <p style="margin:0"><strong style="color:#333">صدر بواسطة:</strong> {context["exported_by"]} — {context["exporter_role"]}</p>
      <p style="margin:2px 0 0">التاريخ: {context["export_datetime"]}</p>
      <p style="margin:2px 0 0">{context["school_name"]} — وثيقة رسمية — SchoolOS</p>
    </div>
    """


def excel_to_response(wb, filename: str):
    """تحويل workbook إلى HttpResponse للتحميل"""
    from io import BytesIO

    from django.http import HttpResponse

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
