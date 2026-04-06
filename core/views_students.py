"""
core/views_students.py
══════════════════════════════════════════════════════════════════════
استيراد / تصدير بيانات الطلاب — Excel
══════════════════════════════════════════════════════════════════════

يشمل:
  - student_import_export  : صفحة إدارة الاستيراد والتصدير (GET/POST)
  - student_export_excel   : تنزيل ملف Excel بكل بيانات الطلاب
  - student_import_template: تنزيل قالب Excel فارغ للاستيراد
"""

from __future__ import annotations

import logging
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from core.permissions import role_required

logger = logging.getLogger(__name__)

# الأعمدة الثابتة لملف الاستيراد/التصدير
EXPORT_COLUMNS = [
    ("الرقم الوطني", 18),
    ("الاسم الكامل", 30),
    ("الصف", 10),
    ("الشعبة", 10),
    ("الجوال", 18),
    ("البريد الإلكتروني", 28),
]

# أعمدة القالب — تُضاف عمود ولي الأمر
TEMPLATE_COLUMNS = [
    ("الرقم الوطني", 18),
    ("الاسم الكامل", 30),
    ("الصف (G7-G12)", 12),
    ("الشعبة (أ/ب/ج...)", 12),
    ("الجوال", 18),
    ("البريد الإلكتروني", 28),
    ("الرقم الوطني ولي الأمر", 20),
    ("اسم ولي الأمر", 28),
    ("جوال ولي الأمر", 18),
    ("بريد ولي الأمر", 28),
    ("صلة القرابة (father/mother/guardian)", 30),
]

MAROON = "8A1538"
WHITE = "FFFFFF"
ALT_BG = "FDF2F5"


# ══════════════════════════════════════════════════════════════════════
# مساعدات Excel
# ══════════════════════════════════════════════════════════════════════


def _make_styles():
    """يُعيد قاموس ستايلات openpyxl مشتركة."""
    import openpyxl  # noqa: F401 — imported for side-effects check
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    return {
        "header_font": Font(name="Arial", bold=True, color=WHITE, size=11),
        "header_fill": PatternFill("solid", fgColor=MAROON),
        "header_align": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "data_align": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "thin_border": Border(
            left=Side(style="thin", color="DDDDDD"),
            right=Side(style="thin", color="DDDDDD"),
            top=Side(style="thin", color="DDDDDD"),
            bottom=Side(style="thin", color="DDDDDD"),
        ),
        "alt_fill": PatternFill("solid", fgColor=ALT_BG),
        "note_font": Font(name="Arial", size=9, color="555555", italic=True),
        "note_fill": PatternFill("solid", fgColor="FFF9E6"),
    }


def _add_header_row(ws, styles, row_num, columns):
    """يرسم صف الرأس باللون الكستنائي."""
    for col_idx, (header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["header_align"]
        cell.border = styles["thin_border"]
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[row_num].height = 26


def _style_data_row(ws, styles, row_num, num_cols, is_alt=False):
    """يطبق ستايل على صف بيانات."""
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.border = styles["thin_border"]
        cell.alignment = styles["data_align"]
        if is_alt:
            cell.fill = styles["alt_fill"]
    ws.row_dimensions[row_num].height = 20


def _setup_workbook(sheet_title, school_name, report_title):
    """ينشئ Workbook بهوية المنصة (4 صفوف رأس)."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.sheet_view.rightToLeft = True

    styles = _make_styles()
    today_str = timezone.now().strftime("%Y/%m/%d")
    from django.conf import settings

    year = getattr(settings, "CURRENT_ACADEMIC_YEAR", "")

    # دعم get_column_letter ديناميكياً

    # الأعمدة ستُحدَّد لاحقاً — نمرر None مؤقتاً
    return wb, ws, styles, year, today_str


def _finalize_workbook(
    wb, ws, styles, num_cols, school_name, report_title, year, today_str, num_data_rows
):
    """يضيف الرأس الاحترافي، يُجمّد، ويُعدّ الطباعة."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.properties import PageSetupProperties

    col_letter = get_column_letter(num_cols)

    bottom_border = Border(bottom=Side(style="medium", color=MAROON))

    # ── الصفوف الثلاثة الأولى للرأس ──────────────────────────────────
    ws.insert_rows(1, 3)

    ws.merge_cells(f"A1:{col_letter}1")
    c = ws["A1"]
    c.value = f"وزارة التربية والتعليم والتعليم العالي — دولة قطر          {today_str}"
    c.font = Font(name="Arial", size=9, color="555555")
    c.fill = PatternFill("solid", fgColor="F5EEF1")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = bottom_border
    ws.row_dimensions[1].height = 22

    ws.merge_cells(f"A2:{col_letter}2")
    c = ws["A2"]
    c.value = school_name
    c.font = Font(name="Arial", bold=True, size=16, color=MAROON)
    c.fill = PatternFill("solid", fgColor="FAFAFA")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = bottom_border
    ws.row_dimensions[2].height = 38

    ws.merge_cells(f"A3:{col_letter}3")
    c = ws["A3"]
    c.value = f"{report_title}   |   السنة الدراسية: {year}"
    c.font = Font(name="Arial", bold=True, size=12, color=MAROON)
    c.fill = PatternFill("solid", fgColor="FDF2F5")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = bottom_border
    ws.row_dimensions[3].height = 26

    # الصف 4 هو الرأس (تم تعبئته قبل insert_rows → أصبح الصف 7 مؤقتاً، لذا نعيد الترتيب)
    # ملاحظة: لتجنب إعادة الترتيب، نضيف الرأس بعد insert_rows مباشرة

    # تجميد بعد الرأس
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{col_letter}4"

    # إعداد طباعة A4
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2
    ws.print_title_rows = "1:4"
    ws.print_area = f"A1:{col_letter}{num_data_rows + 4}"
    ws.oddFooter.center.text = "&P / &N"
    ws.oddFooter.right.text = "&D"


def _wb_to_response(wb, filename):
    """يحوّل Workbook إلى HttpResponse جاهز للتنزيل."""
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# ══════════════════════════════════════════════════════════════════════
# Views
# ══════════════════════════════════════════════════════════════════════


@login_required
@role_required("principal", "vice_admin", "vice_academic", "admin")
def student_import_export(request):
    """
    GET  → صفحة الاستيراد/التصدير
    POST → معالجة ملف الاستيراد
    """
    from django.conf import settings

    from core.models import Membership, Role

    school = request.user.get_school()
    year = getattr(settings, "CURRENT_ACADEMIC_YEAR", "")

    # ── إحصائيات سريعة ──────────────────────────────────────────────
    student_role = Role.objects.filter(name="student").first()
    total_students = 0
    if school and student_role:
        total_students = Membership.objects.filter(
            school=school, role=student_role, is_active=True
        ).count()

    ctx = {
        "school": school,
        "year": year,
        "total_students": total_students,
        "import_result": None,
    }

    if request.method != "POST":
        return render(request, "core/student_import_export.html", ctx)

    # ── POST: استيراد الملف ──────────────────────────────────────────
    uploaded_file = request.FILES.get("student_file")
    if not uploaded_file:
        ctx["import_error"] = "لم يتم اختيار ملف."
        return render(request, "core/student_import_export.html", ctx)

    if not uploaded_file.name.endswith((".xlsx", ".xls")):
        ctx["import_error"] = "يُقبل ملف Excel فقط (.xlsx أو .xls)."
        return render(request, "core/student_import_export.html", ctx)

    try:
        result = _process_import(uploaded_file, school, year)
        ctx["import_result"] = result
    except (OSError, ValueError, KeyError, TypeError) as exc:
        logger.exception("فشل استيراد الطلاب")
        ctx["import_error"] = f"خطأ في قراءة الملف: {exc}"

    return render(request, "core/student_import_export.html", ctx)


# ── ثوابت الاستيراد ──────────────────────────────────────────────────

_IMPORT_RELATION_MAP = {
    "father": "father",
    "mother": "mother",
    "guardian": "guardian",
    "other": "other",
    "أب": "father",
    "والد": "father",
    "أم": "mother",
    "ام": "mother",
    "والدة": "mother",
    "وصي": "guardian",
    "وصية": "guardian",
}

_IMPORT_GRADE_NORMALIZE = {
    "7": "G7",
    "8": "G8",
    "9": "G9",
    "10": "G10",
    "11": "G11",
    "12": "G12",
    "g7": "G7",
    "g8": "G8",
    "g9": "G9",
    "g10": "G10",
    "g11": "G11",
    "g12": "G12",
    "G7": "G7",
    "G8": "G8",
    "G9": "G9",
    "G10": "G10",
    "G11": "G11",
    "G12": "G12",
}


# ── مساعدات الاستيراد ─────────────────────────────────────────────────


def _parse_import_row(row):
    """يحوّل tuple الصف الخام إلى قاموس بأسماء واضحة."""

    def _cell(pos, default=""):
        return str(row[pos]).strip() if len(row) > pos and row[pos] else default

    return {
        "student_nid": _cell(0),
        "full_name": _cell(1),
        "grade_raw": _cell(2),
        "section": _cell(3),
        "phone": _cell(4),
        "email": _cell(5),
        "parent_nid": _cell(6),
        "parent_name": _cell(7),
        "parent_phone": _cell(8),
        "parent_email": _cell(9),
        "relation_raw": _cell(10, "father"),
    }


def _upsert_user(nid, full_name, phone="", email=""):
    """
    get_or_create مستخدم بالرقم الوطني.
    إذا أُنشئ: يضع كلمة المرور ويملأ phone/email.
    إذا كان موجوداً: يُكمل الحقول الفارغة فقط.
    يُعيد (user, created).
    """
    from core.models import CustomUser

    user, created = CustomUser.objects.get_or_create(
        national_id=nid,
        defaults={"full_name": full_name or nid, "is_active": True},
    )
    if created:
        user.set_password(nid)
        if phone:
            user.phone = phone
        if email:
            user.email = email
        user.save()
    else:
        changed = False
        if not user.full_name and full_name:
            user.full_name = full_name
            changed = True
        if not user.phone and phone:
            user.phone = phone
            changed = True
        if not user.email and email:
            user.email = email
            changed = True
        if changed:
            user.save()
    return user, created


def _enroll_student_in_class(student, school, grade_raw, section, stats, row_num):
    """
    يبحث عن الفصل ويسجّل الطالب فيه.
    يُضيف خطأ إلى stats إذا لم يُعثر على الفصل.
    يُعيد True إذا أُنشئ تسجيل جديد.
    """
    from core.models import ClassGroup, StudentEnrollment

    grade = _IMPORT_GRADE_NORMALIZE.get(grade_raw, "")
    if not (grade and section):
        return False

    class_group = ClassGroup.objects.filter(
        school=school, grade=grade, section=section, is_active=True
    ).first()

    if not class_group:
        stats["errors"].append(
            f"سطر {row_num}: الفصل {grade}/{section} غير موجود — تم إنشاء الطالب بدون تسجيل"
        )
        return False

    _, created = StudentEnrollment.objects.get_or_create(
        student=student, class_group=class_group, defaults={"is_active": True}
    )
    return created


def _link_parent_to_student(parent, student, school, relation_raw, stats):
    """
    يُنشئ ParentStudentLink إذا لم يكن موجوداً.
    يُعيد True إذا أُنشئ رابط جديد.
    """
    from core.models import ParentStudentLink

    relation = _IMPORT_RELATION_MAP.get(relation_raw, "father")
    _, created = ParentStudentLink.objects.get_or_create(
        parent=parent,
        student=student,
        school=school,
        defaults={
            "relationship": relation,
            "is_primary": True,
            "can_view_grades": True,
            "can_view_attendance": True,
        },
    )
    return created


def _process_import(uploaded_file, school, year):
    """
    يقرأ ملف Excel ويستورد الطلاب + أولياء الأمور.
    يُعيد dict بإحصائيات النتيجة + قائمة الأخطاء.
    """
    import openpyxl

    from core.models import Membership, Role

    roles = {r.name: r for r in Role.objects.all()}
    student_role = roles.get("student")
    parent_role = roles.get("parent")

    if not student_role or not parent_role:
        return {
            "success": False,
            "errors": ["الأدوار الأساسية (student/parent) غير موجودة — شغّل seed_data أولاً."],
        }

    wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
    ws = wb.active

    # فلترة صفوف البيانات الصالحة (تجاوز الرأس والتعليقات)
    data_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        first = str(row[0]).strip()
        if first.startswith("#") or not first[0].isdigit():
            continue
        data_rows.append(row)

    stats = {
        "students_created": 0,
        "students_existed": 0,
        "parents_created": 0,
        "parents_existed": 0,
        "enrollments_created": 0,
        "links_created": 0,
        "errors": [],
    }

    with transaction.atomic():
        for i, raw_row in enumerate(data_rows, start=2):
            fields = _parse_import_row(raw_row)

            if not fields["student_nid"]:
                stats["errors"].append(f"سطر {i}: الرقم الوطني فارغ — تجاوز")
                continue
            if not fields["full_name"]:
                stats["errors"].append(f"سطر {i}: اسم الطالب {fields['student_nid']} فارغ — تجاوز")
                continue

            # ── الطالب ──────────────────────────────────────────────
            student, s_created = _upsert_user(
                fields["student_nid"],
                fields["full_name"],
                fields["phone"],
                fields["email"],
            )
            stats["students_created" if s_created else "students_existed"] += 1

            if school:
                Membership.objects.get_or_create(
                    user=student,
                    school=school,
                    role=student_role,
                    defaults={"is_active": True},
                )
                if _enroll_student_in_class(
                    student, school, fields["grade_raw"], fields["section"], stats, i
                ):
                    stats["enrollments_created"] += 1

            # ── ولي الأمر (اختياري) ──────────────────────────────────
            if not fields["parent_nid"]:
                continue

            parent, p_created = _upsert_user(
                fields["parent_nid"],
                fields["parent_name"],
                fields["parent_phone"],
                fields["parent_email"],
            )
            stats["parents_created" if p_created else "parents_existed"] += 1

            if school:
                Membership.objects.get_or_create(
                    user=parent,
                    school=school,
                    role=parent_role,
                    defaults={"is_active": True},
                )
                if _link_parent_to_student(parent, student, school, fields["relation_raw"], stats):
                    stats["links_created"] += 1

    return {
        "success": True,
        "total_rows": len(data_rows),
        "students_created": stats["students_created"],
        "students_existed": stats["students_existed"],
        "parents_created": stats["parents_created"],
        "parents_existed": stats["parents_existed"],
        "enrollments_created": stats["enrollments_created"],
        "links_created": stats["links_created"],
        "error_count": len(stats["errors"]),
        "errors": stats["errors"][:20],
    }


@login_required
@role_required("principal", "vice_admin", "vice_academic", "admin")
def student_export_excel(request):
    """
    GET → تنزيل ملف Excel بكل بيانات الطلاب في المدرسة.
    الأعمدة: الرقم الوطني | الاسم | الصف | الشعبة | الجوال | البريد
    """
    from django.conf import settings

    from core.models import Membership, Role, StudentEnrollment

    school = request.user.get_school()
    year = getattr(settings, "CURRENT_ACADEMIC_YEAR", "")

    if not school:
        return HttpResponse("المدرسة غير محددة", status=400)

    student_role = Role.objects.filter(name="student").first()
    if not student_role:
        return HttpResponse("دور الطالب غير موجود", status=400)

    # جلب الطلاب مع تسجيلاتهم دفعةً واحدة (N+1 safe)
    memberships = (
        Membership.objects.filter(school=school, role=student_role, is_active=True)
        .select_related("user")
        .order_by("user__full_name")
    )

    student_ids = [m.user_id for m in memberships]
    enrollments_qs = StudentEnrollment.objects.filter(
        student_id__in=student_ids, is_active=True
    ).select_related("class_group")
    enrollment_map = {enr.student_id: enr for enr in enrollments_qs}

    # ── بناء Workbook ───────────────────────────────────────────────
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.properties import PageSetupProperties

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الطلاب"
    ws.sheet_view.rightToLeft = True

    styles = _make_styles()
    today_str = timezone.now().strftime("%Y/%m/%d")

    num_cols = len(EXPORT_COLUMNS)
    col_letter = get_column_letter(num_cols)

    from openpyxl.styles import Border, Side

    bottom_border = Border(bottom=Side(style="medium", color=MAROON))

    # ── صفوف الرأس الثلاثة ──────────────────────────────────────────
    ws.merge_cells(f"A1:{col_letter}1")
    c = ws["A1"]
    c.value = f"وزارة التربية والتعليم والتعليم العالي — دولة قطر          {today_str}"
    c.font = Font(name="Arial", size=9, color="555555")
    c.fill = PatternFill("solid", fgColor="F5EEF1")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = bottom_border
    ws.row_dimensions[1].height = 22

    ws.merge_cells(f"A2:{col_letter}2")
    c = ws["A2"]
    c.value = school.name
    c.font = Font(name="Arial", bold=True, size=16, color=MAROON)
    c.fill = PatternFill("solid", fgColor="FAFAFA")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = bottom_border
    ws.row_dimensions[2].height = 38

    ws.merge_cells(f"A3:{col_letter}3")
    c = ws["A3"]
    c.value = f"كشف الطلاب الكامل   |   السنة الدراسية: {year}"
    c.font = Font(name="Arial", bold=True, size=12, color=MAROON)
    c.fill = PatternFill("solid", fgColor="FDF2F5")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = bottom_border
    ws.row_dimensions[3].height = 26

    # ── صف رأس الأعمدة (الصف 4) ────────────────────────────────────
    _add_header_row(ws, styles, 4, EXPORT_COLUMNS)

    # ── البيانات (تبدأ من الصف 5) ───────────────────────────────────
    for idx, mem in enumerate(memberships, start=1):
        row_num = idx + 4
        st = mem.user
        enr = enrollment_map.get(st.id)
        grade_display = ""
        section_display = ""
        if enr:
            grade_display = enr.class_group.get_grade_display()
            section_display = enr.class_group.section

        ws.cell(row=row_num, column=1, value=st.national_id or "")
        ws.cell(row=row_num, column=2, value=st.full_name)
        ws.cell(row=row_num, column=3, value=grade_display)
        ws.cell(row=row_num, column=4, value=section_display)
        ws.cell(row=row_num, column=5, value=st.get_phone_decrypted() or "")
        ws.cell(row=row_num, column=6, value=st.email or "")

        _style_data_row(ws, styles, row_num, num_cols, idx % 2 == 0)

    num_data_rows = len(memberships)

    # ── تجميد + فلاتر ────────────────────────────────────────────────
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{col_letter}4"

    # ── إعداد الطباعة A4 ─────────────────────────────────────────────
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2
    ws.print_title_rows = "1:4"
    ws.print_area = f"A1:{col_letter}{num_data_rows + 4}"
    ws.oddFooter.center.text = "&P / &N"
    ws.oddFooter.right.text = "&D"

    filename = f"طلاب_{school.name}_{year}_{today_str.replace('/', '-')}.xlsx"
    return _wb_to_response(wb, filename)


@login_required
@role_required("principal", "vice_admin", "vice_academic", "admin")
def student_import_template(request):
    """
    GET → تنزيل قالب Excel فارغ مع تعليمات الاستيراد.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Border, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.properties import PageSetupProperties

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "استيراد الطلاب"
    ws.sheet_view.rightToLeft = True

    styles = _make_styles()
    today_str = timezone.now().strftime("%Y/%m/%d")

    num_cols = len(TEMPLATE_COLUMNS)
    col_letter = get_column_letter(num_cols)

    bottom_border = Border(bottom=Side(style="medium", color=MAROON))

    # ── صف 1: تعليمات ──────────────────────────────────────────────
    ws.merge_cells(f"A1:{col_letter}1")
    c = ws["A1"]
    c.value = (
        "تعليمات: لا تحذف هذا الصف — ابدأ البيانات من الصف الثالث — "
        "الحقول الإلزامية: الرقم الوطني + الاسم الكامل — "
        "الصفوف المقبولة: G7 G8 G9 G10 G11 G12 — "
        "كلمة المرور الافتراضية = الرقم الوطني"
    )
    c.font = styles["note_font"]
    c.fill = styles["note_fill"]
    c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # ── صف 2: رأس الأعمدة ──────────────────────────────────────────
    _add_header_row(ws, styles, 2, TEMPLATE_COLUMNS)

    # ── صفوف نموذجية (مثال) ─────────────────────────────────────────
    examples = [
        (
            "12345678",
            "محمد أحمد العلي",
            "G9",
            "أ",
            "+97455000001",
            "student@example.com",
            "87654321",
            "أحمد محمد العلي",
            "+97455000002",
            "parent@example.com",
            "father",
        ),
        (
            "12345679",
            "علي محمد السيد",
            "G10",
            "ب",
            "+97455000003",
            "",
            "87654322",
            "محمد سيد الأمين",
            "+97455000004",
            "",
            "father",
        ),
    ]
    from openpyxl.styles import Font as XLFont

    for idx, row_data in enumerate(examples, start=3):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=idx, column=col_idx, value=val)
            cell.border = styles["thin_border"]
            cell.alignment = styles["data_align"]
            # تمييز لوني خفيف للأمثلة
            cell.fill = PatternFill("solid", fgColor="F0F9FF")
            cell.font = XLFont(name="Arial", size=10, color="1E40AF", italic=True)
        ws.row_dimensions[idx].height = 20

    # ── تجميد الصف 2 ────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    # ── إعداد طباعة ─────────────────────────────────────────────────
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

    return _wb_to_response(wb, "قالب_استيراد_الطلاب.xlsx")
