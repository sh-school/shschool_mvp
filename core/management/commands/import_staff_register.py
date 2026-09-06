"""استيرادُ كشف الكادر الرسميّ — كلُّ منتسبٍ بمسمّاه ورقمه.

    JobTitle → Role        بجدولٍ صريحٍ لا بتخمين

كان في المنصّة كادرٌ تدريسيٌّ وحدَه: تسعةٌ وسبعون معلّماً ومنسّقاً بُذروا من
كشفٍ لا يحمل غيرَهم. فلا مشرفَ إداريّاً ولا ملاحظَ طلبةٍ ولا محاسبَ ولا أمينَ
مخزن — سبعةٌ وأربعون موظّفاً خارج النظام، وهم في المدرسة كلَّ يوم.

وهذا الأمرُ يقرأ كشفَ الكادر الكامل (إكسل: الرقم الشخصيّ، الاسم، المسمّى
الوظيفيّ، الرقم الوظيفيّ، البريد، الجوّال) ويُدخل من ليس في القاعدة.

## ثلاثةُ مبادئ

    المسمّى يُحمل كما هو      ولا يُحشر في «إداريّ» ما لا يقابله دور
    الرقمُ الشخصيُّ هو الهويّة  فمن كان في القاعدة يُستكمل ولا يُكرَّر
    المرجعُ يُسجَّل            كشفُ الكادر هو مرجعُ الالتحاق

ولا يكتب شيئاً بلا `--apply`، وهو مُعاوِد: تشغيلُه مرّتين لا يُنشئ عضويّةً ثانية.

    python manage.py import_staff_register --file data/stuff_03.xlsx
    python manage.py import_staff_register --file data/stuff_03.xlsx --apply
"""

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from core.models import CustomUser, Membership, Role, School

#: المسمّى الوظيفيُّ في الكشف ← دورُ المنصّة. وما لم يُذكر هنا يُوقف الاستيرادَ
#: ولا يُخمَّن: دورٌ خاطئٌ يفتح شاشاتٍ لا تخصّ صاحبَه.
TITLE_ROLES = {
    "مدير المدرسة": "principal",
    "نائب المدير للشؤون الاكاديمية": "vice_academic",
    "النائب الإداري": "vice_admin",
    "النائب الاداري": "vice_admin",
    "مشرف اداري": "admin_supervisor",
    "مشرف إداري": "admin_supervisor",
    "الاخصائي الاجتماعي": "social_worker",
    "الأخصائي الاجتماعي": "social_worker",
    "الاخصائي النفسي": "psychologist",
    "الأخصائي النفسي": "psychologist",
    "مرشد أكاديمي": "academic_advisor",
    "مرشد اكاديمي": "academic_advisor",
    "أخصائي الأنشطة المعلمية": "activities_coordinator",
    "أخصائي الأنشطة المدرسية": "activities_coordinator",
    "السكرتير": "secretary",
    "مساعد سكرتير معلمة": "secretary",
    "موظف استقبال": "receptionist",
    "فني تقنية معلومات": "it_technician",
    "مسؤول مصادر التعلم": "librarian",
    "ممرض المدرسة": "nurse",
    "مرافق الدعم": "ese_assistant",
    "معلم تربية خاصة": "ese_teacher",
    "منسق المشاريع الالكترونية": "e_projects_coordinator",
    "ملاحظ طلبه": "student_observer",
    "ملاحظ طلبة": "student_observer",
    "محضر مختبر أحياء": "lab_technician",
    "محضر مختبر فيزياء": "lab_technician",
    "امين مخزن": "storekeeper",
    "أمين مخزن": "storekeeper",
    "محاسب": "accountant",
    "مشرف مقصف": "canteen_supervisor",
    "عامل خدمات": "services_worker",
    "مندوب": "messenger",
}

#: ما بدأ بهذين يُقرأ من بادئته — «معلم رياضيات» و«منسق العلوم» عشراتُ صيغ.
PREFIX_ROLES = (("منسق", "coordinator"), ("معلم", "teacher"))

#: الكادرُ التدريسيُّ — في القاعدة أصلاً، ولا يُستورَد (طلبُ المستخدم 2026-09-06).
TEACHING_ROLES = frozenset(
    {"teacher", "coordinator", "ese_teacher", "e_projects_coordinator"}
)


def role_for(title: str) -> str | None:
    title = " ".join((title or "").split())
    if title in TITLE_ROLES:
        return TITLE_ROLES[title]
    for prefix, role in PREFIX_ROLES:
        if title.startswith(prefix):
            return role
    return None


class Command(BaseCommand):
    help = "يستورد كشفَ الكادر الكامل من إكسل — بلا كتابةٍ إلّا بـ--apply"

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="مسارُ ملفّ الإكسل")
        parser.add_argument(
            "--include-teaching",
            action="store_true",
            help="أدخِل المعلّمين والمنسّقين أيضاً — والافتراضُ الكادرُ الإداريُّ وحدَه",
        )
        parser.add_argument("--sheet", default="", help="اسمُ الورقة — والافتراضُ الأولى")
        parser.add_argument("--school", default="", help="رمزُ المدرسة — والافتراضُ الأولى")
        parser.add_argument(
            "--reference", default="", help="مرجعُ الالتحاق — والافتراضُ اسمُ الملفّ"
        )
        parser.add_argument("--apply", action="store_true")

    def handle(self, *args, **options):
        school = self._school(options["school"])
        rows = self._read(options["file"], options["sheet"])
        reference = options["reference"] or f"كشف الكادر: {options['file'].split('/')[-1]}"

        # الكادرُ التدريسيُّ في القاعدة أصلاً (طلبُ المستخدم 2026-09-06): يُقرأ
        # الكشفُ كلُّه ليُفحص، ولا يُدخل منه إلّا الإداريّون.
        if not options["include_teaching"]:
            rows = [r for r in rows if role_for(r["title"]) not in TEACHING_ROLES]

        unknown = sorted({r["title"] for r in rows if role_for(r["title"]) is None})
        if unknown:
            raise CommandError(
                "مسمّياتٌ لا يقابلها دورٌ في المنصّة — أضِفها إلى الجدول أوّلاً:\n  "
                + "\n  ".join(unknown)
            )

        known = {
            u.national_id: u
            for u in CustomUser.objects.filter(national_id__in=[r["national_id"] for r in rows])
        }
        roles_held = {}
        for m in Membership.objects.filter(school=school, is_active=True).select_related(
            "user", "role"
        ):
            roles_held.setdefault(m.user.national_id, set()).add(m.role.name)

        new, existing, conflicts, skipped = [], [], [], []
        for row in rows:
            if not row["national_id"]:
                skipped.append((row["name"], "بلا رقمٍ شخصيّ في الكشف"))
                continue
            role = role_for(row["title"])
            mine = roles_held.get(row["national_id"], set())
            # عضويّةُ وليّ الأمر ليست تعارضاً: موظّفٌ ابنُه في المدرسة له صفتان،
            # وهي أشهرُ حالةٍ في كشف الكادر لا أندرُها.
            staff_roles = mine - {"student", "parent"}
            if role in mine:
                existing.append(row)
            elif staff_roles:
                # له عضويّةٌ نشطةٌ بدورٍ آخر من الكادر — قرارُ تصحيحٍ لا استيراد.
                conflicts.append((row, role, sorted(staff_roles)))
            else:
                new.append((row, role))

        self.stdout.write(f"المدرسة: {school.name} · الكشف: {len(rows)} سطراً")
        self.stdout.write(
            f"قائمٌ بالفعل: {len(existing)} · جديد: {len(new)} ·"
            f" يُراجَع: {len(conflicts)} · متعذّر: {len(skipped)}"
        )

        for row, role in sorted(new, key=lambda p: (p[1], p[0]["name"])):
            mark = "حسابٌ قائم" if row["national_id"] in known else "حسابٌ جديد"
            self.stdout.write(f"  + {row['title']:<26} {row['name']:<34} [{role}] {mark}")
        for row, role, mine in sorted(conflicts, key=lambda p: p[0]["name"]):
            self.stdout.write(
                self.style.WARNING(
                    f"  ? {row['title']:<26} {row['name']:<34} الكشفُ [{role}]"
                    f" والقاعدةُ [{'، '.join(mine)}] — يُراجَع يدويّاً"
                )
            )
        for name, why in skipped:
            self.stdout.write(self.style.WARNING(f"  ! {name} — {why}"))

        if not options["apply"]:
            self.stdout.write("\nتقريرٌ فقط — أضِف --apply للكتابة.")
            return

        created_users, created_memberships = self._write(school, new, reference)
        self.stdout.write(
            f"\nأُنشئ {created_users} حساباً و{created_memberships} عضويّة — "
            "والحساباتُ بلا كلمة مرورٍ حتّى تُصدَر لها."
        )

    # ── القراءة ──────────────────────────────────────────────────────

    def _school(self, code):
        school = (School.objects.filter(code=code) if code else School.objects.all()).first()
        if school is None:
            raise CommandError("لا مدرسةَ بهذا الرمز.")
        return school

    def _read(self, path, sheet):
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover - بيئةٌ بلا المكتبة
            raise CommandError("openpyxl غيرُ مثبّتة.") from exc

        try:
            book = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except OSError as exc:
            raise CommandError(f"تعذّر فتحُ الملفّ: {path}") from exc

        worksheet = book[sheet] if sheet else book[book.sheetnames[0]]
        raw = list(worksheet.iter_rows(values_only=True))
        book.close()
        if not raw:
            raise CommandError("الورقةُ فارغة.")

        header = [str(c).strip() if c is not None else "" for c in raw[0]]
        columns = {name: index for index, name in enumerate(header)}
        needed = ("national_no", "stuff _name", "job_title")
        missing = [c for c in needed if c not in columns]
        if missing:
            raise CommandError(f"أعمدةٌ ناقصةٌ في الكشف: {'، '.join(missing)}")

        def cell(row, name):
            index = columns.get(name)
            if index is None or index >= len(row) or row[index] is None:
                return ""
            return str(row[index]).strip()

        rows = []
        for row in raw[1:]:
            if not any(row):
                continue
            name = " ".join(cell(row, "stuff _name").split())
            if not name:
                continue
            rows.append(
                {
                    "national_id": cell(row, "national_no"),
                    "name": name,
                    "title": " ".join(cell(row, "job_title").split()),
                    "employee_number": cell(row, "job_no"),
                    "email": cell(row, "email"),
                    "phone": cell(row, "phone_no"),
                }
            )
        return rows

    # ── الكتابة ──────────────────────────────────────────────────────

    @transaction.atomic
    def _write(self, school, new, reference):
        today = timezone.localdate()
        users, memberships = 0, 0
        for row, role_name in new:
            user = CustomUser.objects.filter(national_id=row["national_id"]).first()
            if user is None:
                user = CustomUser(
                    national_id=row["national_id"],
                    full_name=row["name"],
                    email=row["email"],
                    phone=row["phone"],
                    employee_number=row["employee_number"],
                )
                user.set_unusable_password()
                user.full_clean(exclude=["password", "last_login"])
                user.save()
                users += 1
            else:
                changed = []
                for field, value in (
                    ("email", row["email"]),
                    ("phone", row["phone"]),
                    ("employee_number", row["employee_number"]),
                ):
                    if value and not getattr(user, field):
                        setattr(user, field, value)
                        changed.append(field)
                if changed:
                    user.save(update_fields=changed)

            role, _ = Role.objects.get_or_create(school=school, name=role_name)
            membership = Membership(
                user=user,
                school=school,
                role=role,
                joined_at=today,
                job_title=row["title"],
                appointment_reference=reference,
                is_active=True,
            )
            membership.full_clean(exclude=["user", "school", "role"])
            membership.save()
            memberships += 1
            user.invalidate_active_membership()
        return users, memberships
