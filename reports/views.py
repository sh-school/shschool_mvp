"""
reports/views.py
نظام تقارير — PDF + Excel
- PDF: WeasyPrint (الشهادات وكشف النتائج)
- Excel: openpyxl (مثبتة في requirements.txt)
[مهمة 17] إضافة Export Excel لـ 3 تقارير:
  1. كشف نتائج الفصل   → class_results_excel
  2. تقرير الغياب       → attendance_excel
  3. مخالفات السلوك     → behavior_excel
"""
from io import BytesIO
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, Http404
from django.template.loader import render_to_string
from django.utils import timezone


# ── مساعد Excel مشترك ──────────────────────────────────────────────────
def _make_workbook(title):
    """ينشئ Workbook جاهز مع ستايل الهوية القطرية"""
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws.sheet_view.rightToLeft = True   # RTL عربي

    MAROON = "8A1538"
    WHITE  = "FFFFFF"
    LIGHT  = "FDF2F5"

    header_font   = Font(name="Arial", bold=True, color=WHITE, size=11)
    header_fill   = PatternFill("solid", fgColor=MAROON)
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border   = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    alt_fill      = PatternFill("solid", fgColor="FDF2F5")

    styles = {
        "header_font":  header_font,
        "header_fill":  header_fill,
        "header_align": header_align,
        "thin_border":  thin_border,
        "alt_fill":     alt_fill,
        "maroon":       MAROON,
        "white":        WHITE,
    }
    return wb, ws, styles


def _apply_header_row(ws, styles, row_num, columns):
    """يطبّق ستايل رأس الجدول على صف محدد"""
    from openpyxl.styles import Font, PatternFill, Alignment
    for col_idx, (header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font      = styles["header_font"]
        cell.fill      = styles["header_fill"]
        cell.alignment = styles["header_align"]
        cell.border    = styles["thin_border"]
        ws.column_dimensions[
            ws.cell(row=row_num, column=col_idx).column_letter
        ].width = width


def _style_data_row(ws, styles, row_num, num_cols, is_alt=False):
    """يطبّق ستايل صفوف البيانات"""
    from openpyxl.styles import Alignment
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.border    = styles["thin_border"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if is_alt:
            cell.fill = styles["alt_fill"]


def _excel_response(wb, filename):
    """يُعيد HttpResponse جاهز لتنزيل Excel"""
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

from assessments.models import (
    AnnualSubjectResult, StudentSubjectResult,
    SubjectClassSetup, AssessmentPackage,
)
from core.models import CustomUser, StudentEnrollment, ClassGroup, School
from operations.models import StudentAttendance


def _render_pdf(html_string, filename="report.pdf"):
    """تحويل HTML → PDF مع دعم الخطوط العربية — يعيد bytes"""
    from core.pdf_utils import render_pdf_bytes
    return render_pdf_bytes(html_string)


def _get_student_report_data(student, school, year):
    """بيانات تقرير الطالب السنوي الكاملة"""
    annual = AnnualSubjectResult.objects.filter(
        student=student, school=school, academic_year=year
    ).select_related("setup__subject").order_by("setup__subject__name_ar")

    s1_map = {
        r.setup_id: r for r in StudentSubjectResult.objects.filter(
            student=student, school=school, semester="S1"
        )
    }
    s2_map = {
        r.setup_id: r for r in StudentSubjectResult.objects.filter(
            student=student, school=school, semester="S2"
        )
    }

    rows = []
    for ann in annual:
        rows.append({
            "subject":  ann.setup.subject.name_ar,
            "s1":       s1_map.get(ann.setup_id),
            "s2":       s2_map.get(ann.setup_id),
            "annual":   ann,
        })

    total   = annual.count()
    passed  = annual.filter(status="pass").count()
    failed  = annual.filter(status="fail").count()
    grades  = [float(r.annual_total) for r in annual if r.annual_total]
    avg     = round(sum(grades) / len(grades), 2) if grades else None
    rank    = None  # يمكن إضافته لاحقاً

    enrollment = StudentEnrollment.objects.filter(
        student=student, is_active=True
    ).select_related("class_group").first()

    # الغياب
    att = StudentAttendance.objects.filter(student=student, session__school=school)
    absent_total = att.filter(status="absent").count()
    late_total   = att.filter(status="late").count()

    return {
        "student":       student,
        "school":        school,
        "year":          year,
        "enrollment":    enrollment,
        "rows":          rows,
        "total":         total,
        "passed":        passed,
        "failed":        failed,
        "avg":           avg,
        "absent_total":  absent_total,
        "late_total":    late_total,
        "print_date":    timezone.now().date(),
    }


# ── صفحة اختيار التقرير ────────────────────────────────────

@login_required
def reports_index(request):
    school = request.user.get_school()
    year   = request.GET.get("year", "2025-2026")

    # الفصول الدراسية للمدير
    if request.user.is_admin():
        classes = ClassGroup.objects.filter(
            school=school, academic_year=year, is_active=True
        ).order_by("grade", "section")
    else:
        # المعلم: فصوله فقط
        setup_class_ids = SubjectClassSetup.objects.filter(
            school=school, teacher=request.user, academic_year=year
        ).values_list("class_group_id", flat=True)
        classes = ClassGroup.objects.filter(id__in=setup_class_ids)

    return render(request, "reports/index.html", {
        "classes": classes,
        "year":    year,
        "school":  school,
    })


# ── كشف نتائج الفصل الدراسي ────────────────────────────────

@login_required
def class_results_pdf(request, class_id):
    """PDF: كشف نتائج كامل لجميع طلاب فصل"""
    school     = request.user.get_school()
    class_grp  = get_object_or_404(ClassGroup, id=class_id, school=school)
    year       = request.GET.get("year", "2025-2026")
    preview    = request.GET.get("preview", "0") == "1"

    enrollments = StudentEnrollment.objects.filter(
        class_group=class_grp, is_active=True
    ).select_related("student").order_by("student__full_name")

    # بيانات كل طالب
    students_data = []
    for enr in enrollments:
        st      = enr.student
        annual  = AnnualSubjectResult.objects.filter(
            student=st, school=school, academic_year=year
        ).select_related("setup__subject").order_by("setup__subject__name_ar")

        grades  = [float(r.annual_total) for r in annual if r.annual_total]
        avg     = round(sum(grades) / len(grades), 2) if grades else None
        passed  = annual.filter(status="pass").count()
        failed  = annual.filter(status="fail").count()

        students_data.append({
            "student": st,
            "annual":  annual,
            "avg":     avg,
            "passed":  passed,
            "failed":  failed,
            "status":  "ناجح" if failed == 0 and passed > 0 else ("راسب" if failed > 0 else "—"),
        })

    # ترتيب حسب المتوسط
    students_data.sort(key=lambda x: x["avg"] or 0, reverse=True)
    for i, sd in enumerate(students_data, start=1):
        sd["rank"] = i

    # المواد (من أول طالب)
    subjects = []
    if enrollments.exists():
        subjects = list(AnnualSubjectResult.objects.filter(
            student=enrollments.first().student,
            school=school, academic_year=year
        ).select_related("setup__subject").order_by(
            "setup__subject__name_ar"
        ).values_list("setup__subject__name_ar", flat=True))

    ctx = {
        "class_group":    class_grp,
        "school":         school,
        "year":           year,
        "students_data":  students_data,
        "subjects":       subjects,
        "print_date":     timezone.now().date(),
        "total_students": len(students_data),
        "total_passed":   sum(1 for s in students_data if s["failed"] == 0 and s["passed"] > 0),
        "total_failed":   sum(1 for s in students_data if s["failed"] > 0),
    }

    if preview:
        return render(request, "reports/class_results.html", ctx)

    html = render_to_string("reports/class_results.html", ctx, request=request)
    pdf  = _render_pdf(html)
    filename = f"نتائج_{class_grp.get_grade_display()}_{class_grp.section}_{year}.pdf"
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    return resp


# ── تقرير الطالب الفردي ─────────────────────────────────────

@login_required
def student_result_pdf(request, student_id):
    """PDF: تقرير نتيجة طالب واحد مفصّل"""
    school   = request.user.get_school()
    student  = get_object_or_404(CustomUser, id=student_id)
    year     = request.GET.get("year", "2025-2026")
    preview  = request.GET.get("preview", "0") == "1"

    # صلاحية
    if (not request.user.is_admin()
            and not request.user.is_teacher()
            and request.user != student):
        # ولي الأمر
        from core.models import ParentStudentLink
        is_parent = ParentStudentLink.objects.filter(
            parent=request.user, student=student, school=school
        ).exists()
        if not is_parent:
            return HttpResponse("غير مسموح", status=403)

    ctx = _get_student_report_data(student, school, year)

    if preview:
        return render(request, "reports/student_result.html", ctx)

    html = render_to_string("reports/student_result.html", ctx, request=request)
    pdf  = _render_pdf(html)
    filename = f"نتيجة_{student.full_name}_{year}.pdf"
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    return resp


# ── شهادة التقدير (Certificate) ─────────────────────────────

@login_required
def student_certificate_pdf(request, student_id):
    """PDF: شهادة نتيجة سنوية رسمية"""
    school   = request.user.get_school()
    student  = get_object_or_404(CustomUser, id=student_id)
    year     = request.GET.get("year", "2025-2026")
    preview  = request.GET.get("preview", "0") == "1"

    if not request.user.is_admin() and not request.user.is_teacher():
        from core.models import ParentStudentLink
        is_parent = ParentStudentLink.objects.filter(
            parent=request.user, student=student, school=school
        ).exists()
        if not is_parent:
            return HttpResponse("غير مسموح", status=403)

    ctx = _get_student_report_data(student, school, year)

    # تحديد الحالة النهائية
    if ctx["failed"] == 0 and ctx["passed"] > 0:
        ctx["final_status"] = "ناجح"
        ctx["status_color"] = "#15803d"
    elif ctx["failed"] > 0:
        ctx["final_status"] = "راسب"
        ctx["status_color"] = "#dc2626"
    else:
        ctx["final_status"] = "غير مكتمل"
        ctx["status_color"] = "#d97706"

    if preview:
        return render(request, "reports/certificate.html", ctx)

    html = render_to_string("reports/certificate.html", ctx, request=request)
    pdf  = _render_pdf(html)
    filename = f"شهادة_{student.full_name}_{year}.pdf"
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    return resp


# ── طباعة شهادات الفصل بالجملة ─────────────────────────────

@login_required
def class_certificates_pdf(request, class_id):
    """PDF: شهادات جميع طلاب فصل في ملف واحد"""
    if not request.user.is_admin():
        return HttpResponse("غير مسموح", status=403)

    school    = request.user.get_school()
    class_grp = get_object_or_404(ClassGroup, id=class_id, school=school)
    year      = request.GET.get("year", "2025-2026")
    preview   = request.GET.get("preview", "0") == "1"

    enrollments = StudentEnrollment.objects.filter(
        class_group=class_grp, is_active=True
    ).select_related("student").order_by("student__full_name")

    students_ctx = []
    for enr in enrollments:
        ctx = _get_student_report_data(enr.student, school, year)
        if ctx["failed"] == 0 and ctx["passed"] > 0:
            ctx["final_status"] = "ناجح"
            ctx["status_color"] = "#15803d"
        elif ctx["failed"] > 0:
            ctx["final_status"] = "راسب"
            ctx["status_color"] = "#dc2626"
        else:
            ctx["final_status"] = "غير مكتمل"
            ctx["status_color"] = "#d97706"
        students_ctx.append(ctx)

    ctx = {
        "students_ctx": students_ctx,
        "class_group":  class_grp,
        "school":       school,
        "year":         year,
        "print_date":   timezone.now().date(),
    }

    if preview:
        return render(request, "reports/class_certificates.html", ctx)

    html = render_to_string("reports/class_certificates.html", ctx, request=request)
    pdf  = _render_pdf(html)
    filename = f"شهادات_{class_grp.get_grade_display()}_{class_grp.section}_{year}.pdf"
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    return resp


# ── تقرير الغياب ────────────────────────────────────────────

@login_required
def attendance_report_pdf(request, class_id):
    """PDF: تقرير حضور وغياب الفصل"""
    if not request.user.is_admin():
        return HttpResponse("غير مسموح", status=403)

    school    = request.user.get_school()
    class_grp = get_object_or_404(ClassGroup, id=class_id, school=school)
    year      = request.GET.get("year", "2025-2026")
    preview   = request.GET.get("preview", "0") == "1"

    enrollments = StudentEnrollment.objects.filter(
        class_group=class_grp, is_active=True
    ).select_related("student").order_by("student__full_name")

    rows = []
    for enr in enrollments:
        att = StudentAttendance.objects.filter(
            student=enr.student, session__school=school
        )
        total   = att.count()
        present = att.filter(status="present").count()
        absent  = att.filter(status="absent").count()
        late    = att.filter(status="late").count()
        pct     = round(present / total * 100) if total else 0
        rows.append({
            "student": enr.student,
            "total":   total,
            "present": present,
            "absent":  absent,
            "late":    late,
            "pct":     pct,
        })

    ctx = {
        "class_group": class_grp,
        "school":      school,
        "year":        year,
        "rows":        rows,
        "print_date":  timezone.now().date(),
    }

    if preview:
        return render(request, "reports/attendance_report.html", ctx)

    html = render_to_string("reports/attendance_report.html", ctx, request=request)
    pdf  = _render_pdf(html)
    filename = f"غياب_{class_grp.get_grade_display()}_{class_grp.section}_{year}.pdf"
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    return resp



# ══════════════════════════════════════════════════════════════════════
# [مهمة 17] — تقارير Excel
# ══════════════════════════════════════════════════════════════════════

# ── 1. كشف نتائج الفصل Excel ──────────────────────────────────────────

@login_required
def class_results_excel(request, class_id):
    """Excel: كشف نتائج كامل لجميع طلاب فصل — مع كل المواد"""
    if not request.user.is_admin() and not request.user.is_teacher():
        return HttpResponse("غير مسموح", status=403)

    school    = request.user.get_school()
    class_grp = get_object_or_404(ClassGroup, id=class_id, school=school)
    year      = request.GET.get("year", "2025-2026")

    enrollments = StudentEnrollment.objects.filter(
        class_group=class_grp, is_active=True
    ).select_related("student").order_by("student__full_name")

    # الحصول على المواد من أول طالب
    first_student = enrollments.first()
    subjects = []
    if first_student:
        subjects = list(
            AnnualSubjectResult.objects.filter(
                student=first_student.student, school=school, academic_year=year
            ).select_related("setup__subject")
            .order_by("setup__subject__name_ar")
            .values_list("setup__subject__name_ar", flat=True)
        )

    wb, ws, styles = _make_workbook("كشف النتائج")

    # ── صف العنوان ──────────────────────────────────────────────────
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.merge_cells(f"A1:{chr(65 + 5 + len(subjects))}1")
    title_cell = ws["A1"]
    title_cell.value     = f"كشف نتائج — {class_grp} — {year} — {school.name}"
    title_cell.font      = Font(name="Arial", bold=True, color=styles["white"], size=13)
    title_cell.fill      = PatternFill("solid", fgColor=styles["maroon"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── رأس الجدول ──────────────────────────────────────────────────
    columns = (
        [("م", 5), ("اسم الطالب", 28), ("الرقم الوطني", 16)]
        + [(subj[:20], 12) for subj in subjects]
        + [("المتوسط", 10), ("الحالة", 10), ("الترتيب", 8)]
    )
    _apply_header_row(ws, styles, 2, columns)
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    # ── البيانات ─────────────────────────────────────────────────────
    students_data = []
    for enr in enrollments:
        st     = enr.student
        annual = AnnualSubjectResult.objects.filter(
            student=st, school=school, academic_year=year
        ).select_related("setup__subject").order_by("setup__subject__name_ar")

        grade_map = {r.setup.subject.name_ar: r.annual_total for r in annual}
        grades    = [float(g) for g in grade_map.values() if g is not None]
        avg       = round(sum(grades) / len(grades), 2) if grades else 0
        failed    = sum(1 for r in annual if r.status == "fail")
        passed    = sum(1 for r in annual if r.status == "pass")
        status    = "ناجح" if failed == 0 and passed > 0 else ("راسب" if failed > 0 else "—")
        students_data.append((st, grade_map, avg, status))

    # ترتيب حسب المتوسط
    students_data.sort(key=lambda x: x[2], reverse=True)

    for row_idx, (st, grade_map, avg, status) in enumerate(students_data, start=1):
        row_num = row_idx + 2
        is_alt  = row_idx % 2 == 0

        ws.cell(row=row_num, column=1, value=row_idx)
        ws.cell(row=row_num, column=2, value=st.full_name)
        ws.cell(row=row_num, column=3, value=st.national_id)

        for col_offset, subj in enumerate(subjects, start=4):
            grade = grade_map.get(subj)
            cell  = ws.cell(row=row_num, column=col_offset,
                            value=float(grade) if grade else "—")
            # تلوين الراسب
            if grade and float(grade) < 50:
                cell.font = Font(name="Arial", color="DC2626", bold=True)

        ws.cell(row=row_num, column=4 + len(subjects), value=avg)
        status_cell = ws.cell(row=row_num, column=5 + len(subjects), value=status)
        if status == "ناجح":
            status_cell.font = Font(name="Arial", color="15803D", bold=True)
        elif status == "راسب":
            status_cell.font = Font(name="Arial", color="DC2626", bold=True)
        ws.cell(row=row_num, column=6 + len(subjects), value=row_idx)

        _style_data_row(ws, styles, row_num, len(columns), is_alt)
        ws.row_dimensions[row_num].height = 18

    filename = f"نتائج_{class_grp.get_grade_display()}_{class_grp.section}_{year}.xlsx"
    return _excel_response(wb, filename)


# ── 2. تقرير الغياب Excel ─────────────────────────────────────────────

@login_required
def attendance_excel(request, class_id):
    """Excel: تقرير حضور وغياب تفصيلي للفصل"""
    if not request.user.is_admin():
        return HttpResponse("غير مسموح", status=403)

    school    = request.user.get_school()
    class_grp = get_object_or_404(ClassGroup, id=class_id, school=school)
    year      = request.GET.get("year", "2025-2026")

    enrollments = StudentEnrollment.objects.filter(
        class_group=class_grp, is_active=True
    ).select_related("student").order_by("student__full_name")

    wb, ws, styles = _make_workbook("الغياب")

    # العنوان
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.merge_cells("A1:H1")
    tc = ws["A1"]
    tc.value     = f"تقرير الحضور والغياب — {class_grp} — {year} — {school.name}"
    tc.font      = Font(name="Arial", bold=True, color=styles["white"], size=13)
    tc.fill      = PatternFill("solid", fgColor=styles["maroon"])
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    columns = [
        ("م", 5), ("اسم الطالب", 28), ("الرقم الوطني", 16),
        ("إجمالي الحصص", 13), ("حاضر", 10), ("غائب", 10),
        ("متأخر", 10), ("نسبة الحضور %", 14),
    ]
    _apply_header_row(ws, styles, 2, columns)
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    for row_idx, enr in enumerate(enrollments, start=1):
        row_num = row_idx + 2
        is_alt  = row_idx % 2 == 0
        att     = StudentAttendance.objects.filter(
            student=enr.student, session__school=school
        )
        total   = att.count()
        present = att.filter(status="present").count()
        absent  = att.filter(status="absent").count()
        late    = att.filter(status="late").count()
        pct     = round(present / total * 100, 1) if total else 0

        ws.cell(row=row_num, column=1, value=row_idx)
        ws.cell(row=row_num, column=2, value=enr.student.full_name)
        ws.cell(row=row_num, column=3, value=enr.student.national_id)
        ws.cell(row=row_num, column=4, value=total)
        ws.cell(row=row_num, column=5, value=present)

        absent_cell = ws.cell(row=row_num, column=6, value=absent)
        if absent > 10:
            absent_cell.font = Font(name="Arial", color="DC2626", bold=True)

        ws.cell(row=row_num, column=7, value=late)

        pct_cell = ws.cell(row=row_num, column=8, value=pct)
        if pct < 80:
            pct_cell.font = Font(name="Arial", color="DC2626", bold=True)
        elif pct >= 95:
            pct_cell.font = Font(name="Arial", color="15803D", bold=True)

        _style_data_row(ws, styles, row_num, 8, is_alt)
        ws.row_dimensions[row_num].height = 18

    filename = f"غياب_{class_grp.get_grade_display()}_{class_grp.section}_{year}.xlsx"
    return _excel_response(wb, filename)


# ── 3. تقرير السلوك Excel ─────────────────────────────────────────────

@login_required
def behavior_excel(request):
    """Excel: كشف مخالفات السلوك لكل الطلاب"""
    if not request.user.is_admin():
        return HttpResponse("غير مسموح", status=403)

    school = request.user.get_school()
    year   = request.GET.get("year", "2025-2026")

    from core.models import BehaviorInfraction
    infractions = BehaviorInfraction.objects.filter(
        school=school
    ).select_related("student", "reported_by").order_by("-date")

    wb, ws, styles = _make_workbook("السلوك")

    from openpyxl.styles import Font, PatternFill, Alignment
    ws.merge_cells("A1:H1")
    tc = ws["A1"]
    tc.value     = f"تقرير المخالفات السلوكية — {school.name} — {year}"
    tc.font      = Font(name="Arial", bold=True, color=styles["white"], size=13)
    tc.fill      = PatternFill("solid", fgColor=styles["maroon"])
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    LEVEL_LABELS = {1: "درجة 1 — بسيطة", 2: "درجة 2 — متوسطة",
                    3: "درجة 3 — جسيمة", 4: "درجة 4 — شديدة"}
    LEVEL_COLORS = {1: "854D0E", 2: "C2410C", 3: "B91C1C", 4: "BE123C"}

    columns = [
        ("م", 5), ("اسم الطالب", 28), ("الرقم الوطني", 16),
        ("التاريخ", 13), ("الدرجة", 18), ("النقاط المخصومة", 15),
        ("المُبلِّغ", 20), ("الوصف", 40),
    ]
    _apply_header_row(ws, styles, 2, columns)
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    for row_idx, inf in enumerate(infractions, start=1):
        row_num = row_idx + 2
        is_alt  = row_idx % 2 == 0

        ws.cell(row=row_num, column=1, value=row_idx)
        ws.cell(row=row_num, column=2, value=inf.student.full_name)
        ws.cell(row=row_num, column=3, value=inf.student.national_id)
        ws.cell(row=row_num, column=4,
                value=inf.date.strftime("%Y/%m/%d") if inf.date else "")

        level_cell = ws.cell(row=row_num, column=5,
                             value=LEVEL_LABELS.get(inf.level, inf.level))
        level_cell.font = Font(name="Arial",
                               color=LEVEL_COLORS.get(inf.level, "000000"),
                               bold=True)

        ws.cell(row=row_num, column=6, value=inf.points_deducted)
        ws.cell(row=row_num, column=7,
                value=inf.reported_by.full_name if inf.reported_by else "—")
        ws.cell(row=row_num, column=8, value=inf.description)

        _style_data_row(ws, styles, row_num, 8, is_alt)
        ws.row_dimensions[row_num].height = 20

    filename = f"سلوك_{school.name}_{year}.xlsx"
    return _excel_response(wb, filename)
