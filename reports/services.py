"""
reports/services.py
━━━━━━━━━━━━━━━━━━
Business logic لوحدة التقارير — بدون أي HTTP logic

يشمل:
  - ReportDataService  : تجميع بيانات التقارير
  - ExcelService       : إنشاء ملفات Excel باحترافية كاملة
"""

from __future__ import annotations

import logging
from io import BytesIO
from typing import TYPE_CHECKING

from django.conf import settings
from django.http import HttpResponse
from django.utils import timezone

from assessments.models import (
    AnnualSubjectResult,
    StudentSubjectResult,
    SubjectClassSetup,
)
from core.models import StudentEnrollment
from operations.models import StudentAttendance

if TYPE_CHECKING:
    from core.models import ClassGroup, CustomUser, School

logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════
# ReportDataService — تجميع البيانات
# ══════════════════════════════════════════════════════════════════════


class ReportDataService:
    """تجميع بيانات التقارير — بدون أي HTTP logic"""

    @staticmethod
    def get_student_report(
        student: CustomUser,
        school: School,
        year: str = settings.CURRENT_ACADEMIC_YEAR,
    ) -> dict:
        """بيانات تقرير الطالب السنوي الكاملة"""
        annual = (
            AnnualSubjectResult.objects.filter(student=student, school=school, academic_year=year)
            .select_related("setup__subject")
            .order_by("setup__subject__name_ar")
        )

        s1_map = {
            r.setup_id: r
            for r in StudentSubjectResult.objects.filter(
                student=student, school=school, semester="S1"
            )
        }
        s2_map = {
            r.setup_id: r
            for r in StudentSubjectResult.objects.filter(
                student=student, school=school, semester="S2"
            )
        }

        rows = [
            {
                "subject": ann.setup.subject.name_ar,
                "s1": s1_map.get(ann.setup_id),
                "s2": s2_map.get(ann.setup_id),
                "annual": ann,
            }
            for ann in annual
        ]

        total = annual.count()
        passed = annual.filter(status="pass").count()
        failed = annual.filter(status="fail").count()
        grades = [float(r.annual_total) for r in annual if r.annual_total]
        avg = round(sum(grades) / len(grades), 2) if grades else None

        enrollment = (
            StudentEnrollment.objects.filter(student=student, is_active=True)
            .select_related("class_group")
            .first()
        )

        att = StudentAttendance.objects.filter(student=student, session__school=school)
        absent_total = att.filter(status="absent").count()
        late_total = att.filter(status="late").count()

        return {
            "student": student,
            "school": school,
            "year": year,
            "enrollment": enrollment,
            "rows": rows,
            "total": total,
            "passed": passed,
            "failed": failed,
            "avg": avg,
            "absent_total": absent_total,
            "late_total": late_total,
            "print_date": timezone.now().date(),
        }

    @staticmethod
    def get_class_results(
        class_group: ClassGroup,
        school: School,
        year: str = settings.CURRENT_ACADEMIC_YEAR,
    ) -> dict:
        """بيانات كشف نتائج الفصل الكامل مع ترتيب + ملخص"""
        enrollments = list(
            StudentEnrollment.objects.filter(class_group=class_group, is_active=True)
            .select_related("student")
            .order_by("student__full_name")
        )

        setups = list(
            SubjectClassSetup.objects.filter(
                class_group=class_group, school=school, academic_year=year
            )
            .select_related("subject")
            .order_by("subject__name_ar")
        )

        subjects = [s.subject for s in setups]

        # ── جلب كل النتائج السنوية دفعةً واحدة (يُلغي N×M queries) ──
        student_ids = [enr.student_id for enr in enrollments]
        all_annuals = AnnualSubjectResult.objects.filter(
            setup__in=setups,
            student_id__in=student_ids,
            academic_year=year,
        ).select_related("setup__subject")

        # lookup: (student_id, setup_id) → AnnualSubjectResult
        annual_map = {(a.student_id, a.setup_id): a for a in all_annuals}

        student_rows: list = []
        for enr in enrollments:
            row: dict = {"student": enr.student, "grades": {}}
            grades: list = []
            passed = failed = 0
            for setup in setups:
                annual = annual_map.get((enr.student_id, setup.id))
                row["grades"][setup.subject.name_ar] = annual
                if annual:
                    if annual.annual_total:
                        grades.append(float(annual.annual_total))
                    if annual.status == "pass":
                        passed += 1
                    elif annual.status == "fail":
                        failed += 1

            avg = round(sum(grades) / len(grades), 2) if grades else None
            row["avg"] = avg
            row["passed"] = passed
            row["failed"] = failed
            row["status"] = (
                "ناجح" if failed == 0 and passed > 0 else ("راسب" if failed > 0 else "—")
            )
            student_rows.append(row)

        # ترتيب حسب المتوسط
        student_rows.sort(key=lambda x: x["avg"] or 0, reverse=True)
        for i, row in enumerate(student_rows, start=1):
            row["rank"] = i
            row["grades_list"] = [row["grades"].get(s.name_ar) for s in subjects]

        total_passed = sum(1 for r in student_rows if r["failed"] == 0 and r["passed"] > 0)
        total_failed = sum(1 for r in student_rows if r["failed"] > 0)

        return {
            "class_group": class_group,
            "school": school,
            "year": year,
            "subjects": subjects,
            "student_rows": student_rows,
            "total_students": len(student_rows),
            "total_passed": total_passed,
            "total_failed": total_failed,
            "print_date": timezone.now().date(),
        }

    @staticmethod
    def get_attendance_report(
        class_group: ClassGroup,
        school: School,
        year: str = settings.CURRENT_ACADEMIC_YEAR,
    ) -> dict:
        """بيانات تقرير الغياب لفصل"""
        enrollments = list(
            StudentEnrollment.objects.filter(class_group=class_group, is_active=True)
            .select_related("student")
            .order_by("student__full_name")
        )

        # ── جلب كل سجلات الحضور دفعةً واحدة (يُلغي N×4 queries) ──
        student_ids = [enr.student_id for enr in enrollments]
        all_att = StudentAttendance.objects.filter(
            student_id__in=student_ids,
            session__school=school,
        ).values("student_id", "status")

        # تجميع الإحصائيات لكل طالب في Python
        from collections import defaultdict

        att_counts = defaultdict(lambda: {"total": 0, "absent": 0, "late": 0, "excused": 0})
        for rec in all_att:
            sid = rec["student_id"]
            att_counts[sid]["total"] += 1
            if rec["status"] in ("absent", "late", "excused"):
                att_counts[sid][rec["status"]] += 1

        student_rows: list = []
        for enr in enrollments:
            counts = att_counts[enr.student_id]
            total = counts["total"]
            absent = counts["absent"]
            late = counts["late"]
            excused = counts["excused"]
            present = total - absent - late - excused
            pct = round(present / total * 100) if total else 0

            student_rows.append(
                {
                    "student": enr.student,
                    "total_sessions": total,
                    "present": present,
                    "absent": absent,
                    "late": late,
                    "excused": excused,
                    "attendance_pct": pct,
                }
            )

        return {
            "class_group": class_group,
            "school": school,
            "year": year,
            "student_rows": student_rows,
            "print_date": timezone.now().date(),
        }

    @staticmethod
    def get_behavior_report(school: School, year: str = settings.CURRENT_ACADEMIC_YEAR) -> dict:
        """بيانات تقرير السلوك العام"""
        from behavior.models import BehaviorInfraction

        infractions = (
            BehaviorInfraction.objects.filter(school=school)
            .select_related("student", "reported_by")
            .order_by("-date")
        )

        return {
            "school": school,
            "year": year,
            "infractions": infractions,
            "print_date": timezone.now().date(),
        }


# ══════════════════════════════════════════════════════════════════════
# AcademicReportsService — REQ-SH-003 (Client #001, MTG-007)
# ══════════════════════════════════════════════════════════════════════
#
# 4 report types requested by the Shahaniya School principal:
#   1. Quiz reports (by subject, at student or section scope)
#   2. Exam package results (AssessmentPackage comparison across classes)
#   3. Academic progress reports (by section over a time range)
#   4. Monthly behavior + academic report (FLAGSHIP — quiz avg + infractions)
#
# All 4 reports use existing schemas (Assessment, StudentAssessmentGrade,
# BehaviorInfraction) — NO new migrations. Walking-skeleton approach.
# ══════════════════════════════════════════════════════════════════════


class AcademicReportsService:
    """Services for REQ-SH-003 academic reports (4 report types)."""

    # ── Helpers ────────────────────────────────────────────────────

    @staticmethod
    def _month_bounds(year: int, month: int):
        """Return (start_date, end_date) for the given month."""
        from calendar import monthrange
        from datetime import date

        return (
            date(year, month, 1),
            date(year, month, monthrange(year, month)[1]),
        )

    @staticmethod
    def _safe_date(value: str | None):
        """Parse an ISO date string or return None."""
        if not value:
            return None
        try:
            from datetime import date

            parts = value.split("-")
            if len(parts) != 3:
                return None
            return date(int(parts[0]), int(parts[1]), int(parts[2]))
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _avg(values):
        clean = [float(v) for v in values if v is not None]
        if not clean:
            return None
        return round(sum(clean) / len(clean), 2)

    # ── Report 1: Quiz reports by subject ──────────────────────────

    @classmethod
    def get_quiz_reports(
        cls,
        school,
        *,
        subject_id: str | None = None,
        class_group_id: str | None = None,
        student_id: str | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
    ) -> dict:
        """
        Report 1 — Quiz results by subject at student or section scope.
        Uses StudentAssessmentGrade filtered by assessment_type='quiz'.
        """
        from assessments.models import StudentAssessmentGrade
        from core.models import ClassGroup
        from operations.models import Subject

        d_from = cls._safe_date(date_from)
        d_to = cls._safe_date(date_to)

        qs = StudentAssessmentGrade.objects.filter(
            school=school,
            assessment__assessment_type="quiz",
        ).select_related(
            "assessment__package__setup__subject",
            "assessment__package__setup__class_group",
            "student",
        )

        if subject_id:
            qs = qs.filter(assessment__package__setup__subject_id=subject_id)
        if class_group_id:
            qs = qs.filter(assessment__package__setup__class_group_id=class_group_id)
        if student_id:
            qs = qs.filter(student_id=student_id)
        if d_from:
            qs = qs.filter(assessment__date__gte=d_from)
        if d_to:
            qs = qs.filter(assessment__date__lte=d_to)

        qs = qs.order_by("assessment__date", "student__full_name")

        rows = []
        pct_values = []
        for grade in qs:
            pct = grade.grade_pct
            if pct is not None:
                pct_values.append(pct)
            rows.append(
                {
                    "student_name": grade.student.full_name,
                    "subject": grade.assessment.package.setup.subject.name_ar,
                    "class_group": str(grade.assessment.package.setup.class_group),
                    "title": grade.assessment.title,
                    "date": grade.assessment.date,
                    "max_grade": float(grade.assessment.max_grade),
                    "raw_grade": float(grade.grade) if grade.grade is not None else None,
                    "pct": pct,
                    "status": grade.status_label,
                }
            )

        # lookups for filter dropdowns (scoped to the school)
        subjects = list(Subject.objects.filter(school=school).order_by("name_ar"))
        classes = list(
            ClassGroup.objects.filter(school=school, is_active=True).order_by("grade", "section")
        )

        return {
            "school": school,
            "rows": rows,
            "total_rows": len(rows),
            "avg_pct": cls._avg(pct_values),
            "min_pct": min(pct_values) if pct_values else None,
            "max_pct": max(pct_values) if pct_values else None,
            "filters": {
                "subject_id": subject_id or "",
                "class_group_id": class_group_id or "",
                "student_id": student_id or "",
                "date_from": date_from or "",
                "date_to": date_to or "",
            },
            "subjects": subjects,
            "classes": classes,
            "print_date": timezone.now().date(),
        }

    # ── Report 2: Exam results (package comparison) ───────────────

    @classmethod
    def get_exam_results_reports(
        cls,
        school,
        *,
        package_type: str | None = None,
        semester: str | None = None,
        class_group_id: str | None = None,
    ) -> dict:
        """
        Report 2 — AssessmentPackage exam results across classes.
        Compares student totals of a package type (P1..P4, AW) between classes.
        """
        from assessments.models import AssessmentPackage, StudentSubjectResult
        from core.models import ClassGroup

        pkg_qs = AssessmentPackage.objects.filter(school=school).select_related(
            "setup__subject", "setup__class_group"
        )
        if package_type:
            pkg_qs = pkg_qs.filter(package_type=package_type)
        if semester:
            pkg_qs = pkg_qs.filter(semester=semester)
        if class_group_id:
            pkg_qs = pkg_qs.filter(setup__class_group_id=class_group_id)

        pkg_qs = pkg_qs.order_by("semester", "package_type", "setup__subject__name_ar")

        # Map package_type → field on StudentSubjectResult
        field_map = {
            "P1": "p1_score",
            "P2": "p2_score",
            "P3": "p3_score",
            "P4": "p4_score",
            "AW": "p_aw_score",
        }

        rows = []
        by_class: dict = {}
        for pkg in pkg_qs:
            field = field_map.get(pkg.package_type)
            if not field:
                continue
            results = StudentSubjectResult.objects.filter(
                school=school, setup=pkg.setup, semester=pkg.semester
            ).select_related("student")

            scores = []
            for r in results:
                val = getattr(r, field, None)
                if val is None:
                    continue
                scores.append(float(val))

            avg = cls._avg(scores)
            cg_key = str(pkg.setup.class_group)

            row = {
                "package": pkg.get_package_type_display(),
                "package_type": pkg.package_type,
                "semester": pkg.get_semester_display(),
                "class_group": cg_key,
                "subject": pkg.setup.subject.name_ar,
                "students_count": len(scores),
                "avg_score": avg,
                "min_score": min(scores) if scores else None,
                "max_score": max(scores) if scores else None,
                "max_grade": float(pkg.effective_max_grade),
            }
            rows.append(row)

            by_class.setdefault(cg_key, []).append(row)

        # Class-level summary (cross-class comparison)
        class_summary = []
        for cg, crows in by_class.items():
            avgs = [r["avg_score"] for r in crows if r["avg_score"] is not None]
            class_summary.append(
                {
                    "class_group": cg,
                    "packages_count": len(crows),
                    "overall_avg": cls._avg(avgs),
                }
            )

        classes = list(
            ClassGroup.objects.filter(school=school, is_active=True).order_by("grade", "section")
        )

        return {
            "school": school,
            "rows": rows,
            "class_summary": class_summary,
            "total_packages": len(rows),
            "filters": {
                "package_type": package_type or "",
                "semester": semester or "",
                "class_group_id": class_group_id or "",
            },
            "package_type_choices": list(AssessmentPackage.PACKAGE_TYPE),
            "semester_choices": list(AssessmentPackage.SEMESTER),
            "classes": classes,
            "print_date": timezone.now().date(),
        }

    # ── Report 3: Academic progress ────────────────────────────────

    @classmethod
    def get_academic_progress_reports(
        cls,
        school,
        *,
        class_group_id: str | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
    ) -> dict:
        """
        Report 3 — Academic progress for a section over a time range.
        Aggregates all Assessment grades per student within the window.
        """
        from assessments.models import StudentAssessmentGrade
        from core.models import ClassGroup, StudentEnrollment

        d_from = cls._safe_date(date_from)
        d_to = cls._safe_date(date_to)

        class_group = None
        if class_group_id:
            try:
                class_group = ClassGroup.objects.filter(school=school, id=class_group_id).first()
            except (ValueError, TypeError):
                class_group = None

        students = []
        if class_group:
            enrollments = (
                StudentEnrollment.objects.filter(class_group=class_group, is_active=True)
                .select_related("student")
                .order_by("student__full_name")
            )

            student_ids = [e.student_id for e in enrollments]

            grade_qs = StudentAssessmentGrade.objects.filter(
                school=school,
                student_id__in=student_ids,
                assessment__package__setup__class_group=class_group,
            )
            if d_from:
                grade_qs = grade_qs.filter(assessment__date__gte=d_from)
            if d_to:
                grade_qs = grade_qs.filter(assessment__date__lte=d_to)

            # Bucket grades per student
            from collections import defaultdict

            buckets: dict = defaultdict(list)
            for g in grade_qs:
                pct = g.grade_pct
                if pct is not None:
                    buckets[g.student_id].append(pct)

            for enr in enrollments:
                pcts = buckets.get(enr.student_id, [])
                avg = cls._avg(pcts)
                students.append(
                    {
                        "student_name": enr.student.full_name,
                        "assessments_count": len(pcts),
                        "avg_pct": avg,
                        "min_pct": min(pcts) if pcts else None,
                        "max_pct": max(pcts) if pcts else None,
                    }
                )

            # Rank by average
            students.sort(key=lambda s: s["avg_pct"] or 0, reverse=True)
            for i, s in enumerate(students, start=1):
                s["rank"] = i

        classes = list(
            ClassGroup.objects.filter(school=school, is_active=True).order_by("grade", "section")
        )

        overall_pcts = [s["avg_pct"] for s in students if s["avg_pct"] is not None]

        return {
            "school": school,
            "class_group": class_group,
            "students": students,
            "overall_avg": cls._avg(overall_pcts),
            "filters": {
                "class_group_id": class_group_id or "",
                "date_from": date_from or "",
                "date_to": date_to or "",
            },
            "classes": classes,
            "print_date": timezone.now().date(),
        }

    # ── Report 4: Monthly Behavior + Academic (FLAGSHIP) ──────────

    @classmethod
    def get_monthly_behavior_academic_report(
        cls,
        school,
        *,
        month: int,
        year: int,
        scope: str = "section",
        class_group_id: str | None = None,
        student_id: str | None = None,
    ) -> dict:
        """
        Report 4 — FLAGSHIP monthly report: quiz averages + behavior count.
        Level: 'student' (single student) or 'section' (all students in a class).
        Period: one calendar month.
        """
        from assessments.models import StudentAssessmentGrade
        from behavior.models import BehaviorInfraction
        from core.models import ClassGroup, CustomUser, StudentEnrollment

        start_date, end_date = cls._month_bounds(year, month)

        class_group = None
        if class_group_id:
            class_group = ClassGroup.objects.filter(school=school, id=class_group_id).first()

        # Resolve students list
        students: list = []
        if scope == "student" and student_id:
            student = CustomUser.objects.filter(
                id=student_id,
                memberships__school=school,
                memberships__is_active=True,
            ).first()
            if student:
                students = [student]
        elif class_group:
            enrollments = (
                StudentEnrollment.objects.filter(class_group=class_group, is_active=True)
                .select_related("student")
                .order_by("student__full_name")
            )
            students = [e.student for e in enrollments]

        # Batch query: quiz grades within month
        quiz_qs = StudentAssessmentGrade.objects.filter(
            school=school,
            assessment__assessment_type="quiz",
            assessment__date__gte=start_date,
            assessment__date__lte=end_date,
        )
        if students:
            quiz_qs = quiz_qs.filter(student__in=students)

        # Batch query: behavior infractions within month
        beh_qs = BehaviorInfraction.objects.filter(
            school=school,
            date__gte=start_date,
            date__lte=end_date,
        )
        if students:
            beh_qs = beh_qs.filter(student__in=students)

        # Bucket per student
        from collections import defaultdict

        quiz_buckets: dict = defaultdict(list)
        for g in quiz_qs:
            pct = g.grade_pct
            if pct is not None:
                quiz_buckets[g.student_id].append(pct)

        # نظام النقاط ملغى — نحسب فقط عدد المخالفات بوزن حسب الدرجة
        beh_counts: dict = defaultdict(int)
        beh_severity: dict = defaultdict(int)
        _SEVERITY_PER_LEVEL = {1: 2, 2: 5, 3: 10, 4: 25}
        for inf in beh_qs:
            beh_counts[inf.student_id] += 1
            beh_severity[inf.student_id] += _SEVERITY_PER_LEVEL.get(inf.level, 2)

        rows = []
        for st in students:
            pcts = quiz_buckets.get(st.id, [])
            quiz_avg = cls._avg(pcts)
            inf_count = beh_counts.get(st.id, 0)
            severity = beh_severity.get(st.id, 0)

            # Combined score: quiz% weighted 0.7 minus behavior penalty (severity-based)
            if quiz_avg is not None:
                penalty = min(severity, 30) * 0.3
                combined = round(quiz_avg * 0.7 + (100 - penalty) * 0.3, 2)
            else:
                combined = None

            rows.append(
                {
                    "student_name": st.full_name,
                    "student_id": str(st.id),
                    "quiz_count": len(pcts),
                    "quiz_avg": quiz_avg,
                    "behavior_count": inf_count,
                    "combined_score": combined,
                }
            )

        rows.sort(key=lambda r: (r["combined_score"] or 0), reverse=True)

        quiz_avg_all = cls._avg([r["quiz_avg"] for r in rows if r["quiz_avg"] is not None])
        total_infractions = sum(r["behavior_count"] for r in rows)

        classes = list(
            ClassGroup.objects.filter(school=school, is_active=True).order_by("grade", "section")
        )

        return {
            "school": school,
            "period": f"{year}-{month:02d}",
            "month": month,
            "year": year,
            "scope": scope,
            "class_group": class_group,
            "rows": rows,
            "students_count": len(rows),
            "quiz_avg": quiz_avg_all,
            "total_infractions": total_infractions,
            "filters": {
                "month": month,
                "year": year,
                "scope": scope,
                "class_group_id": class_group_id or "",
                "student_id": student_id or "",
            },
            "classes": classes,
            "print_date": timezone.now().date(),
        }


# ══════════════════════════════════════════════════════════════════════
# AcademicReportsExcel — lightweight Excel exports for REQ-SH-003
# ══════════════════════════════════════════════════════════════════════


class AcademicReportsExcel:
    """
    Lightweight Excel exporters for the 4 REQ-SH-003 reports.
    Re-uses ExcelService's header/style helpers.
    """

    @classmethod
    def _build(cls, sheet_title: str, title: str, columns: list, data_rows: list, school) -> bytes:
        """
        Generic builder: creates a workbook with the 4-row pro header,
        the given column widths, and writes the data rows.
        Returns HttpResponse via ExcelService.to_response.
        """
        wb, ws, styles = ExcelService._make_workbook(sheet_title)
        num_cols = len(columns)
        ExcelService._add_professional_header(
            ws,
            school.name if school else "SchoolOS",
            title,
            settings.CURRENT_ACADEMIC_YEAR,
            num_cols,
        )
        ExcelService._add_header_row(ws, styles, 4, columns)
        ws.freeze_panes = "A5"

        for idx, row in enumerate(data_rows, start=1):
            row_num = idx + 4
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_num, column=col_idx, value=value if value is not None else "—")
            ExcelService._style_data_row(ws, styles, row_num, num_cols, idx % 2 == 0)

        ExcelService._apply_protection(ws, num_cols)
        ExcelService._setup_print_a4_portrait(ws, num_cols, len(data_rows))
        return wb

    @classmethod
    def quiz_reports_excel(cls, data: dict, school) -> HttpResponse:
        columns = [
            ("م", 5),
            ("اسم الطالب", 28),
            ("المادة", 20),
            ("الفصل", 14),
            ("عنوان التقييم", 28),
            ("التاريخ", 13),
            ("الدرجة", 10),
            ("من", 8),
            ("النسبة %", 12),
        ]
        data_rows = []
        for i, r in enumerate(data.get("rows", []), start=1):
            data_rows.append(
                [
                    i,
                    r["student_name"],
                    r["subject"],
                    r["class_group"],
                    r["title"],
                    r["date"].strftime("%Y/%m/%d") if r.get("date") else "—",
                    r.get("raw_grade"),
                    r.get("max_grade"),
                    r.get("pct"),
                ]
            )
        wb = cls._build(
            "الاختبارات القصيرة",
            "تقارير الاختبارات القصيرة",
            columns,
            data_rows,
            school,
        )
        return ExcelService.to_response(wb, "quiz_reports.xlsx")

    @classmethod
    def exam_results_excel(cls, data: dict, school) -> HttpResponse:
        columns = [
            ("م", 5),
            ("الباقة", 22),
            ("الفصل الدراسي", 18),
            ("الشعبة", 14),
            ("المادة", 20),
            ("عدد الطلاب", 12),
            ("المتوسط", 12),
            ("الأدنى", 10),
            ("الأعلى", 10),
            ("من", 10),
        ]
        data_rows = []
        for i, r in enumerate(data.get("rows", []), start=1):
            data_rows.append(
                [
                    i,
                    r["package"],
                    r["semester"],
                    r["class_group"],
                    r["subject"],
                    r["students_count"],
                    r["avg_score"],
                    r["min_score"],
                    r["max_score"],
                    r["max_grade"],
                ]
            )
        wb = cls._build(
            "نتائج الاختبارات",
            "تقارير نتائج الاختبارات",
            columns,
            data_rows,
            school,
        )
        return ExcelService.to_response(wb, "exam_results.xlsx")

    @classmethod
    def academic_progress_excel(cls, data: dict, school) -> HttpResponse:
        columns = [
            ("الترتيب", 8),
            ("اسم الطالب", 32),
            ("عدد التقييمات", 14),
            ("المتوسط %", 12),
            ("الأدنى %", 12),
            ("الأعلى %", 12),
        ]
        data_rows = []
        for s in data.get("students", []):
            data_rows.append(
                [
                    s.get("rank"),
                    s["student_name"],
                    s["assessments_count"],
                    s["avg_pct"],
                    s["min_pct"],
                    s["max_pct"],
                ]
            )
        cg = data.get("class_group")
        title = f"تقارير التقدم الأكاديمي — {cg}" if cg else "تقارير التقدم الأكاديمي"
        wb = cls._build(
            "التقدم الأكاديمي",
            title,
            columns,
            data_rows,
            school,
        )
        return ExcelService.to_response(wb, "academic_progress.xlsx")

    @classmethod
    def monthly_behavior_academic_excel(cls, data: dict, school) -> HttpResponse:
        columns = [
            ("م", 5),
            ("اسم الطالب", 32),
            ("عدد الاختبارات", 13),
            ("متوسط الاختبارات %", 17),
            ("عدد المخالفات", 13),
            ("النقاط المخصومة", 15),
            ("التقييم المدمج", 14),
        ]
        data_rows = []
        for i, r in enumerate(data.get("rows", []), start=1):
            data_rows.append(
                [
                    i,
                    r["student_name"],
                    r["quiz_count"],
                    r["quiz_avg"],
                    r["behavior_count"],
                    r["behavior_points"],
                    r["combined_score"],
                ]
            )
        title = f"التقرير السلوكي والتعليمي الشهري — {data.get('period', '')}"
        wb = cls._build(
            "السلوك والتعليم الشهري",
            title,
            columns,
            data_rows,
            school,
        )
        return ExcelService.to_response(wb, f"monthly_ba_{data.get('period', 'report')}.xlsx")


# ══════════════════════════════════════════════════════════════════════
# ExcelService — إنشاء Excel باحترافية كاملة
# ══════════════════════════════════════════════════════════════════════


class ExcelService:
    """
    إنشاء ملفات Excel بهوية قطرية كاملة:
      - رأس 4 صفوف: وزارة + مدرسة + عنوان + أعمدة
      - شعار المدرسة في الرأس
      - حماية الورقة (قراءة فقط — كلمة السر من EXCEL_PROTECTION_PASSWORD)
      - فلاتر تلقائية + تجميد الرأس
      - RTL عربي + صفوف متبادلة + تلوين شرطي
    """

    MAROON = "8A1538"
    WHITE = "FFFFFF"
    ALT_BG = "FDF2F5"
    HEADER1 = "F5EEF1"  # وزارة
    HEADER2 = "FAFAFA"  # مدرسة
    HEADER3 = "FDF2F5"  # عنوان

    # ── بنية تحتية ────────────────────────────────────────────────────

    @classmethod
    def _make_workbook(cls, sheet_title: str) -> tuple:
        """Workbook جديد مع ستايل الهوية القطرية"""
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title
        ws.sheet_view.rightToLeft = True

        styles = {
            "header_font": Font(name="Arial", bold=True, color=cls.WHITE, size=11),
            "header_fill": PatternFill("solid", fgColor=cls.MAROON),
            "header_align": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "data_align": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "thin_border": Border(
                left=Side(style="thin", color="DDDDDD"),
                right=Side(style="thin", color="DDDDDD"),
                top=Side(style="thin", color="DDDDDD"),
                bottom=Side(style="thin", color="DDDDDD"),
            ),
            "alt_fill": PatternFill("solid", fgColor=cls.ALT_BG),
        }
        return wb, ws, styles

    @classmethod
    def _add_professional_header(
        cls, ws: object, school_name: str, report_title: str, year: str, num_cols: int
    ) -> None:
        """
        4 صفوف رأس احترافية:
          الصف 1 — وزارة التربية والتعليم العالي — دولة قطر  + تاريخ الطباعة
          الصف 2 — اسم المدرسة (كبير، كستنائي)
          الصف 3 — عنوان التقرير | السنة الدراسية
          الصف 4 — رأس الأعمدة (يملأه المستدعي عبر _add_header_row)
        """
        from pathlib import Path

        from django.conf import settings
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        col_letter = ws.cell(row=1, column=num_cols).column_letter
        today_str = timezone.now().strftime("%Y/%m/%d")

        bottom_border = Border(bottom=Side(style="medium", color=cls.MAROON))

        # ── صف 1: وزارة التربية + تاريخ الطباعة ─────────────────────
        ws.merge_cells(f"A1:{col_letter}1")
        c = ws["A1"]
        c.value = f"وزارة التربية والتعليم والتعليم العالي — دولة قطر          {today_str}"
        c.font = Font(name="Arial", size=9, color="555555")
        c.fill = PatternFill("solid", fgColor=cls.HEADER1)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bottom_border
        ws.row_dimensions[1].height = 22

        # ── صف 2: اسم المدرسة ────────────────────────────────────────
        ws.merge_cells(f"A2:{col_letter}2")
        c = ws["A2"]
        c.value = school_name
        c.font = Font(name="Arial", bold=True, size=16, color=cls.MAROON)
        c.fill = PatternFill("solid", fgColor=cls.HEADER2)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bottom_border
        ws.row_dimensions[2].height = 38

        # ── صف 3: عنوان التقرير + السنة الدراسية ─────────────────────
        ws.merge_cells(f"A3:{col_letter}3")
        c = ws["A3"]
        c.value = f"{report_title}   |   السنة الدراسية: {year}"
        c.font = Font(name="Arial", bold=True, size=12, color=cls.MAROON)
        c.fill = PatternFill("solid", fgColor=cls.HEADER3)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bottom_border
        ws.row_dimensions[3].height = 26

        # ── شعار المدرسة ──────────────────────────────────────────────
        try:
            from openpyxl.drawing.image import Image as XLImage

            logo_path = Path(settings.BASE_DIR) / "static" / "icons" / "badge-72.png"
            if not logo_path.exists():
                logo_path = Path(settings.BASE_DIR) / "static" / "icons" / "icon-192.png"
            if logo_path.exists():
                img = XLImage(str(logo_path))
                img.width = 54
                img.height = 54
                ws.add_image(img, "A1")
        except (ImportError, OSError, ValueError) as exc:
            logger.debug("Excel logo: %s", exc)

    @classmethod
    def _add_header_row(cls, ws: object, styles: dict, row_num: int, columns: list) -> None:
        """رأس الجدول: قائمة من (عنوان، عرض)"""
        for col_idx, (header, width) in enumerate(columns, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=header)
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
            cell.alignment = styles["header_align"]
            cell.border = styles["thin_border"]
            ws.column_dimensions[cell.column_letter].width = width
        ws.row_dimensions[row_num].height = 26

    @classmethod
    def _style_data_row(
        cls, ws: object, styles: dict, row_num: int, num_cols: int, is_alt: bool = False
    ) -> None:
        """تطبيق ستايل على صف بيانات"""
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = styles["thin_border"]
            cell.alignment = styles["data_align"]
            if is_alt:
                cell.fill = styles["alt_fill"]
        ws.row_dimensions[row_num].height = 20

    @classmethod
    def _apply_protection(cls, ws: object, num_cols: int) -> None:
        """
        حماية الورقة (قراءة فقط) مع السماح بالتصفية والفرز.
        كلمة السر تُقرأ من متغير البيئة EXCEL_PROTECTION_PASSWORD.
        """
        from django.conf import settings

        ws.protection.sheet = True
        ws.protection.password = getattr(settings, "EXCEL_PROTECTION_PASSWORD", "") or "protected"
        ws.protection.autoFilter = False  # يسمح باستخدام الفلاتر
        ws.protection.sort = False  # يسمح بالفرز

    @classmethod
    def _setup_print(
        cls,
        ws: object,
        num_cols: int,
        num_data_rows: int,
        paper: str = "a4",
        orientation: str = "portrait",
    ) -> None:
        """
        إعداد الطباعة — يدعم A4/A3 بوضع عمودي أو أفقي.

        Args:
            ws: ورقة العمل
            num_cols: عدد الأعمدة
            num_data_rows: عدد صفوف البيانات
            paper: "a4" (paperSize=9) أو "a3" (paperSize=8)
            orientation: "portrait" أو "landscape"
        """
        from openpyxl.worksheet.properties import PageSetupProperties

        PAPER_SIZES = {"a4": 9, "a3": 8}
        ws.page_setup.paperSize = PAPER_SIZES.get(paper.lower(), 9)
        ws.page_setup.orientation = orientation

        # ملاءمة العرض في صفحة واحدة — ارتفاع غير محدود
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

        # هوامش (بالبوصة) — A3 أفقي يحتاج هوامش أوسع قليلاً
        if paper.lower() == "a3" and orientation == "landscape":
            ws.page_margins.left = 0.5
            ws.page_margins.right = 0.5
            ws.page_margins.top = 0.5
            ws.page_margins.bottom = 0.5
            ws.page_margins.header = 0.3
            ws.page_margins.footer = 0.3
        else:
            ws.page_margins.left = 0.4
            ws.page_margins.right = 0.4
            ws.page_margins.top = 0.5
            ws.page_margins.bottom = 0.5
            ws.page_margins.header = 0.2
            ws.page_margins.footer = 0.2

        # تكرار أول 4 صفوف (الرأس) في كل صفحة مطبوعة
        ws.print_title_rows = "1:4"

        # منطقة الطباعة
        from openpyxl.utils import get_column_letter

        col_letter = get_column_letter(num_cols)
        last_row = num_data_rows + 4
        ws.print_area = f"A1:{col_letter}{last_row}"

        # هيدر وفوتر الطباعة (Excel format codes)
        ws.oddHeader.center.text = (
            '&"Arial,Bold"&9'
            "\u0645\u062f\u0631\u0633\u0629 \u0627\u0644\u0634\u062d\u0627\u0646\u064a\u0629"
        )
        ws.oddFooter.center.text = "&P / &N"
        ws.oddFooter.right.text = "&D"

    @classmethod
    def _setup_print_a4_portrait(cls, ws: object, num_cols: int, num_data_rows: int) -> None:
        """إعداد الطباعة على ورق A4 عمودي — غلاف توافقي للاستدعاءات القديمة."""
        cls._setup_print(ws, num_cols, num_data_rows, paper="a4", orientation="portrait")

    @classmethod
    def _setup_print_a3_landscape(cls, ws: object, num_cols: int, num_data_rows: int) -> None:
        """إعداد الطباعة على ورق A3 أفقي — غلاف مختصر."""
        cls._setup_print(ws, num_cols, num_data_rows, paper="a3", orientation="landscape")

    @classmethod
    def to_response(cls, wb: object, filename: str) -> HttpResponse:
        """تحويل Workbook إلى HttpResponse جاهز للتنزيل"""
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.read(),
            content_type=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

    # ── التقارير ──────────────────────────────────────────────────────

    @classmethod
    def class_results_excel(
        cls,
        class_group: ClassGroup,
        school: School,
        year: str = settings.CURRENT_ACADEMIC_YEAR,
        paper: str = "a4",
    ) -> HttpResponse:
        """
        Excel كشف نتائج الفصل:
        - رأس 4 صفوف احترافي + شعار
        - ترتيب حسب المتوسط
        - تلوين أحمر للدرجات < 50
        - تلوين الحالة (ناجح/راسب)
        - فلاتر تلقائية + تجميد الرأس + حماية الورقة
        """
        from openpyxl.styles import Font

        data = ReportDataService.get_class_results(class_group, school, year)
        subjects = data["subjects"]
        num_cols = 3 + len(subjects) + 3  # م + اسم + وطني + مواد + متوسط + حالة + ترتيب

        wb, ws, styles = cls._make_workbook("كشف النتائج")

        cls._add_professional_header(
            ws,
            school.name,
            f"كشف نتائج الفصل — {class_group}",
            year,
            num_cols,
        )

        is_a3 = paper.lower() == "a3"
        columns = (
            [
                ("م", 6 if is_a3 else 5),
                ("اسم الطالب", 38 if is_a3 else 28),
                ("الرقم الشخصي", 20 if is_a3 else 16),
            ]
            + [(s.name_ar[:18], 15 if is_a3 else 11) for s in subjects]
            + [
                ("المتوسط", 13 if is_a3 else 10),
                ("الحالة", 13 if is_a3 else 10),
                ("الترتيب", 10 if is_a3 else 8),
            ]
        )
        cls._add_header_row(ws, styles, 4, columns)

        # تجميد بعد صفوف الرأس الأربعة + فلاتر تلقائية
        ws.freeze_panes = "A5"
        col_letter = ws.cell(row=4, column=num_cols).column_letter
        ws.auto_filter.ref = f"A4:{col_letter}4"

        for row in data["student_rows"]:
            rank = row["rank"]
            row_num = rank + 4  # البيانات تبدأ من الصف 5
            is_alt = rank % 2 == 0
            st = row["student"]

            ws.cell(row=row_num, column=1, value=rank)
            ws.cell(row=row_num, column=2, value=st.full_name)
            ws.cell(row=row_num, column=3, value=st.national_id or "")

            for col_off, subj in enumerate(subjects, start=4):
                ann = row["grades"].get(subj.name_ar)
                grade = float(ann.annual_total) if ann and ann.annual_total else None
                cell = ws.cell(
                    row=row_num, column=col_off, value=grade if grade is not None else "—"
                )
                if grade is not None and grade < 50:
                    cell.font = Font(name="Arial", color="DC2626", bold=True)

            ws.cell(
                row=row_num,
                column=4 + len(subjects),
                value=row["avg"] if row["avg"] is not None else "—",
            )

            status_cell = ws.cell(row=row_num, column=5 + len(subjects), value=row["status"])
            if row["status"] == "ناجح":
                status_cell.font = Font(name="Arial", color="15803D", bold=True)
            elif row["status"] == "راسب":
                status_cell.font = Font(name="Arial", color="DC2626", bold=True)

            ws.cell(row=row_num, column=6 + len(subjects), value=rank)
            cls._style_data_row(ws, styles, row_num, num_cols, is_alt)

        cls._apply_protection(ws, num_cols)
        if is_a3:
            cls._setup_print_a3_landscape(ws, num_cols, len(data["student_rows"]))
        else:
            cls._setup_print_a4_portrait(ws, num_cols, len(data["student_rows"]))

        filename = f"نتائج_{class_group.get_grade_display()}_{class_group.section}_{year}.xlsx"
        return cls.to_response(wb, filename)

    @classmethod
    def attendance_excel(
        cls,
        class_group: ClassGroup,
        school: School,
        year: str = settings.CURRENT_ACADEMIC_YEAR,
        paper: str = "a4",
    ) -> HttpResponse:
        """
        Excel تقرير الغياب:
        - رأس 4 صفوف احترافي + شعار
        - أحمر للغياب > 10 حصة
        - أحمر/أخضر لنسبة الحضور
        - فلاتر تلقائية + تجميد الرأس + حماية الورقة
        """
        from openpyxl.styles import Font

        data = ReportDataService.get_attendance_report(class_group, school, year)
        num_cols = 8

        wb, ws, styles = cls._make_workbook("الحضور والغياب")

        cls._add_professional_header(
            ws,
            school.name,
            f"تقرير الحضور والغياب — {class_group}",
            year,
            num_cols,
        )

        is_a3 = paper.lower() == "a3"
        columns = [
            ("م", 6 if is_a3 else 5),
            ("اسم الطالب", 38 if is_a3 else 28),
            ("الرقم الشخصي", 20 if is_a3 else 16),
            ("إجمالي الحصص", 17 if is_a3 else 13),
            ("حاضر", 13 if is_a3 else 10),
            ("غائب", 13 if is_a3 else 10),
            ("متأخر", 13 if is_a3 else 10),
            ("نسبة الحضور %", 18 if is_a3 else 14),
        ]
        cls._add_header_row(ws, styles, 4, columns)

        ws.freeze_panes = "A5"
        ws.auto_filter.ref = "A4:H4"

        for idx, row in enumerate(data["student_rows"], start=1):
            row_num = idx + 4  # البيانات تبدأ من الصف 5
            st = row["student"]
            pct = row["attendance_pct"]

            ws.cell(row=row_num, column=1, value=idx)
            ws.cell(row=row_num, column=2, value=st.full_name)
            ws.cell(row=row_num, column=3, value=st.national_id or "")
            ws.cell(row=row_num, column=4, value=row["total_sessions"])
            ws.cell(row=row_num, column=5, value=row["present"])

            absent_cell = ws.cell(row=row_num, column=6, value=row["absent"])
            if row["absent"] > 10:
                absent_cell.font = Font(name="Arial", color="DC2626", bold=True)

            ws.cell(row=row_num, column=7, value=row["late"])

            pct_cell = ws.cell(row=row_num, column=8, value=pct)
            if pct < 80:
                pct_cell.font = Font(name="Arial", color="DC2626", bold=True)
            elif pct >= 95:
                pct_cell.font = Font(name="Arial", color="15803D", bold=True)

            cls._style_data_row(ws, styles, row_num, num_cols, idx % 2 == 0)

        cls._apply_protection(ws, num_cols)
        if is_a3:
            cls._setup_print_a3_landscape(ws, num_cols, len(data["student_rows"]))
        else:
            cls._setup_print_a4_portrait(ws, num_cols, len(data["student_rows"]))

        filename = f"غياب_{class_group.get_grade_display()}_{class_group.section}_{year}.xlsx"
        return cls.to_response(wb, filename)

    @classmethod
    def behavior_excel(
        cls,
        school: School,
        year: str = settings.CURRENT_ACADEMIC_YEAR,
        paper: str = "a4",
    ) -> HttpResponse:
        """
        Excel تقرير السلوك:
        - رأس 4 صفوف احترافي + شعار
        - تلوين درجة المخالفة (1→4 ألوان متصاعدة)
        - فلاتر تلقائية + تجميد الرأس + حماية الورقة
        """
        from openpyxl.styles import Font

        data = ReportDataService.get_behavior_report(school, year)

        LEVEL_LABELS = {
            1: "درجة 1 — بسيطة",
            2: "درجة 2 — متوسطة",
            3: "درجة 3 — جسيمة",
            4: "درجة 4 — شديدة",
        }
        LEVEL_COLORS = {1: "854D0E", 2: "C2410C", 3: "B91C1C", 4: "BE123C"}
        num_cols = 8

        wb, ws, styles = cls._make_workbook("مخالفات السلوك")

        cls._add_professional_header(
            ws,
            school.name,
            "تقرير المخالفات السلوكية",
            year,
            num_cols,
        )

        is_a3 = paper.lower() == "a3"
        columns = [
            ("م", 6 if is_a3 else 5),
            ("اسم الطالب", 38 if is_a3 else 28),
            ("الرقم الشخصي", 20 if is_a3 else 16),
            ("التاريخ", 17 if is_a3 else 13),
            ("الدرجة", 22 if is_a3 else 18),
            ("المُبلِّغ", 26 if is_a3 else 20),
            ("الوصف", 55 if is_a3 else 40),
        ]
        cls._add_header_row(ws, styles, 4, columns)

        ws.freeze_panes = "A5"
        ws.auto_filter.ref = "A4:G4"

        for idx, inf in enumerate(data["infractions"], start=1):
            row_num = idx + 4  # البيانات تبدأ من الصف 5

            ws.cell(row=row_num, column=1, value=idx)
            ws.cell(row=row_num, column=2, value=inf.student.full_name)
            ws.cell(row=row_num, column=3, value=inf.student.national_id or "")
            ws.cell(row=row_num, column=4, value=inf.date.strftime("%Y/%m/%d") if inf.date else "")

            level_cell = ws.cell(
                row=row_num,
                column=5,
                value=LEVEL_LABELS.get(inf.level, str(inf.level)),
            )
            level_cell.font = Font(
                name="Arial",
                color=LEVEL_COLORS.get(inf.level, "000000"),
                bold=True,
            )

            ws.cell(
                row=row_num, column=6, value=inf.reported_by.full_name if inf.reported_by else "—"
            )
            ws.cell(row=row_num, column=7, value=inf.description or "")

            cls._style_data_row(ws, styles, row_num, num_cols, idx % 2 == 0)

        cls._apply_protection(ws, num_cols)
        if is_a3:
            cls._setup_print_a3_landscape(ws, num_cols, len(data["infractions"]))
        else:
            cls._setup_print_a4_portrait(ws, num_cols, len(data["infractions"]))

        return cls.to_response(wb, f"سلوك_{school.name}_{year}.xlsx")
