"""
student_affairs/views.py — شؤون الطلاب
16 view — يتبع أنماط المشروع الموجودة بالضبط.
"""

import json
from datetime import timedelta

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Count, Exists, OuterRef, Q, Sum
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.utils import timezone
from django.views.decorators.http import require_POST

from assessments.models import AnnualSubjectResult
from behavior.models import BehaviorInfraction
from clinic.models import ClinicVisit, HealthRecord
from core.export_utils import (
    add_excel_footer,
    add_excel_header,
    excel_to_response,
    generate_export_filename,
    get_export_context,
    get_pdf_footer_html,
    get_pdf_header_html,
)
from core.models.academic import ClassGroup, ParentStudentLink, StudentEnrollment
from core.models.access import Membership, Role
from core.models.user import CustomUser, Profile
from core.pdf_utils import render_pdf
from core.permissions import STUDENT_AFFAIRS_MANAGE, STUDENT_DEACTIVATE, role_required
from library.models import BookBorrowing
from operations.models import AbsenceAlert, StudentAttendance

from .models import StudentActivity, StudentTransfer

# الأدوار المسموح لها بالوصول لشؤون الطلاب — مستوردة من core.permissions (MTG-2026-012)


# ═════════════════════════════════════════════════════════════════════
# لوحة شؤون الطلاب — الخطوة 3
# ═════════════════════════════════════════════════════════════════════


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def student_dashboard(request):
    """لوحة شؤون الطلاب — KPIs عبر Service Layer."""
    from .services import StudentService

    school = request.user.get_school()
    today = timezone.localdate()
    year = request.GET.get("year", settings.CURRENT_ACADEMIC_YEAR)

    # ✅ v5.4: StudentService.get_dashboard_context — جميع queries في service layer
    ctx = StudentService.get_dashboard_context(school, year, today=today)
    att = ctx["today_attendance"]

    return render(
        request,
        "student_affairs/dashboard.html",
        {
            "today": today,
            "year": year,
            "current_school": school,
            "total_students": ctx["total_students"],
            "absent_today": att["absent"] or 0,
            "late_today": att["late"] or 0,
            "present_today": att["present"] or 0,
            "behavior_month": ctx["behavior_month"]["total"] or 0,
            "clinic_today": ctx["clinic_today"],
            "pending_transfers": ctx["pending_transfers"],
            "grade_distribution": ctx["grade_distribution"],
            "linked_parents": ctx["parent_link_count"],
            "activities_year": ctx["activities_count"],
            "recent_infractions": ctx["recent_infractions"],
            "recent_transfers": ctx["recent_transfers"],
            "no_parent_count": ctx["no_parent_count"],
            "weekly_tardiness": ctx["weekly_tardiness"],
            "recent_activities": ctx["recent_activities"],
            # Req3 — لوحة اليوم
            "stage_map": ctx.get("stage_map", {}),
            "qatari_pct": ctx.get("qatari_pct", 0),
            "absent_pct": ctx.get("absent_pct", 0),
            "late_pct": ctx.get("late_pct", 0),
            "today_behavior_count": ctx.get("today_behavior_count", 0),
            "absent_list": ctx.get("absent_list", []),
            "late_list": ctx.get("late_list", []),
            "today_infraction_list": ctx.get("today_infraction_list", []),
        },
    )


# ═════════════════════════════════════════════════════════════════════
# سجل الطلاب — الخطوة 4
# ═════════════════════════════════════════════════════════════════════


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def student_list(request):
    """قائمة الطلاب مع بحث وفلتر حسب الصف والشعبة."""
    school = request.user.get_school()
    year = request.GET.get("year", settings.CURRENT_ACADEMIC_YEAR)

    # ── الاستعلام الأساسي: طلاب فعّالون في المدرسة ──
    students = (
        Membership.objects.filter(
            school=school,
            role__name="student",
            is_active=True,
        )
        .select_related("user", "user__profile")
        .order_by("user__full_name")
    )

    # ── إرفاق بيانات التسجيل (الصف + الشعبة) ──
    # نبني dict سريع: user_id → enrollment
    student_ids = list(students.values_list("user_id", flat=True))
    enrollments = {
        e["student_id"]: e
        for e in StudentEnrollment.objects.filter(
            student_id__in=student_ids,
            class_group__academic_year=year,
            is_active=True,
        )
        .select_related("class_group")
        .values(
            "student_id",
            "class_group__grade",
            "class_group__section",
            "class_group_id",
        )
    }

    # ── الفلاتر ──
    q = request.GET.get("q", "").strip()
    grade_filter = request.GET.get("grade", "")
    section_filter = request.GET.get("section", "")
    parent_status = request.GET.get("parent_status", "")

    if q:
        students = students.filter(
            Q(user__full_name__icontains=q) | Q(user__national_id__icontains=q)
        )

    if grade_filter:
        # ✅ subquery مباشر — لا تحميل IDs إلى Python
        grade_enrollment_exists = Exists(
            StudentEnrollment.objects.filter(
                class_group__school=school,
                class_group__academic_year=year,
                class_group__grade=grade_filter,
                is_active=True,
                student_id=OuterRef("user_id"),
            )
        )
        students = students.filter(grade_enrollment_exists)

    if section_filter:
        section_enrollment_exists = Exists(
            StudentEnrollment.objects.filter(
                class_group__school=school,
                class_group__academic_year=year,
                class_group__section=section_filter,
                is_active=True,
                student_id=OuterRef("user_id"),
            )
        )
        students = students.filter(section_enrollment_exists)

    if parent_status:
        # ✅ Exists subquery بدل Python set arithmetic — O(1) ذاكرة
        parent_link_exists = Exists(
            ParentStudentLink.objects.filter(school=school, student_id=OuterRef("user_id"))
        )
        if parent_status == "linked":
            students = students.annotate(_has_parent=parent_link_exists).filter(_has_parent=True)
        elif parent_status == "unlinked":
            students = students.annotate(_has_parent=parent_link_exists).filter(_has_parent=False)

    # ── بناء القائمة مع بيانات التسجيل ──
    student_rows = []
    for m in students[:200]:
        enr = enrollments.get(m.user_id, {})
        student_rows.append(
            {
                "id": m.user_id,
                "full_name": m.user.full_name,
                "national_id": m.user.national_id,
                "phone": m.user.phone,
                "email": m.user.email,
                "gender": getattr(m.user, "profile", None) and m.user.profile.gender or "",
                "grade": enr.get("class_group__grade", "—"),
                "section": enr.get("class_group__section", "—"),
            }
        )

    # ── خيارات الفلتر ──
    available_grades = (
        ClassGroup.objects.filter(school=school, academic_year=year, is_active=True)
        .values_list("grade", flat=True)
        .distinct()
        .order_by("grade")
    )
    available_sections = (
        ClassGroup.objects.filter(school=school, academic_year=year, is_active=True)
        .values_list("section", flat=True)
        .distinct()
        .order_by("section")
    )

    ctx = {
        "students": student_rows,
        "total": len(student_rows),
        "q": q,
        "grade_filter": grade_filter,
        "section_filter": section_filter,
        "parent_status": parent_status,
        "grades": available_grades,
        "sections": available_sections,
        "year": year,
    }

    # HTMX: إرجاع الجدول فقط
    if request.headers.get("HX-Request"):
        return render(request, "student_affairs/_student_table.html", ctx)

    return render(request, "student_affairs/student_list.html", ctx)


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def student_table_partial(request):
    """HTMX partial — يُعيد التوجيه لـ student_list مع نفس المعاملات."""
    return student_list(request)


# ═════════════════════════════════════════════════════════════════════
# تصدير Excel
# ═════════════════════════════════════════════════════════════════════


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def student_export_excel(request):
    """تصدير قائمة الطلاب إلى Excel — مع هيدر وفوتر احترافي."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    school = request.user.get_school()
    year = settings.CURRENT_ACADEMIC_YEAR
    q = request.GET.get("q", "").strip()
    grade_filter = request.GET.get("grade", "")
    section_filter = request.GET.get("section", "")

    ctx = get_export_context(request, "سجل الطلاب")

    # نفس فلترة student_list
    students = (
        Membership.objects.filter(
            school=school,
            role__name="student",
            is_active=True,
        )
        .select_related("user")
        .order_by("user__full_name")
    )

    if q:
        students = students.filter(
            Q(user__full_name__icontains=q) | Q(user__national_id__icontains=q)
        )

    enrollment_data = {}
    for enr in StudentEnrollment.objects.filter(
        class_group__school=school,
        class_group__academic_year=year,
        is_active=True,
    ).values("student_id", "class_group__grade", "class_group__section"):
        enrollment_data[enr["student_id"]] = enr

    if grade_filter:
        enrolled_ids = [
            sid
            for sid, data in enrollment_data.items()
            if data["class_group__grade"] == grade_filter
        ]
        students = students.filter(user_id__in=enrolled_ids)
    if section_filter:
        enrolled_ids = [
            sid
            for sid, data in enrollment_data.items()
            if data.get("class_group__section") == section_filter
        ]
        students = students.filter(user_id__in=enrolled_ids)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "سجل الطلاب"
    ws.sheet_view.rightToLeft = True

    # هيدر احترافي
    headers = ["#", "الاسم الكامل", "الرقم الشخصي", "الصف", "الشعبة", "الجوال", "البريد"]
    num_cols = len(headers)
    data_start = add_excel_header(ws, ctx, num_cols)

    # Header row
    header_fill = PatternFill(start_color="8A1538", end_color="8A1538", fill_type="solid")
    header_font = Font(name="Tajawal", bold=True, color="FFFFFF", size=11)
    cell_font = Font(name="Tajawal", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=data_start, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for i, m in enumerate(students, 1):
        enr = enrollment_data.get(m.user_id, {})
        row_data = [
            i,
            m.user.full_name,
            m.user.national_id,
            enr.get("class_group__grade", "—"),
            enr.get("class_group__section", "—"),
            m.user.phone or "—",
            m.user.email or "—",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=data_start + i, column=col, value=val)
            cell.font = cell_font
            cell.border = thin_border
            if i % 2 == 0:
                cell.fill = PatternFill(start_color="FDF2F5", end_color="FDF2F5", fill_type="solid")

    # Auto-width
    for col_idx in range(1, num_cols + 1):
        max_len = 0
        for row_idx in range(data_start, data_start + students.count() + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            max_len = max(max_len, len(str(cell.value or "")))
        col_letter = chr(64 + col_idx)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # فوتر احترافي
    last_data_row = data_start + students.count()
    add_excel_footer(ws, ctx, last_data_row, num_cols)

    filename = generate_export_filename("students", "list", "xlsx")
    return excel_to_response(wb, filename)


# ═════════════════════════════════════════════════════════════════════
# إضافة / تعديل / تعطيل — الخطوات 5 + 7
# ═════════════════════════════════════════════════════════════════════


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def student_add(request):
    """
    إضافة طالب جديد — مُفوَّض لـ StudentService.create_student().
    الـ View مسؤول فقط عن: قراءة الطلب + تحويل النموذج + عرض النتيجة.
    منطق إنشاء 4 السجلات (User + Profile + Membership + Enrollment) في Service Layer.
    """
    from .forms import StudentAddForm
    from .services import StudentService

    school = request.user.get_school()
    year = settings.CURRENT_ACADEMIC_YEAR

    if request.method == "POST":
        form = StudentAddForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data

            # ── تحديد الشعبة (المطلوب للـ Service) ──
            class_group = ClassGroup.objects.filter(
                school=school,
                grade=cd["grade"],
                section=cd["section"],
                academic_year=year,
                is_active=True,
            ).first()
            if not class_group:
                messages.error(
                    request,
                    f"لا توجد شعبة {cd['section']} في الصف {cd['grade']} للعام {year}.",
                )
                return render(
                    request,
                    "student_affairs/student_form.html",
                    {
                        "form": form,
                        "mode": "add",
                        "year": year,
                        "grades": ClassGroup.GRADES,
                        "school": school,
                    },
                )

            # ── تفويض الإنشاء للـ Service Layer ──
            try:
                user = StudentService.create_student(
                    school,
                    {
                        "national_id": cd["national_id"],
                        "full_name": cd["full_name"],
                        "phone": cd.get("phone", ""),
                        "email": cd.get("email", ""),
                        "gender": cd.get("gender", ""),
                        "birth_date": cd.get("birth_date"),
                        "nationality": cd.get("nationality", ""),
                        "class_group_id": class_group.pk,
                    },
                )
                messages.success(
                    request,
                    f"تم إضافة الطالب {user.full_name} في "
                    f"{class_group.grade}/{class_group.section} بنجاح.",
                )
                return redirect("student_affairs:student_profile", student_id=user.id)

            except ValueError as e:
                messages.error(request, str(e))
            except Exception as e:
                messages.error(request, f"خطأ غير متوقع أثناء إضافة الطالب: {e}")
    else:
        form = StudentAddForm()

    return render(
        request,
        "student_affairs/student_form.html",
        {
            "form": form,
            "mode": "add",
            "year": year,
            "grades": ClassGroup.GRADES,
            "school": school,
        },
    )


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def student_edit(request, student_id):
    """تعديل بيانات طالب موجود."""
    school = request.user.get_school()
    student = get_object_or_404(
        CustomUser,
        id=student_id,
        memberships__school=school,
        memberships__is_active=True,
    )
    year = settings.CURRENT_ACADEMIC_YEAR
    profile = getattr(student, "profile", None)
    enrollment = (
        StudentEnrollment.objects.filter(
            student=student,
            class_group__academic_year=year,
            is_active=True,
        )
        .select_related("class_group")
        .first()
    )

    from .forms import StudentEditForm

    if request.method == "POST":
        form = StudentEditForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            # ── تحديث CustomUser ──
            student.full_name = cd["full_name"]
            student.phone = cd.get("phone", "")
            student.email = cd.get("email", "")
            student.save()

            # ── تحديث Profile ──
            Profile.objects.update_or_create(
                user=student,
                defaults={
                    "birth_date": cd.get("birth_date"),
                    "notes": cd.get("notes", ""),
                },
            )

            # ── تحديث التسجيل (إذا تغيّر الصف/الشعبة) ──
            new_grade = cd["grade"]
            new_section = cd["section"]
            needs_reenroll = (
                not enrollment
                or enrollment.class_group.grade != new_grade
                or enrollment.class_group.section != new_section
            )
            if needs_reenroll:
                new_class = ClassGroup.objects.filter(
                    school=school,
                    grade=new_grade,
                    section=new_section,
                    academic_year=year,
                    is_active=True,
                ).first()
                if new_class:
                    if enrollment:
                        enrollment.is_active = False
                        enrollment.save()
                    StudentEnrollment.objects.create(
                        student=student,
                        class_group=new_class,
                        is_active=True,
                    )
                else:
                    messages.warning(request, f"لا توجد شعبة {new_section} في الصف {new_grade}.")

            messages.success(request, f"تم تحديث بيانات {student.full_name} بنجاح.")
            return redirect("student_affairs:student_profile", student_id=student.id)
    else:
        form = StudentEditForm(
            initial={
                "full_name": student.full_name,
                "phone": student.phone,
                "email": student.email,
                "grade": enrollment.class_group.grade if enrollment else "",
                "section": enrollment.class_group.section if enrollment else "",
                "birth_date": profile.birth_date if profile else None,
                "notes": profile.notes if profile else "",
            }
        )

    return render(
        request,
        "student_affairs/student_form.html",
        {
            "form": form,
            "mode": "edit",
            "student": student,
            "year": year,
            "grades": ClassGroup.GRADES,
            "school": school,
        },
    )


@login_required
@role_required(STUDENT_DEACTIVATE)
@require_POST
def student_deactivate(request, student_id):
    """
    تعطيل طالب — مُفوَّض لـ StudentService.deactivate_student().
    is_active=False فقط — لا حذف فيزيائي (PDPPL: حفظ السجل التاريخي).
    """
    from .services import StudentService

    school = request.user.get_school()
    student = get_object_or_404(
        CustomUser,
        id=student_id,
        memberships__school=school,
        memberships__is_active=True,
    )

    StudentService.deactivate_student(
        student=student,
        school=school,
        user=request.user,
    )

    messages.success(request, f"تم تعطيل الطالب {student.full_name}. البيانات محفوظة ولم تُحذف.")
    return redirect("student_affairs:student_list")


# ═════════════════════════════════════════════════════════════════════
# ملف الطالب الشامل — الخطوة 6
# ═════════════════════════════════════════════════════════════════════


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def student_profile(request, student_id):
    """ملف الطالب الشامل — يجمع بيانات من 7 تطبيقات."""
    school = request.user.get_school()
    student = get_object_or_404(
        CustomUser,
        id=student_id,
        memberships__school=school,
        memberships__is_active=True,
    )
    year = request.GET.get("year", settings.CURRENT_ACADEMIC_YEAR)

    # ── 1. البيانات الشخصية (core) ──
    profile = getattr(student, "profile", None)
    enrollment = (
        StudentEnrollment.objects.filter(
            student=student,
            class_group__academic_year=year,
            is_active=True,
        )
        .select_related("class_group")
        .first()
    )
    parent_links = ParentStudentLink.objects.filter(
        student=student,
        school=school,
    ).select_related("parent")

    # ── 2. الحضور (operations) ──
    attendance_qs = StudentAttendance.objects.filter(
        student=student,
        school=school,
        session__date__year=timezone.localdate().year,
    )
    attendance_summary = {
        "present": attendance_qs.filter(status="present").count(),
        "absent": attendance_qs.filter(status="absent").count(),
        "late": attendance_qs.filter(status="late").count(),
        "excused": attendance_qs.filter(status="excused").count(),
        "total": attendance_qs.count(),
    }
    if attendance_summary["total"] > 0:
        attendance_summary["pct"] = round(
            attendance_summary["present"] / attendance_summary["total"] * 100, 1
        )
    else:
        attendance_summary["pct"] = 0

    # ── 3. السلوك (behavior) ──
    infractions = (
        BehaviorInfraction.objects.filter(student=student, school=school)
        .select_related("violation_category")
        .order_by("-date")
    )
    # نظام النقاط ملغى — summary يعتمد على عدد المخالفات فقط
    behavior_summary = {
        "total": infractions.count(),
        "by_level": {lvl: infractions.filter(level=lvl).count() for lvl in range(1, 5)},
        "recent": infractions[:5],
    }

    # ── 4. العيادة (clinic) — ClinicVisit + HealthRecord مُستورَدان من أعلى الملف ──
    clinic_visits = ClinicVisit.objects.filter(student=student, school=school).order_by(
        "-visit_date"
    )[:5]
    health_record = HealthRecord.objects.filter(student=student).first()

    # ── 5. الدرجات (assessments) — AnnualSubjectResult مُستورَد من أعلى الملف ──
    grades = (
        AnnualSubjectResult.objects.filter(
            student=student,
            school=school,
            academic_year=year,
        )
        .select_related("setup__subject", "setup__class_group")
        .order_by("setup__subject__name_ar")
    )
    grades_summary = grades.aggregate(
        total_subjects=Count("id"),
        passed=Count("id", filter=Q(status="pass")),
        failed=Count("id", filter=Q(status="fail")),
    )

    # ── 6. المكتبة (library) — BookBorrowing مُستورَد من أعلى الملف ──
    borrowings = (
        BookBorrowing.objects.filter(user=student)
        .select_related("book")
        .order_by("-borrow_date")[:5]
    )

    # ── 7. الأنشطة (student_affairs) ──
    activities = StudentActivity.objects.filter(student=student, school=school).order_by("-date")[
        :10
    ]

    # ── الانتقالات ──
    transfers = StudentTransfer.objects.filter(student=student, school=school).order_by(
        "-created_at"
    )[:5]

    return render(
        request,
        "student_affairs/student_profile.html",
        {
            "student": student,
            "profile": profile,
            "enrollment": enrollment,
            "parent_links": parent_links,
            "attendance": attendance_summary,
            "behavior": behavior_summary,
            "clinic_visits": clinic_visits,
            "health_record": health_record,
            "grades": grades,
            "grades_summary": grades_summary,
            "borrowings": borrowings,
            "activities": activities,
            "transfers": transfers,
            "year": year,
        },
    )


# ═════════════════════════════════════════════════════════════════════
# الانتقالات — الخطوة 8
# ═════════════════════════════════════════════════════════════════════


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def transfer_list(request):
    """قائمة الانتقالات مع فلتر حسب الحالة والاتجاه."""
    school = request.user.get_school()
    transfers = (
        StudentTransfer.objects.filter(school=school)
        .select_related("student")
        .order_by("-created_at")
    )

    status_filter = request.GET.get("status", "")
    direction_filter = request.GET.get("direction", "")
    if status_filter:
        transfers = transfers.filter(status=status_filter)
    if direction_filter:
        transfers = transfers.filter(direction=direction_filter)

    return render(
        request,
        "student_affairs/transfer_list.html",
        {
            "transfers": transfers[:100],
            "status_filter": status_filter,
            "direction_filter": direction_filter,
            "status_choices": StudentTransfer.STATUS_CHOICES,
            "direction_choices": StudentTransfer.DIRECTION_CHOICES,
        },
    )


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def transfer_create(request):
    """تسجيل طلب انتقال جديد."""
    school = request.user.get_school()

    from .forms import TransferForm

    if request.method == "POST":
        form = TransferForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            student = get_object_or_404(
                CustomUser,
                id=cd["student_id"],
                memberships__school=school,
                memberships__is_active=True,
            )
            StudentTransfer.objects.create(
                school=school,
                student=student,
                direction=cd["direction"],
                other_school_name=cd["other_school_name"],
                from_grade=cd.get("from_grade", ""),
                to_grade=cd.get("to_grade", ""),
                transfer_date=cd["transfer_date"],
                reason=cd.get("reason", ""),
                academic_year=settings.CURRENT_ACADEMIC_YEAR,
                created_by=request.user,
                updated_by=request.user,
            )
            messages.success(request, f"تم تسجيل طلب انتقال {student.full_name} بنجاح.")
            return redirect("student_affairs:transfer_list")
    else:
        form = TransferForm()

    # قائمة الطلاب للاختيار
    students = (
        Membership.objects.filter(school=school, role__name="student", is_active=True)
        .select_related("user")
        .order_by("user__full_name")
    )
    return render(
        request,
        "student_affairs/transfer_form.html",
        {
            "form": form,
            "students": students,
        },
    )


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def transfer_detail(request, pk):
    """تفاصيل طلب انتقال."""
    school = request.user.get_school()
    transfer = get_object_or_404(StudentTransfer, pk=pk, school=school)
    return render(request, "student_affairs/transfer_detail.html", {"transfer": transfer})


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
@require_POST
def transfer_review(request, pk):
    """مراجعة طلب انتقال — موافقة / رفض / إتمام."""
    school = request.user.get_school()
    transfer = get_object_or_404(StudentTransfer, pk=pk, school=school)

    from .forms import TransferReviewForm

    form = TransferReviewForm(request.POST)
    if form.is_valid():
        action = form.cleaned_data["action"]
        notes = form.cleaned_data.get("notes", "")

        transfer.status = action
        transfer.notes = notes
        transfer.updated_by = request.user
        transfer.save()

        # إذا اكتمل الانتقال الصادر → تعطيل الطالب
        if action == "completed" and transfer.direction == "out":
            Membership.objects.filter(
                user=transfer.student,
                school=school,
                role__name="student",
                is_active=True,
            ).update(is_active=False)
            StudentEnrollment.objects.filter(
                student=transfer.student,
                is_active=True,
            ).update(is_active=False)

        status_label = dict(StudentTransfer.STATUS_CHOICES).get(action, action)
        messages.success(request, f"تم تحديث حالة الانتقال إلى: {status_label}")

    return redirect("student_affairs:transfer_detail", pk=pk)


# ═════════════════════════════════════════════════════════════════════
# الحضور والسلوك (ملخصات)
# ═════════════════════════════════════════════════════════════════════


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def attendance_overview(request):
    """إحصائيات الحضور والغياب — شاملة مع Trends."""
    school = request.user.get_school()
    today = timezone.localdate()
    year = request.GET.get("year", settings.CURRENT_ACADEMIC_YEAR)
    grade_filter = request.GET.get("grade", "")

    # ── إحصائيات اليوم ──
    today_qs = StudentAttendance.objects.filter(school=school, session__date=today)
    total_today = today_qs.count()
    present = today_qs.filter(status="present").count()
    absent = today_qs.filter(status="absent").count()
    late = today_qs.filter(status="late").count()
    excused = today_qs.filter(status="excused").count()
    pct = round(present * 100 / total_today) if total_today else 0

    summary = {
        "present": present,
        "absent": absent,
        "late": late,
        "excused": excused,
        "total": total_today,
        "pct": pct,
    }

    # ── أكثر 20 طالب غياباً (آخر 30 يوم) ──
    thirty_days_ago = today - timedelta(days=30)
    worst_students_qs = (
        StudentAttendance.objects.filter(
            school=school,
            status="absent",
            session__date__gte=thirty_days_ago,
        )
        .values("student__id", "student__full_name")
        .annotate(absence_count=Count("id"))
        .order_by("-absence_count")[:20]
    )

    # ── توزيع حسب الصف (الحضور اليوم) ──
    class_breakdown = (
        StudentAttendance.objects.filter(school=school, session__date=today)
        .values("session__class_group__grade")
        .annotate(
            total=Count("id"),
            present_count=Count("id", filter=Q(status="present")),
            absent_count=Count("id", filter=Q(status="absent")),
            late_count=Count("id", filter=Q(status="late")),
        )
        .order_by("session__class_group__grade")
    )

    # ── بيانات Chart (آخر 14 يوم) ──
    chart_labels = []
    chart_present = []
    chart_absent = []
    for i in range(13, -1, -1):
        d = today - timedelta(days=i)
        day_qs = StudentAttendance.objects.filter(school=school, session__date=d)
        total_d = day_qs.count()
        pres = day_qs.filter(status="present").count()
        abs_d = day_qs.filter(status="absent").count()
        chart_labels.append(d.strftime("%m/%d"))
        chart_present.append(round(pres * 100 / total_d) if total_d else 0)
        chart_absent.append(round(abs_d * 100 / total_d) if total_d else 0)

    # ── تنبيهات الغياب المتكرر — AbsenceAlert مُستورَد من أعلى الملف ──
    alerts = (
        AbsenceAlert.objects.filter(school=school, status="pending")
        .select_related("student")
        .order_by("-absence_count")[:10]
    )

    # ── الصفوف المتاحة للفلتر ──
    grades = ClassGroup.GRADES

    return render(
        request,
        "student_affairs/attendance_overview.html",
        {
            "summary": summary,
            "today": today,
            "year": year,
            "worst_students": worst_students_qs,
            "class_breakdown": class_breakdown,
            "chart_labels_json": json.dumps(chart_labels),
            "chart_present_json": json.dumps(chart_present),
            "chart_absent_json": json.dumps(chart_absent),
            "alerts": alerts,
            "grades": grades,
            "grade_filter": grade_filter,
        },
    )


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def attendance_export_excel(request):
    """تصدير إحصائيات الغياب — أكثر الطلاب غياباً (آخر 30 يوم) + حضور اليوم."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    school = request.user.get_school()
    today = timezone.localdate()
    thirty_ago = today - timedelta(days=30)

    ctx = get_export_context(request, "تقرير الحضور والغياب")

    # أكثر الطلاب غياباً — Count مُستورَد من أعلى الملف
    absence_data = (
        StudentAttendance.objects.filter(
            school=school,
            status="absent",
            session__date__gte=thirty_ago,
        )
        .values("student__full_name", "student__national_id")
        .annotate(absence_count=Count("id"))
        .order_by("-absence_count")
    )

    # سجل الحضور اليومي
    today_records = (
        StudentAttendance.objects.filter(school=school, session__date=today)
        .select_related("student", "session__class_group")
        .order_by("session__class_group__grade", "student__full_name")
    )

    # أنماط مشتركة
    header_fill = PatternFill(start_color="8A1538", end_color="8A1538", fill_type="solid")
    header_font = Font(name="Tajawal", bold=True, color="FFFFFF", size=11)
    cell_font = Font(name="Tajawal", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    wb = openpyxl.Workbook()

    # ── Sheet 1: الغياب المتكرر ──
    ws1 = wb.active
    ws1.title = "الغياب المتكرر"
    ws1.sheet_view.rightToLeft = True

    s1_headers = ["#", "اسم الطالب", "الرقم الشخصي", "أيام الغياب (30 يوم)"]
    s1_num_cols = len(s1_headers)
    s1_data_start = add_excel_header(ws1, ctx, s1_num_cols)

    for col, h in enumerate(s1_headers, 1):
        cell = ws1.cell(row=s1_data_start, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    absence_count_total = 0
    for i, rec in enumerate(absence_data, 1):
        absence_count_total = i
        row_data = [i, rec["student__full_name"], rec["student__national_id"], rec["absence_count"]]
        for col, val in enumerate(row_data, 1):
            cell = ws1.cell(row=s1_data_start + i, column=col, value=val)
            cell.font = cell_font
            cell.border = thin_border
            if i % 2 == 0:
                cell.fill = PatternFill(start_color="FDF2F5", end_color="FDF2F5", fill_type="solid")

    for col_idx in range(1, 5):  # 4 columns
        max_len = 0
        for row_idx in range(s1_data_start, s1_data_start + len(absence_data) + 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            max_len = max(max_len, len(str(cell.value or "")))
        ws1.column_dimensions[chr(64 + col_idx)].width = min(max_len + 4, 40)

    add_excel_footer(ws1, ctx, s1_data_start + absence_count_total, s1_num_cols)

    # ── Sheet 2: سجل حضور اليوم ──
    ws2 = wb.create_sheet("حضور اليوم")
    ws2.sheet_view.rightToLeft = True

    s2_headers = ["#", "اسم الطالب", "الصف", "الشعبة", "الحالة"]
    s2_num_cols = len(s2_headers)
    s2_data_start = add_excel_header(ws2, ctx, s2_num_cols)

    for col, h in enumerate(s2_headers, 1):
        cell = ws2.cell(row=s2_data_start, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    status_map = {"present": "حاضر", "absent": "غائب", "late": "متأخر", "excused": "معذور"}
    today_count = 0
    for i, rec in enumerate(today_records, 1):
        today_count = i
        row_data = [
            i,
            rec.student.full_name,
            rec.session.class_group.grade,
            rec.session.class_group.section,
            status_map.get(rec.status, rec.status),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=s2_data_start + i, column=col, value=val)
            cell.font = cell_font
            cell.border = thin_border
            if rec.status == "absent":
                cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            elif rec.status == "late":
                cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            elif i % 2 == 0:
                cell.fill = PatternFill(start_color="FDF2F5", end_color="FDF2F5", fill_type="solid")

    for col_idx in range(1, 6):  # 5 columns
        max_len = 0
        for row_idx in range(s2_data_start, s2_data_start + today_count + 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            max_len = max(max_len, len(str(cell.value or "")))
        ws2.column_dimensions[chr(64 + col_idx)].width = min(max_len + 4, 40)

    add_excel_footer(ws2, ctx, s2_data_start + today_count, s2_num_cols)

    filename = generate_export_filename("attendance", "stats", "xlsx")
    return excel_to_response(wb, filename)


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def behavior_overview(request):
    """ملخص سلوك الطلاب — إحصائيات شاملة."""
    school = request.user.get_school()
    today = timezone.localdate()
    grade_filter = request.GET.get("grade", "")

    # ── مخالفات السنة الحالية ──
    year_infractions = BehaviorInfraction.objects.filter(
        school=school,
        date__year=today.year,
    )
    total_infractions = year_infractions.count()
    unresolved = year_infractions.filter(is_resolved=False).count()

    # ── عدد الطلاب المخالفين (فريد) ──
    students_with_infractions = year_infractions.values("student").distinct().count()

    # ── إجمالي الطلاب المسجلين ──
    total_students = Membership.objects.filter(
        school=school,
        role__name="student",
        is_active=True,
    ).count()
    infraction_pct = (
        round(students_with_infractions * 100 / total_students) if total_students else 0
    )

    # ── توزيع حسب درجة المخالفة (1-4) — نظام النقاط ملغى ──
    degree_distribution = (
        year_infractions.values("violation_category__degree")
        .annotate(count=Count("id"))
        .order_by("violation_category__degree")
    )
    degree_map = {}
    for row in degree_distribution:
        deg = row["violation_category__degree"]
        if deg:
            degree_map[deg] = {"count": row["count"]}

    # ── أكثر 15 طالب مخالفات — مُرتَّبة حسب العدد ──
    worst_students = (
        year_infractions.values("student__id", "student__full_name")
        .annotate(infraction_count=Count("id"))
        .order_by("-infraction_count")[:15]
    )

    # ── اتجاه المخالفات الشهري (آخر 6 أشهر) ──
    chart_labels = []
    chart_data = []
    for i in range(5, -1, -1):
        month_start = (today.replace(day=1) - timedelta(days=30 * i)).replace(day=1)
        if i > 0:
            next_month = (month_start + timedelta(days=32)).replace(day=1)
        else:
            next_month = today + timedelta(days=1)
        count = BehaviorInfraction.objects.filter(
            school=school,
            date__gte=month_start,
            date__lt=next_month,
        ).count()
        chart_labels.append(month_start.strftime("%b"))
        chart_data.append(count)

    # ── مخالفات اليوم ──
    today_infractions = year_infractions.filter(date=today).count()

    # ── الصفوف المتاحة للفلتر ──
    grades = ClassGroup.GRADES

    return render(
        request,
        "student_affairs/behavior_overview.html",
        {
            "today": today,
            "total_infractions": total_infractions,
            "unresolved": unresolved,
            "students_with_infractions": students_with_infractions,
            "total_students": total_students,
            "infraction_pct": infraction_pct,
            "today_infractions": today_infractions,
            "degree_map": degree_map,
            "worst_students": worst_students,
            "chart_labels_json": json.dumps(chart_labels),
            "chart_data_json": json.dumps(chart_data),
            "grades": grades,
            "grade_filter": grade_filter,
        },
    )


# ═════════════════════════════════════════════════════════════════════
# الأنشطة والإنجازات — الخطوة 9
# ═════════════════════════════════════════════════════════════════════


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def activity_list(request):
    """قائمة الأنشطة والإنجازات مع فلتر."""
    school = request.user.get_school()
    year = request.GET.get("year", settings.CURRENT_ACADEMIC_YEAR)
    activities = (
        StudentActivity.objects.filter(school=school, academic_year=year)
        .select_related("student")
        .order_by("-date")
    )

    type_filter = request.GET.get("type", "")
    scope_filter = request.GET.get("scope", "")
    if type_filter:
        activities = activities.filter(activity_type=type_filter)
    if scope_filter:
        activities = activities.filter(scope=scope_filter)

    return render(
        request,
        "student_affairs/activity_list.html",
        {
            "activities": activities[:200],
            "type_filter": type_filter,
            "scope_filter": scope_filter,
            "type_choices": StudentActivity.TYPE_CHOICES,
            "scope_choices": StudentActivity.SCOPE_CHOICES,
            "year": year,
        },
    )


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def activity_add(request):
    """تسجيل نشاط أو إنجاز جديد."""
    school = request.user.get_school()

    from .forms import ActivityForm

    if request.method == "POST":
        form = ActivityForm(request.POST, request.FILES)
        if form.is_valid():
            cd = form.cleaned_data
            student = get_object_or_404(
                CustomUser,
                id=cd["student_id"],
                memberships__school=school,
                memberships__is_active=True,
            )
            StudentActivity.objects.create(
                school=school,
                student=student,
                activity_type=cd["activity_type"],
                title=cd["title"],
                description=cd.get("description", ""),
                scope=cd["scope"],
                date=cd["date"],
                academic_year=settings.CURRENT_ACADEMIC_YEAR,
                recorded_by=request.user,
                attachment=cd.get("attachment"),
            )
            messages.success(
                request, f"تم تسجيل النشاط «{cd['title']}» للطالب {student.full_name}."
            )
            return redirect("student_affairs:activity_list")
    else:
        form = ActivityForm()

    students = (
        Membership.objects.filter(school=school, role__name="student", is_active=True)
        .select_related("user")
        .order_by("user__full_name")
    )
    return render(
        request,
        "student_affairs/activity_form.html",
        {
            "form": form,
            "students": students,
            "mode": "add",
        },
    )


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def activity_edit(request, pk):
    """تعديل نشاط."""
    school = request.user.get_school()
    activity = get_object_or_404(StudentActivity, pk=pk, school=school)

    from .forms import ActivityForm

    if request.method == "POST":
        form = ActivityForm(request.POST, request.FILES)
        if form.is_valid():
            cd = form.cleaned_data
            activity.activity_type = cd["activity_type"]
            activity.title = cd["title"]
            activity.description = cd.get("description", "")
            activity.scope = cd["scope"]
            activity.date = cd["date"]
            if cd.get("attachment"):
                activity.attachment = cd["attachment"]
            activity.save()
            messages.success(request, f"تم تحديث النشاط «{activity.title}».")
            return redirect("student_affairs:activity_list")
    else:
        form = ActivityForm(
            initial={
                "student_id": activity.student_id,
                "activity_type": activity.activity_type,
                "title": activity.title,
                "description": activity.description,
                "scope": activity.scope,
                "date": activity.date,
            }
        )

    students = (
        Membership.objects.filter(school=school, role__name="student", is_active=True)
        .select_related("user")
        .order_by("user__full_name")
    )
    return render(
        request,
        "student_affairs/activity_form.html",
        {
            "form": form,
            "students": students,
            "mode": "edit",
            "activity": activity,
        },
    )


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
@require_POST
def activity_delete(request, pk):
    """حذف نشاط."""
    school = request.user.get_school()
    activity = get_object_or_404(StudentActivity, pk=pk, school=school)
    title = activity.title
    activity.delete()
    messages.success(request, f"تم حذف النشاط «{title}».")
    return redirect("student_affairs:activity_list")


# ═════════════════════════════════════════════════════════════════════
# إضافة ولي أمر جديد + ربطه بطالب
# ═════════════════════════════════════════════════════════════════════


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def parent_add(request):
    """إضافة ولي أمر جديد وربطه بطالب — ينشئ 3 سجلات ذرّياً (User + Membership + ParentStudentLink)."""

    from .forms import ParentAddForm

    school = request.user.get_school()
    year = settings.CURRENT_ACADEMIC_YEAR

    # ✅ subquery مباشر — بدون تحميل student_ids إلى Python memory
    students = (
        CustomUser.objects.filter(
            enrollments__class_group__school=school,
            enrollments__class_group__academic_year=year,
            enrollments__is_active=True,
        )
        .distinct()
        .order_by("full_name")
    )

    if request.method == "POST":
        form = ParentAddForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            student = get_object_or_404(CustomUser, id=cd["student_id"])

            existing = CustomUser.objects.filter(national_id=cd["national_id"]).first()

            try:
                with transaction.atomic():
                    if existing:
                        parent = existing
                        # تأكد من وجود Membership كـ parent
                        parent_role, _ = Role.objects.get_or_create(school=school, name="parent")
                        Membership.objects.get_or_create(
                            user=parent,
                            school=school,
                            role=parent_role,
                            defaults={"is_active": True},
                        )
                    else:
                        # إنشاء حساب جديد
                        parent = CustomUser.objects.create_user(
                            national_id=cd["national_id"],
                            password=cd["national_id"],
                            full_name=cd["full_name"],
                            phone=cd.get("phone", ""),
                            email=cd.get("email", ""),
                            must_change_password=True,
                        )
                        Profile.objects.get_or_create(user=parent)
                        parent_role, _ = Role.objects.get_or_create(school=school, name="parent")
                        Membership.objects.create(
                            user=parent,
                            school=school,
                            role=parent_role,
                            is_active=True,
                        )

                    # ربط ولي الأمر بالطالب
                    link, created = ParentStudentLink.objects.get_or_create(
                        school=school,
                        parent=parent,
                        student=student,
                        defaults={
                            "relationship": cd["relationship"],
                            "can_view_grades": True,
                            "can_view_attendance": True,
                        },
                    )

                if created:
                    messages.success(
                        request,
                        f"تم إضافة {parent.full_name} وربطه بالطالب {student.full_name} بنجاح.",
                    )
                else:
                    messages.warning(
                        request,
                        f"الربط بين {parent.full_name} والطالب {student.full_name} موجود مسبقاً.",
                    )
                return redirect("manage_parent_links")

            except Exception as e:
                messages.error(request, f"خطأ أثناء إضافة ولي الأمر: {e}")
    else:
        form = ParentAddForm()

    return render(
        request,
        "student_affairs/parent_form.html",
        {
            "form": form,
            "students": students,
        },
    )


# ═════════════════════════════════════════════════════════════════════
# تصدير ملف الطالب PDF
# ═════════════════════════════════════════════════════════════════════


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def student_profile_pdf(request, student_id):
    """ملف الطالب الشامل — PDF للطباعة (A4)."""
    school = request.user.get_school()
    student = get_object_or_404(
        CustomUser,
        id=student_id,
        memberships__school=school,
        memberships__is_active=True,
    )
    year = request.GET.get("year", settings.CURRENT_ACADEMIC_YEAR)

    # enrollment
    enrollment = (
        StudentEnrollment.objects.filter(
            student=student,
            class_group__academic_year=year,
            is_active=True,
        )
        .select_related("class_group")
        .first()
    )

    # حضور آخر 30 يوم
    today = timezone.localdate()
    thirty_ago = today - timedelta(days=30)
    attendance = (
        StudentAttendance.objects.filter(
            school=school,
            student=student,
            session__date__gte=thirty_ago,
        )
        .select_related("session__subject")
        .order_by("-session__date")
    )

    att_summary = {
        "total": attendance.count(),
        "present": attendance.filter(status="present").count(),
        "absent": attendance.filter(status="absent").count(),
        "late": attendance.filter(status="late").count(),
    }

    # سلوك — نظام النقاط ملغى
    infractions = (
        BehaviorInfraction.objects.filter(school=school, student=student)
        .select_related("violation_category")
        .order_by("-date")[:20]
    )

    # درجات — AnnualSubjectResult مُستورَد من أعلى الملف
    grades = (
        AnnualSubjectResult.objects.filter(
            student=student,
            school=school,
            academic_year=year,
        )
        .select_related("setup__subject")
        .order_by("setup__subject__name_ar")
    )

    # أنشطة
    activities = StudentActivity.objects.filter(school=school, student=student).order_by("-date")[
        :10
    ]

    # أولياء الأمور
    parent_links = ParentStudentLink.objects.filter(
        school=school,
        student=student,
    ).select_related("parent")

    ctx = get_export_context(request, "ملف الطالب الشامل")

    html_string = render_to_string(
        "student_affairs/student_profile_pdf.html",
        {
            "student": student,
            "school": school,
            "enrollment": enrollment,
            "attendance": attendance[:15],
            "att_summary": att_summary,
            "infractions": infractions,
            "grades": grades,
            "activities": activities,
            "parent_links": parent_links,
            "today": today,
            "year": year,
            **ctx,
        },
    )

    return render_pdf(html_string, f"student_{student.full_name}.pdf")


# ═════════════════════════════════════════════════════════════════════
# التأخر الصباحي
# ═════════════════════════════════════════════════════════════════════


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def tardiness_list(request):
    """قائمة الطلاب المتأخرين — مفلترة حسب التاريخ والصف."""
    school = request.user.get_school()

    date_str = request.GET.get("date")
    if date_str:
        try:
            from datetime import date as date_type

            selected_date = date_type.fromisoformat(date_str)
        except ValueError:
            selected_date = timezone.localdate()
    else:
        selected_date = timezone.localdate()

    grade_filter = request.GET.get("grade", "")
    section_filter = request.GET.get("section", "")

    late_qs = StudentAttendance.objects.filter(
        school=school,
        status="late",
        session__date=selected_date,
    )
    if grade_filter:
        late_qs = late_qs.filter(session__class_group__grade=grade_filter)
    if section_filter:
        late_qs = late_qs.filter(session__class_group__section=section_filter)

    late_records = late_qs.select_related(
        "student", "session__class_group", "session__subject"
    ).order_by("session__class_group__grade", "session__class_group__section", "student__full_name")

    total_late = late_records.count()

    # KPIs إضافية
    total_students_today = (
        StudentAttendance.objects.filter(
            school=school,
            session__date=selected_date,
        )
        .values("student")
        .distinct()
        .count()
    )
    late_pct = round(total_late * 100 / total_students_today) if total_students_today else 0

    # توزيع التأخر حسب الصف
    class_breakdown = (
        late_qs.values("session__class_group__grade", "session__class_group__section")
        .annotate(count=Count("id"))
        .order_by("session__class_group__grade", "session__class_group__section")
    )

    # التأخر هذا الأسبوع
    week_start = selected_date - timedelta(days=selected_date.weekday())
    weekly_late = StudentAttendance.objects.filter(
        school=school,
        status="late",
        session__date__gte=week_start,
        session__date__lte=selected_date,
    ).count()

    grades = ClassGroup.GRADES

    return render(
        request,
        "student_affairs/tardiness_list.html",
        {
            "late_records": late_records,
            "selected_date": selected_date,
            "total_late": total_late,
            "total_students_today": total_students_today,
            "late_pct": late_pct,
            "class_breakdown": class_breakdown,
            "weekly_late": weekly_late,
            "grades": grades,
            "grade_filter": grade_filter,
            "section_filter": section_filter,
        },
    )


# ═════════════════════════════════════════════════════════════════════
# تصديرات Excel إضافية — سلوك + تأخر + أنشطة
# ═════════════════════════════════════════════════════════════════════


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def behavior_export_excel(request):
    """تصدير إحصائيات السلوك — المخالفات + أكثر الطلاب."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    school = request.user.get_school()
    ctx = get_export_context(request, "تقرير السلوك الطلابي")

    # بيانات
    infractions = (
        BehaviorInfraction.objects.filter(school=school)
        .values("student__full_name", "student__national_id")
        .annotate(count=Count("id"))
        .order_by("-count")
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "السلوك"
    ws.sheet_view.rightToLeft = True

    headers = ["#", "اسم الطالب", "الرقم الشخصي", "عدد المخالفات"]
    num_cols = len(headers)
    data_start = add_excel_header(ws, ctx, num_cols)

    header_fill = PatternFill(start_color="8A1538", end_color="8A1538", fill_type="solid")
    header_font = Font(name="Tajawal", bold=True, color="FFFFFF", size=11)
    cell_font = Font(name="Tajawal", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=data_start, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row_count = 0
    for i, rec in enumerate(infractions, 1):
        row_data = [
            i,
            rec["student__full_name"],
            rec["student__national_id"],
            rec["count"],
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=data_start + i, column=col, value=val)
            cell.font = cell_font
            cell.border = thin_border
            if i % 2 == 0:
                cell.fill = PatternFill(start_color="FDF2F5", end_color="FDF2F5", fill_type="solid")
        row_count = i

    for col_idx in range(1, num_cols + 1):
        max_len = 0
        for r in range(data_start, data_start + row_count + 1):
            cell = ws.cell(row=r, column=col_idx)
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[chr(64 + col_idx)].width = min(max_len + 4, 40)

    add_excel_footer(ws, ctx, data_start + row_count, num_cols)
    return excel_to_response(wb, generate_export_filename("behavior", "stats", "xlsx"))


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def tardiness_export_excel(request):
    """تصدير قائمة المتأخرين ليوم محدد."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    school = request.user.get_school()

    date_str = request.GET.get("date")
    if date_str:
        try:
            from datetime import date as date_type

            selected_date = date_type.fromisoformat(date_str)
        except ValueError:
            selected_date = timezone.localdate()
    else:
        selected_date = timezone.localdate()

    ctx = get_export_context(
        request, f"تقرير التأخر الصباحي — {selected_date.strftime('%d/%m/%Y')}"
    )

    late_records = (
        StudentAttendance.objects.filter(
            school=school,
            status="late",
            session__date=selected_date,
        )
        .select_related("student", "session__class_group", "session__subject")
        .order_by("session__class_group__grade", "student__full_name")
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "التأخر"
    ws.sheet_view.rightToLeft = True

    headers = ["#", "اسم الطالب", "الصف", "الشعبة", "المادة", "وقت الحصة"]
    num_cols = len(headers)
    data_start = add_excel_header(ws, ctx, num_cols)

    header_fill = PatternFill(start_color="8A1538", end_color="8A1538", fill_type="solid")
    header_font = Font(name="Tajawal", bold=True, color="FFFFFF", size=11)
    cell_font = Font(name="Tajawal", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=data_start, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row_count = 0
    for i, rec in enumerate(late_records, 1):
        subject_name = rec.session.subject.name_ar if rec.session.subject else "—"
        time_str = rec.session.start_time.strftime("%H:%M") if rec.session.start_time else "—"
        row_data = [
            i,
            rec.student.full_name,
            rec.session.class_group.grade,
            rec.session.class_group.section,
            subject_name,
            time_str,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=data_start + i, column=col, value=val)
            cell.font = cell_font
            cell.border = thin_border
            if i % 2 == 0:
                cell.fill = PatternFill(start_color="FDF2F5", end_color="FDF2F5", fill_type="solid")
        row_count = i

    for col_idx in range(1, num_cols + 1):
        max_len = 0
        for r in range(data_start, data_start + row_count + 1):
            cell = ws.cell(row=r, column=col_idx)
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[chr(64 + col_idx)].width = min(max_len + 4, 40)

    add_excel_footer(ws, ctx, data_start + row_count, num_cols)
    return excel_to_response(wb, generate_export_filename("tardiness", "daily", "xlsx"))


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def activities_export_excel(request):
    """تصدير قائمة الأنشطة والإنجازات."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    school = request.user.get_school()
    ctx = get_export_context(request, "تقرير الأنشطة والإنجازات")

    type_map = dict(StudentActivity.TYPE_CHOICES)
    scope_map = dict(StudentActivity.SCOPE_CHOICES)

    activities = (
        StudentActivity.objects.filter(school=school).select_related("student").order_by("-date")
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الأنشطة"
    ws.sheet_view.rightToLeft = True

    headers = ["#", "اسم الطالب", "النشاط", "النوع", "النطاق", "التاريخ"]
    num_cols = len(headers)
    data_start = add_excel_header(ws, ctx, num_cols)

    header_fill = PatternFill(start_color="8A1538", end_color="8A1538", fill_type="solid")
    header_font = Font(name="Tajawal", bold=True, color="FFFFFF", size=11)
    cell_font = Font(name="Tajawal", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=data_start, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row_count = 0
    for i, act in enumerate(activities, 1):
        row_data = [
            i,
            act.student.full_name,
            act.title,
            type_map.get(act.activity_type, act.activity_type),
            scope_map.get(act.scope, act.scope),
            act.date.strftime("%d/%m/%Y") if act.date else "—",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=data_start + i, column=col, value=val)
            cell.font = cell_font
            cell.border = thin_border
            if i % 2 == 0:
                cell.fill = PatternFill(start_color="FDF2F5", end_color="FDF2F5", fill_type="solid")
        row_count = i

    for col_idx in range(1, num_cols + 1):
        max_len = 0
        for r in range(data_start, data_start + row_count + 1):
            cell = ws.cell(row=r, column=col_idx)
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[chr(64 + col_idx)].width = min(max_len + 4, 40)

    add_excel_footer(ws, ctx, data_start + row_count, num_cols)
    return excel_to_response(wb, generate_export_filename("activities", "list", "xlsx"))


# ═════════════════════════════════════════════════════════════════════
# تصديرات PDF — حضور + سلوك + تأخر
# ═════════════════════════════════════════════════════════════════════


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def attendance_overview_pdf(request):
    """تصدير إحصائيات الحضور والغياب — PDF."""
    school = request.user.get_school()
    today = timezone.localdate()

    # ── إحصائيات اليوم ──
    today_qs = StudentAttendance.objects.filter(school=school, session__date=today)
    total_today = today_qs.count()
    present = today_qs.filter(status="present").count()
    absent = today_qs.filter(status="absent").count()
    late = today_qs.filter(status="late").count()
    excused = today_qs.filter(status="excused").count()
    pct = round(present * 100 / total_today) if total_today else 0

    summary = {
        "present": present,
        "absent": absent,
        "late": late,
        "excused": excused,
        "total": total_today,
        "pct": pct,
    }

    # ── أكثر 20 طالب غياباً (آخر 30 يوم) ──
    thirty_days_ago = today - timedelta(days=30)
    worst_students = (
        StudentAttendance.objects.filter(
            school=school,
            status="absent",
            session__date__gte=thirty_days_ago,
        )
        .values("student__id", "student__full_name")
        .annotate(absence_count=Count("id"))
        .order_by("-absence_count")[:20]
    )

    ctx = get_export_context(request, "تقرير الحضور والغياب")
    pdf_header = get_pdf_header_html(ctx)
    pdf_footer = get_pdf_footer_html(ctx)

    html = render_to_string(
        "student_affairs/attendance_overview_pdf.html",
        {
            "summary": summary,
            "today": today,
            "worst_students": worst_students,
            "pdf_header": pdf_header,
            "pdf_footer": pdf_footer,
            **ctx,
        },
    )

    filename = generate_export_filename("attendance", "overview", "pdf")
    return render_pdf(html, filename, paper_size="A4")


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def behavior_overview_pdf(request):
    """تصدير ملخص السلوك — PDF."""
    school = request.user.get_school()
    today = timezone.localdate()

    # ── مخالفات السنة الحالية ──
    year_infractions = BehaviorInfraction.objects.filter(
        school=school,
        date__year=today.year,
    )
    total_infractions = year_infractions.count()
    unresolved = year_infractions.filter(is_resolved=False).count()

    # ── نسبة المخالفين ──
    students_with_infractions = year_infractions.values("student").distinct().count()
    total_students = Membership.objects.filter(
        school=school,
        role__name="student",
        is_active=True,
    ).count()
    infraction_pct = (
        round(students_with_infractions * 100 / total_students) if total_students else 0
    )

    # ── توزيع حسب الدرجة — نظام النقاط ملغى ──
    degree_distribution = (
        year_infractions.values("violation_category__degree")
        .annotate(count=Count("id"))
        .order_by("violation_category__degree")
    )
    degree_rows = []
    for row in degree_distribution:
        deg = row["violation_category__degree"]
        if deg:
            degree_rows.append({"degree": deg, "count": row["count"]})

    # ── أكثر 15 طالب مخالفات — مرتبة حسب العدد ──
    worst_students = (
        year_infractions.values("student__id", "student__full_name")
        .annotate(infraction_count=Count("id"))
        .order_by("-infraction_count")[:15]
    )

    ctx = get_export_context(request, "تقرير السلوك")
    pdf_header = get_pdf_header_html(ctx)
    pdf_footer = get_pdf_footer_html(ctx)

    html = render_to_string(
        "student_affairs/behavior_overview_pdf.html",
        {
            "today": today,
            "total_infractions": total_infractions,
            "unresolved": unresolved,
            "infraction_pct": infraction_pct,
            "degree_rows": degree_rows,
            "worst_students": worst_students,
            "pdf_header": pdf_header,
            "pdf_footer": pdf_footer,
            **ctx,
        },
    )

    filename = generate_export_filename("behavior", "overview", "pdf")
    return render_pdf(html, filename, paper_size="A4")


@login_required
@role_required(STUDENT_AFFAIRS_MANAGE)
def tardiness_pdf(request):
    """تصدير قائمة المتأخرين — PDF."""
    school = request.user.get_school()

    date_str = request.GET.get("date")
    if date_str:
        try:
            from datetime import date as date_type

            selected_date = date_type.fromisoformat(date_str)
        except ValueError:
            selected_date = timezone.localdate()
    else:
        selected_date = timezone.localdate()

    late_records = (
        StudentAttendance.objects.filter(
            school=school,
            status="late",
            session__date=selected_date,
        )
        .select_related("student", "session__class_group", "session__subject")
        .order_by(
            "session__class_group__grade", "session__class_group__section", "student__full_name"
        )
    )

    total_late = late_records.count()

    # نسبة التأخر
    total_students_today = (
        StudentAttendance.objects.filter(
            school=school,
            session__date=selected_date,
        )
        .values("student")
        .distinct()
        .count()
    )
    late_pct = round(total_late * 100 / total_students_today) if total_students_today else 0

    # التأخر هذا الأسبوع
    week_start = selected_date - timedelta(days=selected_date.weekday())
    weekly_late = StudentAttendance.objects.filter(
        school=school,
        status="late",
        session__date__gte=week_start,
        session__date__lte=selected_date,
    ).count()

    ctx = get_export_context(request, "تقرير التأخر الصباحي")
    pdf_header = get_pdf_header_html(ctx)
    pdf_footer = get_pdf_footer_html(ctx)

    html = render_to_string(
        "student_affairs/tardiness_pdf.html",
        {
            "selected_date": selected_date,
            "late_records": late_records,
            "total_late": total_late,
            "late_pct": late_pct,
            "weekly_late": weekly_late,
            "pdf_header": pdf_header,
            "pdf_footer": pdf_footer,
            **ctx,
        },
    )

    filename = generate_export_filename("tardiness", "list", "pdf")
    return render_pdf(html, filename, paper_size="A4")
